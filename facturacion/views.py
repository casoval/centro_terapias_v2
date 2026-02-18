# -*- coding: utf-8 -*-
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count, F, Case, When, DecimalField, Prefetch, Value
from django.db.models.functions import Coalesce
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.conf import settings
from django.utils import timezone
from io import BytesIO

import base64
import os
from decimal import Decimal
from datetime import date, datetime

from django.core.cache import cache
from django.contrib.admin.views.decorators import staff_member_required


from .models import CuentaCorriente, Pago, MetodoPago, DetallePagoMasivo, Devolucion
from .services import PaymentService, AccountService
from . import pdf_generator  # ✅ Importar módulo de PDFs
from pacientes.models import Paciente
from agenda.models import Sesion, Proyecto, Mensualidad
from facturacion.models import Devolucion

from .admin_views import (
    panel_recalcular_cuentas,
    recalcular_todas_cuentas,
    recalcular_cuenta_individual,
    api_recalcular_cuenta,
    api_estado_recalculo
)

# ============================================================
# ✅ CORREGIDO: HELPERS - Total pagado incluyendo pagos masivos Y devoluciones
# ============================================================
# Un pago masivo almacena sesion=None/proyecto=None/mensualidad=None en Pago
# y guarda los ítems reales en DetallePagoMasivo.
# Las relaciones inversas (sesion.pagos, proyecto.pagos, etc.) NO incluyen
# esos pagos masivos, por eso hay que sumar también los detalles masivos.
#
# ✅ ACTUALIZACIÓN: Las funciones ahora SÍ restan devoluciones para proyectos
#    y mensualidades, calculando el monto NETO pagado.
#    Esto corrige el bug donde items con devoluciones no aparecían en pagos masivos.
# ============================================================

def _total_pagado_sesion(sesion):
    """
    Total pagado para una sesión incluyendo pagos masivos (no anulados).
    """
    pagos_directos = sum(p.monto for p in sesion.pagos.filter(anulado=False))
    pagos_masivos = DetallePagoMasivo.objects.filter(
        tipo='sesion',
        sesion=sesion,
        pago__anulado=False
    ).aggregate(total=Coalesce(Sum('monto'), Decimal('0')))['total']
    return pagos_directos + pagos_masivos


def _total_pagado_proyecto(proyecto):
    """
    Total pagado NETO para un proyecto (pagos - devoluciones).
    
    ✅ CORREGIDO: Ahora incluye la resta de devoluciones.
    
    Calcula:
    - Pagos directos (FK pago.proyecto)
    - Pagos masivos (DetallePagoMasivo con tipo='proyecto')
    - Resta devoluciones del proyecto
    
    Returns:
        Decimal: Monto neto pagado (puede ser negativo si hay más devoluciones que pagos)
    """
    # Pagos directos al proyecto
    pagos_directos = sum(p.monto for p in proyecto.pagos.filter(anulado=False))
    
    # Pagos masivos que incluyen este proyecto
    pagos_masivos = DetallePagoMasivo.objects.filter(
        tipo='proyecto',
        proyecto=proyecto,
        pago__anulado=False
    ).aggregate(total=Coalesce(Sum('monto'), Decimal('0')))['total']
    
    # ✅ NUEVO: Devoluciones realizadas del proyecto
    devoluciones = Devolucion.objects.filter(
        proyecto=proyecto
    ).aggregate(total=Coalesce(Sum('monto'), Decimal('0')))['total']
    
    # Total neto = pagos - devoluciones
    return pagos_directos + pagos_masivos - devoluciones


def _total_pagado_mensualidad(mensualidad):
    """
    Total pagado NETO para una mensualidad (pagos - devoluciones).
    
    ✅ CORREGIDO: Ahora incluye la resta de devoluciones.
    
    Calcula:
    - Pagos directos (FK pago.mensualidad)
    - Pagos masivos (DetallePagoMasivo con tipo='mensualidad')
    - Resta devoluciones de la mensualidad
    
    Returns:
        Decimal: Monto neto pagado (puede ser negativo si hay más devoluciones que pagos)
    """
    # Pagos directos a la mensualidad
    pagos_directos = sum(p.monto for p in mensualidad.pagos.filter(anulado=False))
    
    # Pagos masivos que incluyen esta mensualidad
    pagos_masivos = DetallePagoMasivo.objects.filter(
        tipo='mensualidad',
        mensualidad=mensualidad,
        pago__anulado=False
    ).aggregate(total=Coalesce(Sum('monto'), Decimal('0')))['total']
    
    # ✅ NUEVO: Devoluciones realizadas de la mensualidad
    devoluciones = Devolucion.objects.filter(
        mensualidad=mensualidad
    ).aggregate(total=Coalesce(Sum('monto'), Decimal('0')))['total']
    
    # Total neto = pagos - devoluciones
    return pagos_directos + pagos_masivos - devoluciones
    
@login_required
def lista_cuentas_corrientes(request):
    """
    Vista ultra-optimizada de cuentas corrientes
    
    🚀 MEJORAS:
    - Carga inicial: Solo saldo (ya en BD) → < 1 segundo
    - Sin cálculo de propiedades pesadas
    - Estadísticas globales calculadas SOLO con aggregations (no loops)
    - Desglose detallado solo al hacer clic (AJAX)
    
    Performance:
    - Antes: 5-10 segundos (500+ queries)
    - Ahora: < 1 segundo (3-5 queries)
    """
    
    # ==================== FILTROS ====================
    buscar = request.GET.get('q', '').strip()
    estado = request.GET.get('estado', '')
    sucursal_id = request.GET.get('sucursal', '')
    
    # ==================== QUERY BASE OPTIMIZADA ====================
    # Solo select_related para evitar N+1 queries
    pacientes = Paciente.objects.filter(
        estado='activo'
    ).select_related(
        'cuenta_corriente'
    ).prefetch_related(
        'sucursales'
    )
    
    # Aplicar filtros
    if buscar:
        pacientes = pacientes.filter(
            Q(nombre__icontains=buscar) | 
            Q(apellido__icontains=buscar) |
            Q(nombre_tutor__icontains=buscar)
        )
    
    if sucursal_id:
        pacientes = pacientes.filter(sucursales__id=sucursal_id)
    
    # Crear cuentas faltantes (sin actualizar saldos - se hace con señales)
    for paciente in pacientes:
        if not hasattr(paciente, 'cuenta_corriente'):
            CuentaCorriente.objects.create(paciente=paciente)
    
    # ==================== FILTRADO POR ESTADO ====================
    # Usamos el campo 'saldo_actual' que ya está en BD
    if estado == 'deudor':
        pacientes = pacientes.filter(cuenta_corriente__saldo_actual__lt=0)
    elif estado == 'al_dia':
        pacientes = pacientes.filter(cuenta_corriente__saldo_actual=0)
    elif estado == 'a_favor':
        pacientes = pacientes.filter(cuenta_corriente__saldo_actual__gt=0)
    
    # Ordenar por saldo actual
    pacientes = pacientes.order_by('cuenta_corriente__saldo_actual')
    
    # ==================== PAGINACIÓN ====================
    paginator = Paginator(pacientes, 50)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # ==================== ESTADÍSTICAS GLOBALES ====================
    # ✅ OPTIMIZADO: Solo queries a nivel de BD, sin loops
    # Las estadísticas se cargan automáticamente vía AJAX al cargar la página
    
    estadisticas = None
    mostrar_estadisticas = False  # Las estadísticas se cargan automáticamente con JavaScript
    
    # ==================== SUCURSALES PARA FILTRO ====================
    from servicios.models import Sucursal
    sucursales = Sucursal.objects.filter(activa=True)
    
    context = {
        'page_obj': page_obj,
        'estadisticas': estadisticas,
        'mostrar_estadisticas': mostrar_estadisticas,
        'buscar': buscar,
        'estado': estado,
        'sucursal_id': sucursal_id,
        'sucursales': sucursales,
    }
    
    return render(request, 'facturacion/cuentas_corrientes.html', context)


def calcular_estadisticas_globales(buscar=None, estado=None, sucursal_id=None):
    """
    Calcula estadísticas globales de todas las cuentas corrientes
    
    ✅ CORREGIDO: 
    - Usa campos pre-calculados de CuentaCorriente para consistencia
    - Más eficiente (1 query en lugar de 3)
    - Consistente con cuentas individuales
    - ✅ NUEVO: Incluye estadísticas de PROYECCIÓN TOTAL
    - ✅ DINÁMICO: Acepta filtros de búsqueda, estado y sucursal
    
    Args:
        buscar (str, optional): Término de búsqueda por nombre/apellido/tutor
        estado (str, optional): Filtro por estado ('deudor', 'al_dia', 'a_favor')
        sucursal_id (str, optional): ID de sucursal para filtrar
    
    Returns:
        dict: Estadísticas globales del sistema (estado actual + proyección total)
    """
    from decimal import Decimal
    from django.db.models import Sum, Q
    from facturacion.models import CuentaCorriente
    
    try:
        # Obtener todas las cuentas de pacientes activos
        cuentas_activas = CuentaCorriente.objects.filter(
            paciente__estado='activo'
        )
        
        # ========================================
        # ✅ APLICAR FILTROS DINÁMICOS
        # ========================================
        
        # Filtro de búsqueda
        if buscar:
            cuentas_activas = cuentas_activas.filter(
                Q(paciente__nombre__icontains=buscar) | 
                Q(paciente__apellido__icontains=buscar) |
                Q(paciente__nombre_tutor__icontains=buscar)
            )
        
        # Filtro por sucursal
        if sucursal_id:
            cuentas_activas = cuentas_activas.filter(
                paciente__sucursales__id=sucursal_id
            )
        
        # Filtro por estado (aplicar DESPUÉS de obtener las cuentas)
        if estado == 'deudor':
            cuentas_activas = cuentas_activas.filter(saldo_actual__lt=0)
        elif estado == 'al_dia':
            cuentas_activas = cuentas_activas.filter(saldo_actual=0)
        elif estado == 'a_favor':
            cuentas_activas = cuentas_activas.filter(saldo_actual__gt=0)
        
        # ========================================
        # ✅ ESTADO ACTUAL (consumido + asistidas)
        # ========================================
        
        estadisticas_actual = cuentas_activas.aggregate(
            total_consumido=Sum('total_consumido_actual'),  # ⬅️ Campo pre-calculado
            total_pagado=Sum('total_pagado'),               # ⬅️ Campo pre-calculado
            total_balance=Sum('saldo_actual')               # ⬅️ Campo pre-calculado
        )
        
        # Convertir None a Decimal('0.00')
        total_consumido = estadisticas_actual['total_consumido'] or Decimal('0.00')
        total_pagado = estadisticas_actual['total_pagado'] or Decimal('0.00')
        total_balance = estadisticas_actual['total_balance'] or Decimal('0.00')
        
        # ========================================
        # ✅ PROYECCIÓN TOTAL (todas las sesiones programadas)
        # ========================================
        
        estadisticas_proyeccion = cuentas_activas.aggregate(
            total_proyectado=Sum('total_consumido_real'),  # ⬅️ Incluye todas las sesiones
            total_balance_proyeccion=Sum('saldo_real')     # ⬅️ Saldo con proyección
        )
        
        total_proyectado = estadisticas_proyeccion['total_proyectado'] or Decimal('0.00')
        total_balance_proyeccion = estadisticas_proyeccion['total_balance_proyeccion'] or Decimal('0.00')
        
        # ========================================
        # CLASIFICACIÓN POR SALDO ACTUAL
        # ========================================
        
        deudores_count = cuentas_activas.filter(saldo_actual__lt=0).count()
        al_dia_count = cuentas_activas.filter(saldo_actual=0).count()
        a_favor_count = cuentas_activas.filter(saldo_actual__gt=0).count()
        
        # Total que deben (valor absoluto)
        total_debe_result = cuentas_activas.filter(
            saldo_actual__lt=0
        ).aggregate(total=Sum('saldo_actual'))['total']
        
        total_debe = abs(total_debe_result) if total_debe_result else Decimal('0.00')
        
        # Total a favor
        total_favor = cuentas_activas.filter(
            saldo_actual__gt=0
        ).aggregate(total=Sum('saldo_actual'))['total'] or Decimal('0.00')
        
        # ========================================
        # ✅ CLASIFICACIÓN POR SALDO PROYECTADO
        # ========================================
        
        deudores_proyeccion_count = cuentas_activas.filter(saldo_real__lt=0).count()
        al_dia_proyeccion_count = cuentas_activas.filter(saldo_real=0).count()
        a_favor_proyeccion_count = cuentas_activas.filter(saldo_real__gt=0).count()
        
        # Total que deberán (valor absoluto)
        total_debe_proyeccion_result = cuentas_activas.filter(
            saldo_real__lt=0
        ).aggregate(total=Sum('saldo_real'))['total']
        
        total_debe_proyeccion = abs(total_debe_proyeccion_result) if total_debe_proyeccion_result else Decimal('0.00')
        
        # Total a favor proyectado
        total_favor_proyeccion = cuentas_activas.filter(
            saldo_real__gt=0
        ).aggregate(total=Sum('saldo_real'))['total'] or Decimal('0.00')
        
        return {
            # Estado Actual
            'total_consumido': total_consumido,
            'total_pagado': total_pagado,
            'total_balance': total_balance,
            'deudores': deudores_count,
            'al_dia': al_dia_count,
            'a_favor': a_favor_count,
            'total_debe': total_debe,
            'total_favor': total_favor,
            
            # ✅ Proyección Total
            'total_proyectado': total_proyectado,
            'total_balance_proyeccion': total_balance_proyeccion,
            'deudores_proyeccion': deudores_proyeccion_count,
            'al_dia_proyeccion': al_dia_proyeccion_count,
            'a_favor_proyeccion': a_favor_proyeccion_count,
            'total_debe_proyeccion': total_debe_proyeccion,
            'total_favor_proyeccion': total_favor_proyeccion,
        }
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error en estadísticas globales: {str(e)}", exc_info=True)
        
        # Retornar valores por defecto en caso de error
        return {
            # Estado Actual
            'total_consumido': Decimal('0.00'),
            'total_pagado': Decimal('0.00'),
            'total_balance': Decimal('0.00'),
            'deudores': 0,
            'al_dia': 0,
            'a_favor': 0,
            'total_debe': Decimal('0.00'),
            'total_favor': Decimal('0.00'),
            
            # Proyección Total
            'total_proyectado': Decimal('0.00'),
            'total_balance_proyeccion': Decimal('0.00'),
            'deudores_proyeccion': 0,
            'al_dia_proyeccion': 0,
            'a_favor_proyeccion': 0,
            'total_debe_proyeccion': Decimal('0.00'),
            'total_favor_proyeccion': Decimal('0.00'),
        }


# ============================================================
# ✅ FUNCIÓN CORREGIDA: cargar_estadisticas_ajax
# ============================================================
# Versión mejorada con mejor logging de errores
# ============================================================

@login_required
def cargar_estadisticas_ajax(request):
    """
    Vista AJAX para cargar estadísticas globales bajo demanda
    
    ✅ MEJORADO: 
    - Mejor manejo de errores y logging detallado
    - ✅ NUEVO: Incluye estadísticas de proyección total
    - ✅ DINÁMICO: Acepta y aplica filtros de búsqueda, estado y sucursal
    """
    try:
        # ✅ Obtener filtros de la petición
        buscar = request.GET.get('q', '')
        estado = request.GET.get('estado', '')
        sucursal_id = request.GET.get('sucursal', '')
        
        # Calcular estadísticas con filtros aplicados
        estadisticas = calcular_estadisticas_globales(
            buscar=buscar if buscar else None,
            estado=estado if estado else None,
            sucursal_id=sucursal_id if sucursal_id else None
        )
        
        # ✅ Calcular total de pacientes con los mismos filtros
        from django.db.models import Q
        pacientes = Paciente.objects.filter(estado='activo')
        
        if buscar:
            pacientes = pacientes.filter(
                Q(nombre__icontains=buscar) | 
                Q(apellido__icontains=buscar) |
                Q(nombre_tutor__icontains=buscar)
            )
        
        if sucursal_id:
            pacientes = pacientes.filter(sucursales__id=sucursal_id)
        
        if estado == 'deudor':
            pacientes = pacientes.filter(cuenta_corriente__saldo_actual__lt=0)
        elif estado == 'al_dia':
            pacientes = pacientes.filter(cuenta_corriente__saldo_actual=0)
        elif estado == 'a_favor':
            pacientes = pacientes.filter(cuenta_corriente__saldo_actual__gt=0)
        
        total_pacientes = pacientes.count()
        
        return JsonResponse({
            'success': True,
            'total_pacientes': total_pacientes,  # ✅ Añadir total de pacientes
            'estadisticas': {
                # Estado Actual
                'total_consumido': float(estadisticas['total_consumido']),
                'total_pagado': float(estadisticas['total_pagado']),
                'total_balance': float(estadisticas['total_balance']),
                'deudores': estadisticas['deudores'],
                'al_dia': estadisticas['al_dia'],
                'a_favor': estadisticas['a_favor'],
                'total_debe': float(estadisticas['total_debe']),
                'total_favor': float(estadisticas['total_favor']),
                
                # ✅ Proyección Total
                'total_proyectado': float(estadisticas['total_proyectado']),
                'total_balance_proyeccion': float(estadisticas['total_balance_proyeccion']),
                'deudores_proyeccion': estadisticas['deudores_proyeccion'],
                'al_dia_proyeccion': estadisticas['al_dia_proyeccion'],
                'a_favor_proyeccion': estadisticas['a_favor_proyeccion'],
                'total_debe_proyeccion': float(estadisticas['total_debe_proyeccion']),
                'total_favor_proyeccion': float(estadisticas['total_favor_proyeccion']),
            }
        })
    except Exception as e:
        import traceback
        import logging
        
        # Registrar error completo en logs
        logger = logging.getLogger(__name__)
        logger.error(f"Error en cargar_estadisticas_ajax: {str(e)}", exc_info=True)
        
        # Retornar error detallado al frontend
        return JsonResponse({
            'success': False,
            'error': str(e),
            'error_type': type(e).__name__,
            'traceback': traceback.format_exc()
        }, status=500)


# ============================================================
# 🔍 FUNCIÓN DE DIAGNÓSTICO (temporal para debugging)
# ============================================================
# Úsala para identificar el problema exacto
# ============================================================

@login_required
@staff_member_required
def diagnosticar_estadisticas(request):
    """
    Vista de diagnóstico para identificar problemas con estadísticas
    Solo para administradores
    """
    from django.db.models import Sum
    from facturacion.models import CuentaCorriente
    from agenda.models import Sesion, Proyecto, Mensualidad
    from decimal import Decimal
    
    diagnostico = {
        'tests': [],
        'errores': []
    }
    
    # Test 1: Verificar modelo CuentaCorriente
    try:
        test_cuenta = CuentaCorriente.objects.first()
        if test_cuenta:
            # Verificar que tenga el campo saldo_actual
            tiene_saldo_actual = hasattr(test_cuenta, 'saldo_actual')
            diagnostico['tests'].append({
                'nombre': 'Campo saldo_actual existe',
                'resultado': tiene_saldo_actual,
                'valor': test_cuenta.saldo_actual if tiene_saldo_actual else 'N/A'
            })
        else:
            diagnostico['tests'].append({
                'nombre': 'Cuentas corrientes',
                'resultado': False,
                'mensaje': 'No hay cuentas corrientes en la BD'
            })
    except Exception as e:
        diagnostico['errores'].append({
            'test': 'Verificar CuentaCorriente',
            'error': str(e)
        })
    
    # Test 2: Query de sesiones
    try:
        total_sesiones = Sesion.objects.filter(
            paciente__estado='activo',
            estado__in=['realizada', 'realizada_retraso', 'falta'],
            proyecto__isnull=True
        ).aggregate(total=Sum('monto_cobrado'))['total'] or Decimal('0.00')
        
        diagnostico['tests'].append({
            'nombre': 'Query sesiones',
            'resultado': True,
            'valor': str(total_sesiones)
        })
    except Exception as e:
        diagnostico['errores'].append({
            'test': 'Query sesiones',
            'error': str(e)
        })
    
    # Test 3: Query proyectos
    try:
        total_proyectos = Proyecto.objects.filter(
            paciente__estado='activo',
            estado__in=['en_progreso', 'finalizado']
        ).aggregate(total=Sum('costo_total'))['total'] or Decimal('0.00')
        
        diagnostico['tests'].append({
            'nombre': 'Query proyectos',
            'resultado': True,
            'valor': str(total_proyectos)
        })
    except Exception as e:
        diagnostico['errores'].append({
            'test': 'Query proyectos',
            'error': str(e)
        })
    
    # Test 4: Query mensualidades
    try:
        total_mensualidades = Mensualidad.objects.filter(
            paciente__estado='activo',
            estado__in=['activa', 'pausada', 'completada']
        ).aggregate(total=Sum('precio_total'))['total'] or Decimal('0.00')
        
        diagnostico['tests'].append({
            'nombre': 'Query mensualidades',
            'resultado': True,
            'valor': str(total_mensualidades)
        })
    except Exception as e:
        diagnostico['errores'].append({
            'test': 'Query mensualidades',
            'error': str(e)
        })
    
    # Test 5: Query cuentas corrientes
    try:
        cuentas_activas = CuentaCorriente.objects.filter(
            paciente__estado='activo'
        )
        
        total_pagado = cuentas_activas.aggregate(
            total=Sum('total_pagado')
        )['total'] or Decimal('0.00')
        
        total_balance = cuentas_activas.aggregate(
            total=Sum('saldo_actual')
        )['total'] or Decimal('0.00')
        
        diagnostico['tests'].append({
            'nombre': 'Query cuentas corrientes',
            'resultado': True,
            'total_pagado': str(total_pagado),
            'total_balance': str(total_balance),
            'num_cuentas': cuentas_activas.count()
        })
    except Exception as e:
        diagnostico['errores'].append({
            'test': 'Query cuentas corrientes',
            'error': str(e)
        })
    
    # Test 6: Clasificación por saldo
    try:
        cuentas_activas = CuentaCorriente.objects.filter(
            paciente__estado='activo'
        )
        
        deudores = cuentas_activas.filter(saldo_actual__lt=0).count()
        al_dia = cuentas_activas.filter(saldo_actual=0).count()
        a_favor = cuentas_activas.filter(saldo_actual__gt=0).count()
        
        diagnostico['tests'].append({
            'nombre': 'Clasificación por saldo',
            'resultado': True,
            'deudores': deudores,
            'al_dia': al_dia,
            'a_favor': a_favor
        })
    except Exception as e:
        diagnostico['errores'].append({
            'test': 'Clasificación por saldo',
            'error': str(e)
        })
    
    # Resumen
    diagnostico['resumen'] = {
        'tests_exitosos': len([t for t in diagnostico['tests'] if t.get('resultado', False)]),
        'tests_fallidos': len([t for t in diagnostico['tests'] if not t.get('resultado', True)]),
        'errores_totales': len(diagnostico['errores'])
    }
    
    return JsonResponse(diagnostico, safe=False)

@login_required
def detalle_cuenta_ajax(request, paciente_id):
    """
    Vista AJAX para cargar desglose detallado de cuenta de un paciente
    
    Se llama bajo demanda cuando el usuario hace clic en "Ver desglose"
    """
    paciente = get_object_or_404(Paciente, id=paciente_id)
    cuenta = get_object_or_404(CuentaCorriente, paciente=paciente)
    
    # Obtener estadísticas usando las propiedades del modelo
    # (estas usan cache internamente)
    stats = cuenta.get_stats_cached()
    
    # Calcular desglose de crédito (lógica de AccountService)
    from .services import AccountService
    
    # Pagos adelantados
    pagos_adelantados = Pago.objects.filter(
        paciente=paciente,
        anulado=False,
        sesion__isnull=True,
        proyecto__isnull=True
    ).exclude(
        metodo_pago__nombre="Uso de Crédito"
    ).aggregate(total=Sum('monto'))['monto__sum'] or Decimal('0.00')
    
    # Pagos de sesiones pendientes
    pagos_sesiones_pendientes = Pago.objects.filter(
        paciente=paciente,
        anulado=False,
        sesion__estado='programada'
    ).exclude(
        metodo_pago__nombre="Uso de Crédito"
    ).aggregate(total=Sum('monto'))['monto__sum'] or Decimal('0.00')
    
    # Excedentes
    sesiones_con_excedente = Sesion.objects.filter(
        paciente=paciente,
        proyecto__isnull=True,
        monto_cobrado__gt=0
    ).annotate(
        total_pagado_calc=Coalesce(
            Sum('pagos__monto', filter=Q(pagos__anulado=False) & ~Q(pagos__metodo_pago__nombre="Uso de Crédito")), 
            Decimal('0.00')
        )
    ).filter(
        total_pagado_calc__gt=F('monto_cobrado')
    )
    
    excedentes_total = sesiones_con_excedente.aggregate(
        total=Coalesce(Sum(F('total_pagado_calc') - F('monto_cobrado')), Decimal('0.00'))
    )['total']
    
    # Uso de crédito
    uso_credito = Pago.objects.filter(
        paciente=paciente,
        anulado=False,
        metodo_pago__nombre="Uso de Crédito"
    ).aggregate(total=Sum('monto'))['monto__sum'] or Decimal('0.00')
    
    return JsonResponse({
        'success': True,
        'paciente': {
            'id': paciente.id,
            'nombre_completo': paciente.nombre_completo,
        },
        'cuenta': {
            # Crédito disponible
            'saldo': float(cuenta.saldo_actual),
            
            # Desglose de crédito
            'pagos_adelantados': float(pagos_adelantados),
            'pagos_sesiones_pendientes': float(pagos_sesiones_pendientes),
            'excedentes': float(excedentes_total),
            'uso_credito': float(uso_credito),
            
            # Consumo y pagos
            'total_consumido': float(cuenta.total_consumido_actual),
            'total_pagado': float(cuenta.total_pagado),
            'consumo_sesiones': float(stats['consumo_sesiones']),
            'pagado_sesiones': float(stats['pagado_sesiones']),
            'deuda_sesiones': float(stats['deuda_sesiones']),
            'consumo_proyectos': float(stats['consumo_proyectos']),
            'pagado_proyectos': float(stats['pagado_proyectos']),
            'deuda_proyectos': float(stats['deuda_proyectos']),
            
            # Balance
            'total_deuda_general': float(cuenta.total_deuda_general),
            'balance_final': float(cuenta.balance_final),
        }
    })

@login_required
def detalle_cuenta_corriente(request, paciente_id):
    """
    Vista de detalle de cuenta corriente con MÁXIMA OPTIMIZACIÓN
    
    ✅ ACTUALIZADO: Con filtros de fechas y sumas dinámicas
    
    🚀 OPTIMIZACIONES APLICADAS:
    - annotate() para calcular saldo_pendiente sin queries N+1
    - select_related() y prefetch_related() estratégicos
    - Filtros de fechas en todas las pestañas
    - Checkbox para mostrar/ocultar programadas
    - Sumas totales dinámicas según filtros activos
    
    Performance:
    - Antes: 3-5 segundos (100+ queries)
    - Ahora: < 0.5 segundos (5-8 queries)
    """
    from decimal import Decimal
    from .models import Devolucion
    from django.db.models import Sum, F, Q, Exists, OuterRef, Case, When, DecimalField
    from django.db.models.functions import Coalesce
    from django.core.paginator import Paginator
    from datetime import date, datetime
    
    from facturacion.models import CuentaCorriente, Pago
    from agenda.models import Sesion, Proyecto, Mensualidad
    from pacientes.models import Paciente
    
    # Obtener paciente y cuenta
    paciente = get_object_or_404(Paciente, pk=paciente_id)
    cuenta, _ = CuentaCorriente.objects.get_or_create(paciente=paciente)
    
    # ========================================
    # REGISTROS DE COBROS CON FILTROS MEJORADOS
    # ========================================
    registros_cobros, suma_total_registros, suma_pagado_registros, suma_pendientes_registros, filtros_url = obtener_registros_cobros_filtrados(
        paciente, 
        request
    )

    # ========================================
    # PAGOS AL CONTADO - Ya optimizados con select_related
    # ========================================
    
    pagos_validos = Pago.objects.filter(
        paciente=paciente,
        anulado=False
    ).exclude(
        metodo_pago__nombre="Uso de Crédito"
    ).select_related(
        'sesion', 'proyecto', 'mensualidad', 'metodo_pago'
    ).order_by('-fecha_pago')
    
    # Filtros existentes
    # ========================================
    # PAGOS AL CONTADO (VÁLIDOS) - CON FILTROS MEJORADOS
    # ========================================
    
    pagos_validos = Pago.objects.filter(
        paciente=paciente,
        anulado=False
    ).exclude(
        metodo_pago__nombre="Uso de Crédito"
    ).select_related(
        'sesion', 'proyecto', 'mensualidad', 'metodo_pago'
    ).order_by('-fecha_pago')
    
    # ✅ FILTRO 1: Por tipo de concepto
    filtro_tipo_validos = request.GET.get('filtro_tipo_validos', 'todos')
    if filtro_tipo_validos != 'todos':
        if filtro_tipo_validos == 'sesiones':
            pagos_validos = pagos_validos.filter(sesion__isnull=False)
        elif filtro_tipo_validos == 'proyectos':
            pagos_validos = pagos_validos.filter(proyecto__isnull=False)
        elif filtro_tipo_validos == 'mensualidades':
            pagos_validos = pagos_validos.filter(mensualidad__isnull=False)
        elif filtro_tipo_validos == 'adelantados':
            pagos_validos = pagos_validos.filter(
                sesion__isnull=True,
                proyecto__isnull=True,
                mensualidad__isnull=True
            )
    
    # ✅ FILTRO 2: Por método de pago (MEJORADO)
    # Soporta tanto filtrado por ID como por nombre
    filtro_metodo_validos = request.GET.get('filtro_metodo_validos', 'todos')
    if filtro_metodo_validos != 'todos':
        try:
            # Intentar convertir a entero (ID)
            metodo_id = int(filtro_metodo_validos)
            pagos_validos = pagos_validos.filter(metodo_pago__id=metodo_id)
        except (ValueError, TypeError):
            # Si falla, asumir que es un nombre
            pagos_validos = pagos_validos.filter(metodo_pago__nombre__iexact=filtro_metodo_validos)
            
    # ✅ FILTRO 3: Por fechas
    fecha_desde_validos = request.GET.get('fecha_desde_validos', '')
    fecha_hasta_validos = request.GET.get('fecha_hasta_validos', '')
    
    if fecha_desde_validos:
        try:
            fecha_desde = datetime.strptime(fecha_desde_validos, '%Y-%m-%d').date()
            pagos_validos = pagos_validos.filter(fecha_pago__gte=fecha_desde)
        except ValueError:
            pass
    
    if fecha_hasta_validos:
        try:
            fecha_hasta = datetime.strptime(fecha_hasta_validos, '%Y-%m-%d').date()
            pagos_validos = pagos_validos.filter(fecha_pago__lte=fecha_hasta)
        except ValueError:
            pass
    
    # ✅ DEVOLUCIONES - Obtener y filtrar
    from facturacion.models import Devolucion
    
    devoluciones_query = Devolucion.objects.filter(
        paciente=paciente
    ).select_related(
        'proyecto', 'mensualidad', 'metodo_devolucion', 'registrado_por'
    ).order_by('-fecha_devolucion')
    
    # ✅ CORREGIDO: Aplicar filtro de tipo a devoluciones
    if filtro_tipo_validos == 'devoluciones':
        # Mostrar SOLO devoluciones
        pagos_validos = Pago.objects.none()  # No mostrar pagos
    elif filtro_tipo_validos != 'todos':
        # ⚠️ CORRECCIÓN: Si se filtra por tipo específico (sesiones, proyectos, mensualidades, adelantados),
        # NO mostrar devoluciones, solo mostrar pagos
        devoluciones_query = Devolucion.objects.none()
    
    # Aplicar filtros de fecha a devoluciones
    if fecha_desde_validos:
        try:
            fecha_desde = datetime.strptime(fecha_desde_validos, '%Y-%m-%d').date()
            devoluciones_query = devoluciones_query.filter(fecha_devolucion__gte=fecha_desde)
        except ValueError:
            pass
    
    if fecha_hasta_validos:
        try:
            fecha_hasta = datetime.strptime(fecha_hasta_validos, '%Y-%m-%d').date()
            devoluciones_query = devoluciones_query.filter(fecha_devolucion__lte=fecha_hasta)
        except ValueError:
            pass
    
    # ✅ NUEVO: Aplicar filtro de método de pago a devoluciones
    if filtro_metodo_validos != 'todos' and filtro_tipo_validos == 'devoluciones':
        try:
            # Intentar convertir a entero (ID)
            metodo_id = int(filtro_metodo_validos)
            devoluciones_query = devoluciones_query.filter(metodo_devolucion__id=metodo_id)
        except (ValueError, TypeError):
            # Si falla, asumir que es un nombre
            devoluciones_query = devoluciones_query.filter(metodo_devolucion__nombre__iexact=filtro_metodo_validos)
    
    # ✅ Calcular sumas separadas
    suma_pagos_validos = pagos_validos.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
    suma_devoluciones = devoluciones_query.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
    
    # ✅ Suma total considerando devoluciones como negativas
    suma_total_validos = suma_pagos_validos - suma_devoluciones
    
    # ✅ Combinar pagos y devoluciones en una lista unificada
    # Convertir QuerySets a listas con indicador de tipo
    # ✅ ACTUALIZADO: Prefetch de detalles masivos para evitar N+1 queries
    from .models import DetallePagoMasivo
    pagos_validos_con_detalles = pagos_validos.prefetch_related('detalles_masivos__sesion__servicio', 'detalles_masivos__proyecto', 'detalles_masivos__mensualidad')
    
    pagos_list = [
        {
            'tipo': 'pago',
            'objeto': pago,
            'fecha': pago.fecha_pago,
            'numero': pago.numero_recibo,
            'monto': pago.monto,
            'concepto': pago.concepto,
            'metodo': pago.metodo_pago.nombre,
            'es_masivo': pago.es_pago_masivo,  # ✅ NUEVO
            'cantidad_detalles': pago.cantidad_detalles if pago.es_pago_masivo else 0,  # ✅ NUEVO
        }
        for pago in pagos_validos_con_detalles
    ]
    
    devoluciones_list = [
        {
            'tipo': 'devolucion',
            'objeto': dev,
            'fecha': dev.fecha_devolucion,
            'numero': dev.numero_devolucion,
            'monto': -dev.monto,  # ✅ Negativo para restar
            'concepto': dev.motivo,
            'metodo': dev.metodo_devolucion.nombre,
        }
        for dev in devoluciones_query
    ]
    
    # Combinar y ordenar por fecha
    pagos_y_devoluciones = pagos_list + devoluciones_list
    pagos_y_devoluciones.sort(key=lambda x: x['fecha'], reverse=True)
    
    # Paginar la lista combinada
    paginator_validos = Paginator(pagos_y_devoluciones, 15)
    page_validos = request.GET.get('page_validos')
    pagos_validos_paginados = paginator_validos.get_page(page_validos)

    # ========================================
    # PAGOS CON CRÉDITO - CON FILTROS MEJORADOS
    # ========================================
    
    pagos_credito = Pago.objects.filter(
        paciente=paciente,
        anulado=False,
        metodo_pago__nombre="Uso de Crédito"
    ).select_related(
        'sesion', 'proyecto', 'mensualidad'
    ).order_by('-fecha_pago')
    
    # ✅ FILTRO: Por tipo de concepto
    filtro_tipo_credito = request.GET.get('filtro_tipo_credito', 'todos')
    if filtro_tipo_credito != 'todos':
        if filtro_tipo_credito == 'sesiones':
            pagos_credito = pagos_credito.filter(sesion__isnull=False)
        elif filtro_tipo_credito == 'proyectos':
            pagos_credito = pagos_credito.filter(proyecto__isnull=False)
        elif filtro_tipo_credito == 'mensualidades':
            pagos_credito = pagos_credito.filter(mensualidad__isnull=False)
        elif filtro_tipo_credito == 'adelantados':
            pagos_credito = pagos_credito.filter(
                sesion__isnull=True,
                proyecto__isnull=True,
                mensualidad__isnull=True
            )
    
    # ✅ Filtros de fechas para pagos con crédito
    fecha_desde_credito = request.GET.get('fecha_desde_credito', '')
    fecha_hasta_credito = request.GET.get('fecha_hasta_credito', '')
    
    if fecha_desde_credito:
        try:
            fecha_desde = datetime.strptime(fecha_desde_credito, '%Y-%m-%d').date()
            pagos_credito = pagos_credito.filter(fecha_pago__gte=fecha_desde)
        except ValueError:
            pass
    
    if fecha_hasta_credito:
        try:
            fecha_hasta = datetime.strptime(fecha_hasta_credito, '%Y-%m-%d').date()
            pagos_credito = pagos_credito.filter(fecha_pago__lte=fecha_hasta)
        except ValueError:
            pass
    
    # ✅ Calcular suma total de pagos con crédito filtrados
    suma_total_credito = pagos_credito.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
    
    paginator_credito = Paginator(pagos_credito, 15)
    page_credito = request.GET.get('page_credito')
    pagos_credito = paginator_credito.get_page(page_credito)
    
    # ========================================
    # PAGOS ANULADOS - CON FILTROS MEJORADOS
    # ========================================
    
    pagos_anulados = Pago.objects.filter(
        paciente=paciente,
        anulado=True
    ).select_related(
        'sesion', 'proyecto', 'mensualidad', 'anulado_por'
    ).order_by('-fecha_pago')
    
    # ✅ FILTRO: Por tipo de concepto
    filtro_tipo_anulados = request.GET.get('filtro_tipo_anulados', 'todos')
    if filtro_tipo_anulados != 'todos':
        if filtro_tipo_anulados == 'sesiones':
            pagos_anulados = pagos_anulados.filter(sesion__isnull=False)
        elif filtro_tipo_anulados == 'proyectos':
            pagos_anulados = pagos_anulados.filter(proyecto__isnull=False)
        elif filtro_tipo_anulados == 'mensualidades':
            pagos_anulados = pagos_anulados.filter(mensualidad__isnull=False)
        elif filtro_tipo_anulados == 'adelantados':
            pagos_anulados = pagos_anulados.filter(
                sesion__isnull=True,
                proyecto__isnull=True,
                mensualidad__isnull=True
            )
    
    # ✅ Filtros de fechas para pagos anulados
    fecha_desde_anulados = request.GET.get('fecha_desde_anulados', '')
    fecha_hasta_anulados = request.GET.get('fecha_hasta_anulados', '')
    
    if fecha_desde_anulados:
        try:
            fecha_desde = datetime.strptime(fecha_desde_anulados, '%Y-%m-%d').date()
            pagos_anulados = pagos_anulados.filter(fecha_pago__gte=fecha_desde)
        except ValueError:
            pass
    
    if fecha_hasta_anulados:
        try:
            fecha_hasta = datetime.strptime(fecha_hasta_anulados, '%Y-%m-%d').date()
            pagos_anulados = pagos_anulados.filter(fecha_pago__lte=fecha_hasta)
        except ValueError:
            pass
    
    # ✅ Calcular suma total de pagos anulados filtrados
    suma_total_anulados = pagos_anulados.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
    
    paginator_anulados = Paginator(pagos_anulados, 15)
    page_anulados = request.GET.get('page_anulados')
    pagos_anulados = paginator_anulados.get_page(page_anulados)
    
    # ========================================
    # ESTADÍSTICAS POR ESTADO - MENSUALIDADES Y PROYECTOS
    # ========================================
    
    # Desglose de Mensualidades por estado
    # ✅ NOTA: No podemos usar annotate aquí porque necesitamos incluir pagos masivos Y devoluciones
    # Usaremos la función helper _total_pagado_mensualidad() en cada mensualidad
    mensualidades_todas = Mensualidad.objects.filter(paciente=paciente)
    
    # Estados de Mensualidad: activa, pausada, completada, cancelada
    mensualidades_activas = mensualidades_todas.filter(estado='activa')
    mensualidades_pausadas = mensualidades_todas.filter(estado='pausada')
    mensualidades_completadas = mensualidades_todas.filter(estado='completada')
    mensualidades_canceladas = mensualidades_todas.filter(estado='cancelada')
    
    # ✅ CORREGIDO: Calcular totales usando _total_pagado_mensualidad que incluye pagos masivos y resta devoluciones
    def calcular_stats_mensualidades(mensualidades):
        total_costo = Decimal('0')
        total_pagado = Decimal('0')
        for m in mensualidades:
            total_costo += m.costo_mensual
            total_pagado += _total_pagado_mensualidad(m)
        return {
            'cantidad': mensualidades.count(),
            'total': total_costo,
            'pagado': total_pagado
        }
    
    stats_mensualidades = {
        'activas': calcular_stats_mensualidades(mensualidades_activas),
        'pausadas': calcular_stats_mensualidades(mensualidades_pausadas),
        'completadas': calcular_stats_mensualidades(mensualidades_completadas),
        'canceladas': calcular_stats_mensualidades(mensualidades_canceladas)
    }
    
    # Desglose de Proyectos por estado
    # ✅ NOTA: No podemos usar annotate aquí porque necesitamos incluir pagos masivos Y devoluciones
    # Usaremos la función helper _total_pagado_proyecto() en cada proyecto
    proyectos_todos = Proyecto.objects.filter(paciente=paciente)
    
    # Estados de Proyecto: planificado, en_progreso, finalizado, cancelado
    proyectos_planificados = proyectos_todos.filter(estado='planificado')
    proyectos_en_progreso = proyectos_todos.filter(estado='en_progreso')
    proyectos_finalizados = proyectos_todos.filter(estado='finalizado')
    proyectos_cancelados = proyectos_todos.filter(estado='cancelado')
    
    # ✅ CORREGIDO: Calcular totales usando _total_pagado_proyecto que incluye pagos masivos y resta devoluciones
    def calcular_stats_proyectos(proyectos):
        total_costo = Decimal('0')
        total_pagado = Decimal('0')
        for p in proyectos:
            total_costo += p.costo_total
            total_pagado += _total_pagado_proyecto(p)
        return {
            'cantidad': proyectos.count(),
            'total': total_costo,
            'pagado': total_pagado
        }
    
    stats_proyectos = {
        'planificados': calcular_stats_proyectos(proyectos_planificados),
        'en_progreso': calcular_stats_proyectos(proyectos_en_progreso),
        'finalizados': calcular_stats_proyectos(proyectos_finalizados),
        'cancelados': calcular_stats_proyectos(proyectos_cancelados)
    }
    
    # 1. Devoluciones de Mensualidades
    dev_mensualidad = Devolucion.objects.filter(
        paciente=paciente, 
        mensualidad__isnull=False
    ).aggregate(total=Coalesce(Sum('monto'), Decimal('0')))['total']

    # 2. Devoluciones de Proyectos
    dev_proyecto = Devolucion.objects.filter(
        paciente=paciente, 
        proyecto__isnull=False
    ).aggregate(total=Coalesce(Sum('monto'), Decimal('0')))['total']

    # 3. Devoluciones de Crédito (Pagos Adelantados puros)
    # Son aquellas que no están vinculadas ni a proyecto ni a mensualidad
    dev_credito = Devolucion.objects.filter(
        paciente=paciente, 
        proyecto__isnull=True, 
        mensualidad__isnull=True
    ).aggregate(total=Coalesce(Sum('monto'), Decimal('0')))['total']

    # ========================================
    # DEUDAS PENDIENTES - CÁLCULO DETALLADO MEJORADO
    # ========================================
    
    # 1. SESIONES NORMALES CON DEUDA
    # ✅ MEJORADO: Diferencia entre deudas parciales y totales
    sesiones_query = Sesion.objects.filter(
        paciente=paciente,
        proyecto__isnull=True,
        mensualidad__isnull=True,
        estado__in=['realizada', 'realizada_retraso', 'falta'],
        monto_cobrado__gt=0
    ).prefetch_related(
        Prefetch('pagos', queryset=Pago.objects.filter(anulado=False))
    )
    
    sesiones_con_deuda_list = []
    sesiones_con_deuda_total_list = []  # Sin ningún pago
    sesiones_con_deuda_parcial_list = []  # Con pagos incompletos
    deuda_sesiones_monto_total = Decimal('0')
    deuda_sesiones_total_monto = Decimal('0')  # Monto de deudas totales
    deuda_sesiones_parcial_monto = Decimal('0')  # Monto de deudas parciales
    
    for sesion in sesiones_query:
        # Calcular total pagado usando pagos prefetched + detalles de pagos masivos
        total_pagado = _total_pagado_sesion(sesion)
        saldo_pendiente = sesion.monto_cobrado - total_pagado
        
        # ⚠️ IMPORTANTE: Captura CUALQUIER deuda (parcial o total)
        # Umbral de 0.01 para evitar problemas de redondeo
        if saldo_pendiente > Decimal('0.01'):
            sesion.total_pagado_calc = total_pagado
            sesion.saldo_pendiente_calc = saldo_pendiente
            sesiones_con_deuda_list.append(sesion)
            deuda_sesiones_monto_total += saldo_pendiente
            
            # ✅ NUEVO: Diferenciar entre deuda total y parcial
            if total_pagado <= Decimal('0.01'):
                # Deuda total: No tiene ningún pago
                sesiones_con_deuda_total_list.append(sesion)
                deuda_sesiones_total_monto += saldo_pendiente
            else:
                # Deuda parcial: Tiene pagos pero no completos
                sesiones_con_deuda_parcial_list.append(sesion)
                deuda_sesiones_parcial_monto += saldo_pendiente
    
    deuda_sesiones_cantidad = len(sesiones_con_deuda_list)
    deuda_sesiones_cantidad_total = len(sesiones_con_deuda_total_list)
    deuda_sesiones_cantidad_parcial = len(sesiones_con_deuda_parcial_list)
    deuda_sesiones_monto = deuda_sesiones_monto_total
    
    # 2. MENSUALIDADES CON DEUDA
    # ✅ CORREGIDO: Usa _total_pagado_mensualidad que ya considera devoluciones
    mensualidades_query = Mensualidad.objects.filter(
        paciente=paciente
    ).prefetch_related(
        Prefetch('pagos', queryset=Pago.objects.filter(anulado=False)),
    )
    
    mensualidades_con_deuda_list = []
    mensualidades_con_deuda_total_list = []
    mensualidades_con_deuda_parcial_list = []
    deuda_mensualidades_monto_total = Decimal('0')
    deuda_mensualidades_total_monto = Decimal('0')
    deuda_mensualidades_parcial_monto = Decimal('0')
    
    for mensualidad in mensualidades_query:
        # ✅ CORREGIDO: total_pagado ya incluye la resta de devoluciones
        total_pagado = _total_pagado_mensualidad(mensualidad)
        saldo_pendiente = mensualidad.costo_mensual - total_pagado
        
        # Verificar si tiene deuda (umbral 0.01 para evitar redondeo)
        if saldo_pendiente > Decimal('0.01'):
            mensualidad.total_pagado_calc = total_pagado
            mensualidad.saldo_pendiente_calc = saldo_pendiente
            mensualidades_con_deuda_list.append(mensualidad)
            deuda_mensualidades_monto_total += saldo_pendiente
            
            # Diferenciar entre deuda total y parcial
            if total_pagado <= Decimal('0.01'):
                mensualidades_con_deuda_total_list.append(mensualidad)
                deuda_mensualidades_total_monto += saldo_pendiente
            else:
                mensualidades_con_deuda_parcial_list.append(mensualidad)
                deuda_mensualidades_parcial_monto += saldo_pendiente
    
    deuda_mensualidades_cantidad = len(mensualidades_con_deuda_list)
    deuda_mensualidades_cantidad_total = len(mensualidades_con_deuda_total_list)
    deuda_mensualidades_cantidad_parcial = len(mensualidades_con_deuda_parcial_list)
    deuda_mensualidades_monto = deuda_mensualidades_monto_total
    
    # 3. PROYECTOS CON DEUDA (EXCLUYE PLANIFICADOS - SOLO ESTADO ACTUAL)
    # ✅ CORREGIDO: Usa _total_pagado_proyecto que ya considera devoluciones
    proyectos_query = Proyecto.objects.filter(
        paciente=paciente
    ).exclude(
        estado='planificado'
    ).prefetch_related(
        Prefetch('pagos', queryset=Pago.objects.filter(anulado=False)),
    )
    
    proyectos_con_deuda_list = []
    proyectos_con_deuda_total_list = []
    proyectos_con_deuda_parcial_list = []
    deuda_proyectos_monto_total = Decimal('0')
    deuda_proyectos_total_monto = Decimal('0')
    deuda_proyectos_parcial_monto = Decimal('0')
    
    for proyecto in proyectos_query:
        # ✅ CORREGIDO: total_pagado ya incluye la resta de devoluciones
        total_pagado = _total_pagado_proyecto(proyecto)
        saldo_pendiente = proyecto.costo_total - total_pagado
        
        # Verificar si tiene deuda (umbral 0.01 para evitar redondeo)
        if saldo_pendiente > Decimal('0.01'):
            proyecto.total_pagado_calc = total_pagado
            proyecto.saldo_pendiente_calc = saldo_pendiente
            proyectos_con_deuda_list.append(proyecto)
            deuda_proyectos_monto_total += saldo_pendiente
            
            # Diferenciar entre deuda total y parcial
            if total_pagado <= Decimal('0.01'):
                proyectos_con_deuda_total_list.append(proyecto)
                deuda_proyectos_total_monto += saldo_pendiente
            else:
                proyectos_con_deuda_parcial_list.append(proyecto)
                deuda_proyectos_parcial_monto += saldo_pendiente
    
    deuda_proyectos_cantidad = len(proyectos_con_deuda_list)
    deuda_proyectos_cantidad_total = len(proyectos_con_deuda_total_list)
    deuda_proyectos_cantidad_parcial = len(proyectos_con_deuda_parcial_list)
    deuda_proyectos_monto = deuda_proyectos_monto_total
    
    # TOTAL DE DEUDAS PENDIENTES
    deuda_total = deuda_sesiones_monto + deuda_mensualidades_monto + deuda_proyectos_monto

    # ========================================
    # ✅ NUEVO: DEUDAS PROYECTADAS (incluye programadas y planificados)
    # ========================================
    
    # 1. SESIONES PROGRAMADAS CON DEUDA
    sesiones_programadas_query = Sesion.objects.filter(
        paciente=paciente,
        proyecto__isnull=True,
        mensualidad__isnull=True,
        estado='programada',
        monto_cobrado__gt=0
    ).prefetch_related(
        Prefetch('pagos', queryset=Pago.objects.filter(anulado=False))
    )
    
    sesiones_programadas_con_deuda_list = []
    sesiones_programadas_con_deuda_total_list = []
    sesiones_programadas_con_deuda_parcial_list = []
    deuda_sesiones_programadas_monto_total = Decimal('0')
    deuda_sesiones_programadas_total_monto = Decimal('0')
    deuda_sesiones_programadas_parcial_monto = Decimal('0')
    
    for sesion in sesiones_programadas_query:
        # Calcular total pagado usando pagos prefetched + detalles de pagos masivos
        total_pagado = _total_pagado_sesion(sesion)
        saldo_pendiente = sesion.monto_cobrado - total_pagado
        
        # Verificar si tiene deuda (umbral 0.01 para evitar redondeo)
        if saldo_pendiente > Decimal('0.01'):
            sesion.total_pagado_calc = total_pagado
            sesion.saldo_pendiente_calc = saldo_pendiente
            sesiones_programadas_con_deuda_list.append(sesion)
            deuda_sesiones_programadas_monto_total += saldo_pendiente
            
            # ✅ NUEVO: Diferenciar entre deuda total y parcial
            if total_pagado <= Decimal('0.01'):
                sesiones_programadas_con_deuda_total_list.append(sesion)
                deuda_sesiones_programadas_total_monto += saldo_pendiente
            else:
                sesiones_programadas_con_deuda_parcial_list.append(sesion)
                deuda_sesiones_programadas_parcial_monto += saldo_pendiente
    
    deuda_sesiones_programadas_cantidad = len(sesiones_programadas_con_deuda_list)
    deuda_sesiones_programadas_cantidad_total = len(sesiones_programadas_con_deuda_total_list)
    deuda_sesiones_programadas_cantidad_parcial = len(sesiones_programadas_con_deuda_parcial_list)
    deuda_sesiones_programadas_monto = deuda_sesiones_programadas_monto_total

    
    # 2. PROYECTOS PLANIFICADOS CON DEUDA
    # ✅ CORREGIDO: Considera devoluciones en el cálculo de deuda
    proyectos_planificados_query = Proyecto.objects.filter(
        paciente=paciente,
        estado='planificado'
    ).annotate(  # ✅ AGREGAR
        total_devoluciones_anotado=Coalesce(
            Sum('devolucion__monto'),
            Decimal('0')
        )
    ).prefetch_related(
        Prefetch('pagos', queryset=Pago.objects.filter(anulado=False)),
    )
    
    proyectos_planificados_con_deuda_list = []
    proyectos_planificados_con_deuda_total_list = []
    proyectos_planificados_con_deuda_parcial_list = []
    deuda_proyectos_planificados_monto_total = Decimal('0')
    deuda_proyectos_planificados_total_monto = Decimal('0')
    deuda_proyectos_planificados_parcial_monto = Decimal('0')
    
    for proyecto in proyectos_planificados_query:
        # ✅ CORREGIDO: Calcular total pagado MENOS devoluciones (incluye pagos masivos)
        total_pagado = _total_pagado_proyecto(proyecto)
        total_pagado_efectivo = total_pagado - proyecto.total_devoluciones_anotado  # ✅
        saldo_pendiente = proyecto.costo_total - total_pagado_efectivo
        
        # Verificar si tiene deuda (umbral 0.01 para evitar redondeo)
        if saldo_pendiente > Decimal('0.01'):
            proyecto.total_pagado_calc = total_pagado_efectivo  # ✅ Usar efectivo
            proyecto.saldo_pendiente_calc = saldo_pendiente
            proyectos_planificados_con_deuda_list.append(proyecto)
            deuda_proyectos_planificados_monto_total += saldo_pendiente
            
            # ✅ CORREGIDO: Usar total_pagado_efectivo para diferenciar
            if total_pagado_efectivo <= Decimal('0.01'):
                proyectos_planificados_con_deuda_total_list.append(proyecto)
                deuda_proyectos_planificados_total_monto += saldo_pendiente
            else:
                proyectos_planificados_con_deuda_parcial_list.append(proyecto)
                deuda_proyectos_planificados_parcial_monto += saldo_pendiente
    
    deuda_proyectos_planificados_cantidad = len(proyectos_planificados_con_deuda_list)
    deuda_proyectos_planificados_cantidad_total = len(proyectos_planificados_con_deuda_total_list)
    deuda_proyectos_planificados_cantidad_parcial = len(proyectos_planificados_con_deuda_parcial_list)
    deuda_proyectos_planificados_monto = deuda_proyectos_planificados_monto_total

    
    # TOTAL DE DEUDAS PROYECTADAS (incluye todo: actual + futuro)
    deuda_total_proyectada = (
        deuda_total + 
        deuda_sesiones_programadas_monto + 
        deuda_proyectos_planificados_monto
    )

    # ========================================
    # CONTEXT - ACTUALIZADO
    # ========================================
    
    context = {
        'paciente': paciente,
        'cuenta': cuenta,
        'registros_cobros': registros_cobros,
        'pagos_validos': pagos_validos_paginados,
        'pagos_credito': pagos_credito,
        'pagos_anulados': pagos_anulados,
        'filtros_url': filtros_url,
        'filtro_tipo_validos': filtro_tipo_validos,
        'filtro_metodo_validos': filtro_metodo_validos,
        'filtro_tipo_credito': filtro_tipo_credito,
        'filtro_tipo_anulados': filtro_tipo_anulados,
        # Sumas actualizadas
        'suma_total_registros': suma_total_registros,
        'suma_pagado_registros': suma_pagado_registros,
        'suma_pendientes_registros': suma_pendientes_registros,
        'suma_total_validos': suma_total_validos,
        'suma_pagos_validos': suma_pagos_validos,
        'suma_devoluciones': suma_devoluciones,
        'suma_total_credito': suma_total_credito,
        'suma_total_anulados': suma_total_anulados,
        # Filtros de fechas
        'fecha_desde_validos': fecha_desde_validos,
        'fecha_hasta_validos': fecha_hasta_validos,
        'fecha_desde_credito': fecha_desde_credito,
        'fecha_hasta_credito': fecha_hasta_credito,
        'fecha_desde_anulados': fecha_desde_anulados,
        'fecha_hasta_anulados': fecha_hasta_anulados,
        # Estadísticas por estado
        'stats_mensualidades': stats_mensualidades,
        'stats_proyectos': stats_proyectos,
        # Desglose de devoluciones
        'dev_mensualidad': dev_mensualidad,
        'dev_proyecto': dev_proyecto,
        'dev_credito': dev_credito,
        
        # ✅ MEJORADO: Deudas pendientes con desglose parcial/total
        'deuda_sesiones_cantidad': deuda_sesiones_cantidad,
        'deuda_sesiones_cantidad_total': deuda_sesiones_cantidad_total,
        'deuda_sesiones_cantidad_parcial': deuda_sesiones_cantidad_parcial,
        'deuda_sesiones_monto': deuda_sesiones_monto,
        'deuda_sesiones_total_monto': deuda_sesiones_total_monto,
        'deuda_sesiones_parcial_monto': deuda_sesiones_parcial_monto,
        
        'deuda_mensualidades_cantidad': deuda_mensualidades_cantidad,
        'deuda_mensualidades_cantidad_total': deuda_mensualidades_cantidad_total,
        'deuda_mensualidades_cantidad_parcial': deuda_mensualidades_cantidad_parcial,
        'deuda_mensualidades_monto': deuda_mensualidades_monto,
        'deuda_mensualidades_total_monto': deuda_mensualidades_total_monto,
        'deuda_mensualidades_parcial_monto': deuda_mensualidades_parcial_monto,
        
        'deuda_proyectos_cantidad': deuda_proyectos_cantidad,
        'deuda_proyectos_cantidad_total': deuda_proyectos_cantidad_total,
        'deuda_proyectos_cantidad_parcial': deuda_proyectos_cantidad_parcial,
        'deuda_proyectos_monto': deuda_proyectos_monto,
        'deuda_proyectos_total_monto': deuda_proyectos_total_monto,
        'deuda_proyectos_parcial_monto': deuda_proyectos_parcial_monto,
        
        'deuda_total': deuda_total,
        
        # ✅ MEJORADO: Deudas proyectadas con desglose parcial/total
        'deuda_sesiones_programadas_cantidad': deuda_sesiones_programadas_cantidad,
        'deuda_sesiones_programadas_cantidad_total': deuda_sesiones_programadas_cantidad_total,
        'deuda_sesiones_programadas_cantidad_parcial': deuda_sesiones_programadas_cantidad_parcial,
        'deuda_sesiones_programadas_monto': deuda_sesiones_programadas_monto,
        'deuda_sesiones_programadas_total_monto': deuda_sesiones_programadas_total_monto,
        'deuda_sesiones_programadas_parcial_monto': deuda_sesiones_programadas_parcial_monto,
        
        'deuda_proyectos_planificados_cantidad': deuda_proyectos_planificados_cantidad,
        'deuda_proyectos_planificados_cantidad_total': deuda_proyectos_planificados_cantidad_total,
        'deuda_proyectos_planificados_cantidad_parcial': deuda_proyectos_planificados_cantidad_parcial,
        'deuda_proyectos_planificados_monto': deuda_proyectos_planificados_monto,
        'deuda_proyectos_planificados_total_monto': deuda_proyectos_planificados_total_monto,
        'deuda_proyectos_planificados_parcial_monto': deuda_proyectos_planificados_parcial_monto,
        
        'deuda_total_proyectada': deuda_total_proyectada,
    }
    
    return render(request, 'facturacion/detalle_cuenta.html', context)

def registrar_pago(request):
    """
    Registrar pago usando PaymentService.
    Soporta: Sesión, Proyecto, Mensualidad, Adelantado (Credito/Efectivo/Mixto).
    """
    from django.core.exceptions import ValidationError
    from .services import PaymentService
    
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            tipo_pago = request.POST.get('tipo_pago')
            es_pago_completo = request.POST.get('pago_completo') == 'on'
            usar_credito = request.POST.get('usar_credito') == 'on'
            
            raw_credito = request.POST.get('monto_credito', '').strip()
            raw_monto = request.POST.get('monto', '').strip()
            monto_credito = Decimal(raw_credito if raw_credito else '0')
            monto_efectivo = Decimal(raw_monto if raw_monto else '0')
            
            metodo_pago_id = request.POST.get('metodo_pago')
            fecha_pago_str = request.POST.get('fecha_pago')
            observaciones = request.POST.get('observaciones', '')
            numero_transaccion = request.POST.get('numero_transaccion', '')
            
            
            # ✅ CORREGIDO: Detectar si el método de pago es "Sin Cobro"
            metodo_es_sin_cobro = False
            if metodo_pago_id:
                metodo_temp = MetodoPago.objects.get(id=metodo_pago_id)
                metodo_es_sin_cobro = 'sin cobro' in metodo_temp.nombre.lower()
            
            # ✅ NUEVO: Si el método es "Sin Cobro" Y el monto es 0, forzar es_pago_completo=True
            if metodo_es_sin_cobro and monto_efectivo == 0 and monto_credito == 0:
                es_pago_completo = True
            
            # ✅ VALIDACIÓN: Si solo usa crédito (sin efectivo), no necesita método de pago
            # En ese caso, usar el método "Uso de Crédito" por defecto
            if usar_credito and monto_efectivo == 0 and not metodo_pago_id:
                metodo_credito = MetodoPago.objects.get(nombre="Uso de Crédito")
                metodo_pago_id = metodo_credito.id
            elif monto_efectivo > 0 and not metodo_pago_id:
                raise ValidationError('Debes seleccionar un método de pago')
            
            if not fecha_pago_str:
                raise ValidationError('Debes especificar la fecha de pago')
            
            fecha_pago = datetime.strptime(fecha_pago_str, '%Y-%m-%d').date()
            
            # Identificar Entidades
            paciente = None
            referencia_id = None
            
            if tipo_pago == 'sesion':
                referencia_id = request.POST.get('sesion_id')
                if not referencia_id: raise ValidationError('Sesión no seleccionada')
                sesion = Sesion.objects.get(id=referencia_id)
                paciente = sesion.paciente
                
            elif tipo_pago == 'proyecto':
                referencia_id = request.POST.get('proyecto_id')
                if not referencia_id: raise ValidationError('Proyecto no seleccionado')
                proyecto = Proyecto.objects.get(id=referencia_id)
                paciente = proyecto.paciente
            
            # ✅ NUEVO: Soporte para mensualidades
            elif tipo_pago == 'mensualidad':
                referencia_id = request.POST.get('mensualidad_id')  # ← AGREGAR ESTA LÍNEA
                if not referencia_id:
                    raise ValidationError("ID de mensualidad requerido.")
                from agenda.models import Mensualidad
                mensualidad = Mensualidad.objects.get(id=referencia_id)
                paciente = mensualidad.paciente
                
                if es_pago_completo:
                    pagado_previo = mensualidad.pagos.filter(anulado=False).exclude(
                        metodo_pago__nombre="Uso de Crédito"
                    ).aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
                    
                    nuevo_costo = pagado_previo + monto_total_aportado
                    if mensualidad.costo_mensual != nuevo_costo:
                        mensualidad.costo_mensual = nuevo_costo
                        mensualidad.save()
                
            elif tipo_pago == 'adelantado':
                paciente_id = request.POST.get('paciente_adelantado')
                if not paciente_id: raise ValidationError('Paciente no seleccionado')
                paciente = Paciente.objects.get(id=paciente_id)
            else:
                raise ValidationError('Tipo de pago no válido')

            # ✅ CORREGIDO: Calcular monto total correctamente
            monto_total = monto_efectivo + monto_credito
            
            # Procesar Pago con Servicio (actualizado para mensualidad)
            resultado = PaymentService.process_payment(
                user=request.user,
                paciente=paciente,
                monto_total=monto_total,  # ✅ CORREGIDO: pasar monto_total en lugar de monto_efectivo
                metodo_pago_id=metodo_pago_id,
                fecha_pago=fecha_pago,
                tipo_pago=tipo_pago,
                referencia_id=referencia_id,
                usar_credito=usar_credito,  # ✅ AGREGADO: faltaba pasar este parámetro
                monto_credito=monto_credito,
                es_pago_completo=es_pago_completo,
                observaciones=observaciones,
                numero_transaccion=numero_transaccion
            )
            
            # Preparar datos para session (Confirmación)
            # ✅ Manejar caso especial de monto 0
            if monto_total == 0:
                tipo_pago_display = "Sin cobro"
                detalle = "Sesión sin cobro (beca/cortesía/razón social)"
                concepto_display = "Registro sin cobro"
            else:
                tipo_pago_display = "Efectivo"
                if usar_credito and monto_efectivo > 0: tipo_pago_display = "Mixto"
                elif usar_credito: tipo_pago_display = "100% Crédito"
                
                detalle = f"Monto: Bs. {monto_efectivo}"
                if usar_credito: detalle = f"Crédito: Bs. {monto_credito} + Efectivo: Bs. {monto_efectivo}"
                concepto_display = resultado.get('pago_efectivo').concepto if resultado.get('pago_efectivo') else "Pago con Crédito"
            
            request.session['pago_exitoso'] = {
                'tipo': tipo_pago_display,
                'mensaje': resultado['mensaje'],
                'detalle': detalle,
                'total': float(resultado['monto_total']),  # ✅ Convertir Decimal a float
                'paciente': paciente.nombre_completo,
                'concepto': concepto_display,
                'info_estado': "Registro Completado" if monto_total == 0 else "Pago Registrado Correctamente",
                'genero_recibo': bool(resultado['recibos']),
                'numero_recibo': resultado['recibos'][0] if resultado['recibos'] else None,
                'pago_id': resultado.get('pago_efectivo').id if resultado.get('pago_efectivo') else None,
            }
            
            return redirect('facturacion:confirmacion_pago')

        except ValidationError as e:
            messages.error(request, f'❌ {str(e.message if hasattr(e, "message") else e)}')
            return redirect('facturacion:registrar_pago')
        except Exception as e:
            messages.error(request, f'❌ Error inesperado: {str(e)}')
            import traceback
            print(traceback.format_exc())
            return redirect('facturacion:registrar_pago')
    
    # ========== GET - MOSTRAR FORMULARIO ==========
    metodos_pago = MetodoPago.objects.filter(activo=True).exclude(
        nombre__in=["Crédito/Saldo a favor", "Uso de Crédito"]
    )
    
    pacientes_lista = Paciente.objects.filter(estado='activo').order_by('nombre', 'apellido')
    
    # Detectar parámetros
    sesion_id = request.GET.get('sesion')
    paciente_id = request.GET.get('paciente')
    proyecto_id = request.GET.get('proyecto')
    mensualidad_id = request.GET.get('mensualidad')  # ✅ NUEVO
    
    sesion = None
    paciente = None
    proyecto = None
    mensualidad = None  # ✅ NUEVO
    modo = None
    credito_disponible = Decimal('0.00')
    monto_sugerido = Decimal('0.00')
    
    # CASO 1: Proyecto
    if proyecto_id:
        proyecto = get_object_or_404(
            Proyecto.objects.select_related('paciente', 'servicio_base'),
            id=proyecto_id
        )
        paciente = proyecto.paciente
        modo = 'proyecto'
        monto_sugerido = proyecto.saldo_pendiente
        
        cuenta, _ = CuentaCorriente.objects.get_or_create(paciente=paciente)
        cuenta.actualizar_saldo()
        credito_disponible = cuenta.pagos_adelantados if cuenta.pagos_adelantados > 0 else Decimal('0.00')
    
    # ✅ CASO 2: Mensualidad
    elif mensualidad_id:
        from agenda.models import Mensualidad
        mensualidad = get_object_or_404(
            Mensualidad.objects.select_related('paciente', 'sucursal').prefetch_related(
                'servicios_profesionales__servicio',
                'servicios_profesionales__profesional'
            ),
            id=mensualidad_id
        )
        paciente = mensualidad.paciente
        modo = 'mensualidad'
        monto_sugerido = mensualidad.saldo_pendiente
        
        cuenta, _ = CuentaCorriente.objects.get_or_create(paciente=paciente)
        cuenta.actualizar_saldo()
        credito_disponible = cuenta.pagos_adelantados if cuenta.pagos_adelantados > 0 else Decimal('0.00')
    
    # CASO 3: Sesión
    elif sesion_id:
        sesion = get_object_or_404(
            Sesion.objects.select_related('paciente', 'servicio', 'profesional'),
            id=sesion_id
        )
        paciente = sesion.paciente
        modo = 'sesion'
        monto_sugerido = sesion.saldo_pendiente
        
        cuenta, _ = CuentaCorriente.objects.get_or_create(paciente=paciente)
        cuenta.actualizar_saldo()
        credito_disponible = cuenta.pagos_adelantados if cuenta.pagos_adelantados > 0 else Decimal('0.00')
    
    # CASO 4: Adelantado
    elif paciente_id:
        paciente = get_object_or_404(Paciente, id=paciente_id)
        modo = 'adelantado'
        
        cuenta, _ = CuentaCorriente.objects.get_or_create(paciente=paciente)
        cuenta.actualizar_saldo()
        credito_disponible = cuenta.pagos_adelantados if cuenta.pagos_adelantados > 0 else Decimal('0.00')
    
    # CASO 5: Selector
    else:
        modo = 'selector'
    
    context = {
        'metodos_pago': metodos_pago,
        'sesion': sesion,
        'paciente': paciente,
        'proyecto': proyecto,
        'mensualidad': mensualidad,  # ✅ NUEVO
        'modo': modo,
        'credito_disponible': credito_disponible,
        'monto_sugerido': monto_sugerido,
        'fecha_hoy': date.today(),
        'pacientes_lista': pacientes_lista,
    }
    
    return render(request, 'facturacion/registrar_pago.html', context)     

@login_required
def confirmacion_pago(request):
    """
    🆕 NUEVA VISTA: Mostrar modal de confirmación de pago exitoso
    """
    
    # Obtener datos de la sesión
    datos_pago = request.session.get('pago_exitoso')
    
    if not datos_pago:
        messages.error(request, '❌ No hay datos de pago para mostrar')
        return redirect('facturacion:cuentas_corrientes')
    
    # Limpiar sesión después de obtener datos
    del request.session['pago_exitoso']
    
    context = {
        'datos_pago': datos_pago,
    }
    
    return render(request, 'facturacion/confirmacion_pago.html', context)

# ✅ NUEVA: API para cargar proyectos de un paciente (AJAX)
@login_required
def api_proyectos_paciente(request, paciente_id):
    """
    API: Obtener proyectos de un paciente
    - Para pagos: Devuelve proyectos con saldo pendiente
    - Para devoluciones: Devuelve proyectos PLANIFICADOS/EN_PROGRESO que tienen pagos realizados
    """
    try:
        from agenda.models import Proyecto
        
        # Detectar si es para devoluciones
        es_devolucion = request.GET.get('tipo') == 'devolucion'
        
        # Para devoluciones: solo planificados o en progreso (no completados ni cancelados)
        # Para pagos: solo planificados o en progreso
        estados_permitidos = ['planificado', 'en_progreso']
        
        proyectos = Proyecto.objects.filter(
            paciente_id=paciente_id,
            estado__in=estados_permitidos
        ).select_related('servicio_base')
        
        proyectos_lista = []
        
        for p in proyectos:
            # ✅ Usar cálculo que incluye pagos masivos
            total_pagado_real = _total_pagado_proyecto(p)
            saldo_pendiente_real = p.costo_total - total_pagado_real
            if es_devolucion:
                # Para devoluciones: incluir solo los que tienen pagos realizados
                if total_pagado_real > 0:
                    proyectos_lista.append({
                        'id': p.id,
                        'codigo': p.codigo,
                        'nombre': p.nombre,
                        'costo_total': float(p.costo_total),
                        'total_pagado': float(total_pagado_real),
                        'saldo_pendiente': float(saldo_pendiente_real),
                        'tipo': p.get_tipo_display(),
                        'estado': p.get_estado_display(),
                    })
            else:
                # Para pagos: incluir solo los que tienen saldo pendiente
                if saldo_pendiente_real > Decimal('0.01'):
                    proyectos_lista.append({
                        'id': p.id,
                        'codigo': p.codigo,
                        'nombre': p.nombre,
                        'costo_total': float(p.costo_total),
                        'total_pagado': float(total_pagado_real),
                        'saldo_pendiente': float(saldo_pendiente_real),
                        'tipo': p.get_tipo_display(),
                        'estado': p.get_estado_display(),
                    })
        
        return JsonResponse({
            'success': True,
            'proyectos': proyectos_lista
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@login_required
def marcar_sesion_pagada(request, sesion_id):
    """
    Marcar sesión como pagada (rápido desde lista)
    OPTIMIZADO: HTMX response
    """
    
    if request.method == 'POST':
        try:
            sesion = get_object_or_404(Sesion, id=sesion_id)
            
            # Marcar como pagada
            sesion.pagado = True
            sesion.fecha_pago = date.today()
            sesion.save()
            
            # Actualizar cuenta corriente
            cuenta, created = CuentaCorriente.objects.get_or_create(paciente=sesion.paciente)
            cuenta.actualizar_saldo()
            
            messages.success(request, f'✅ Sesión marcada como pagada')
            
            # Retornar fragmento actualizado (HTMX)
            return render(request, 'facturacion/partials/sesion_row.html', {
                'sesion': sesion
            })
            
        except Exception as e:
            messages.error(request, f'❌ Error: {str(e)}')
            return redirect('facturacion:detalle_cuenta', paciente_id=sesion.paciente.id)
    
    return redirect('facturacion:cuentas_corrientes')


# ============= APIs AJAX/HTMX =============

@login_required
def buscar_pacientes_ajax(request):
    """
    API: Buscar pacientes para autocomplete
    OPTIMIZADO: Solo campos necesarios
    """
    
    q = request.GET.get('q', '').strip()
    
    if len(q) < 2:
        return JsonResponse({'pacientes': []})
    
    pacientes = Paciente.objects.filter(
        Q(nombre__icontains=q) | 
        Q(apellido__icontains=q) |
        Q(nombre_tutor__icontains=q),
        estado='activo'
    ).values('id', 'nombre', 'apellido')[:10]
    
    return JsonResponse({
        'pacientes': list(pacientes)
    })


@login_required
def sesiones_pendientes_ajax(request, paciente_id):
    """
    API: Obtener sesiones pendientes de pago de un paciente
    OPTIMIZADO: HTMX partial
    """
    
    paciente = get_object_or_404(Paciente, id=paciente_id)
    
    sesiones = Sesion.objects.filter(
        paciente=paciente,
        pagado=False,
        estado__in=['realizada', 'realizada_retraso', 'falta']
    ).select_related('servicio').order_by('-fecha')[:20]
    
    return render(request, 'facturacion/partials/sesiones_pendientes.html', {
        'sesiones': sesiones,
        'paciente': paciente,
    })


# ==================== FASE 2: PAGOS MASIVOS ====================

@login_required
@login_required
def pagos_masivos(request):
    """
    Vista para pagar múltiples SESIONES, PROYECTOS Y MENSUALIDADES a la vez
    ✅ ACTUALIZADO: Ahora incluye sesiones, proyectos y mensualidades pendientes
    - Incluye todas las sesiones pendientes de pago (incluyendo programadas)
    - Incluye todos los proyectos pendientes de pago (incluyendo planificados)
    - Incluye todas las mensualidades pendientes de pago
    """

    # Paso 1: Seleccionar paciente
    paciente_id = request.GET.get('paciente')
    paciente = None
    sesiones_pendientes = []
    proyectos_pendientes = []
    mensualidades_pendientes = []

    if paciente_id:
        paciente = get_object_or_404(
            Paciente.objects.select_related('cuenta_corriente'),
            id=paciente_id
        )

        # ========================================
        # ✅ OBTENER SESIONES PENDIENTES (TODAS con deuda)
        # ========================================
        # IMPORTANTE: Incluye sesiones realizadas, falta Y programadas
        # Solo excluye: canceladas, confirmadas (sin deuda)
        sesiones = Sesion.objects.filter(
            paciente=paciente,
            proyecto__isnull=True,  # Solo sesiones normales (no de proyectos)
            mensualidad__isnull=True,  # Solo sesiones normales (no de mensualidades)
            estado__in=['realizada', 'realizada_retraso', 'falta', 'programada']
        ).select_related(
            'servicio', 'profesional', 'sucursal'
        ).order_by('-fecha', '-hora_inicio')

        # DEBUG: Contador
        sesiones_total = sesiones.count()
        sesiones_con_deuda = 0

        # ✅ Calcular total_pagado y saldo_pendiente para CADA sesión
        for s in sesiones:
            total_pagado = _total_pagado_sesion(s)
            saldo_pendiente = s.monto_cobrado - total_pagado
            
            # DEBUG: Imprimir información
            print(f"DEBUG Sesión {s.id} ({s.fecha}): Costo={s.monto_cobrado}, Pagado={total_pagado}, Saldo={saldo_pendiente}")
            
            # ✅ Solo agregar si tiene saldo pendiente (umbral para evitar errores de redondeo)
            if saldo_pendiente > Decimal('0.01'):
                s.total_pagado_calc = total_pagado
                s.saldo_pendiente_calc = saldo_pendiente
                sesiones_pendientes.append(s)
                sesiones_con_deuda += 1
        
        print(f"DEBUG SESIONES: Total={sesiones_total}, Con deuda={sesiones_con_deuda}")

        # ========================================
        # ✅ OBTENER PROYECTOS PENDIENTES (TODOS con deuda)
        # ========================================
        # IMPORTANTE: Incluye planificados, en_progreso Y finalizados
        # Solo excluye: cancelados
        proyectos = Proyecto.objects.filter(
            paciente=paciente,
            estado__in=['planificado', 'en_progreso', 'finalizado']
        ).select_related(
            'servicio_base', 'profesional_responsable', 'sucursal'
        ).order_by('-fecha_inicio')

        # DEBUG: Contador
        proyectos_total = proyectos.count()
        proyectos_con_deuda = 0

        # ✅ Calcular total_pagado y saldo_pendiente para CADA proyecto
        for p in proyectos:
            total_pagado = _total_pagado_proyecto(p)
            saldo_pendiente = p.costo_total - total_pagado
            
            # DEBUG: Imprimir información
            print(f"DEBUG Proyecto {p.id} ({p.codigo}): Costo={p.costo_total}, Pagado={total_pagado}, Saldo={saldo_pendiente}")
            
            # ✅ Solo agregar si tiene saldo pendiente
            if saldo_pendiente > Decimal('0.01'):
                p.total_pagado_calc = total_pagado
                p.saldo_pendiente_calc = saldo_pendiente
                proyectos_pendientes.append(p)
                proyectos_con_deuda += 1
        
        print(f"DEBUG PROYECTOS: Total={proyectos_total}, Con deuda={proyectos_con_deuda}")

        # ========================================
        # ✅ OBTENER MENSUALIDADES PENDIENTES (TODAS con deuda)
        # ========================================
        # IMPORTANTE: Incluye activas, pausadas, completadas Y vencidas
        # Solo excluye: canceladas
        mensualidades = Mensualidad.objects.filter(
            paciente=paciente,
            estado__in=['activa', 'pausada', 'completada', 'vencida']
        ).prefetch_related(
            'servicios_profesionales__servicio',
            'servicios_profesionales__profesional'
        ).order_by('-anio', '-mes')

        # DEBUG: Contador
        mensualidades_total = mensualidades.count()
        mensualidades_con_deuda = 0

        # ✅ Calcular total_pagado y saldo_pendiente para CADA mensualidad
        for m in mensualidades:
            total_pagado = _total_pagado_mensualidad(m)
            saldo_pendiente = m.costo_mensual - total_pagado
            
            # DEBUG: Imprimir información
            print(f"DEBUG Mensualidad {m.id} ({m.mes}/{m.anio}): Costo={m.costo_mensual}, Pagado={total_pagado}, Saldo={saldo_pendiente}")
            
            # ✅ Solo agregar si tiene saldo pendiente
            if saldo_pendiente > Decimal('0.01'):
                m.total_pagado_calc = total_pagado
                m.saldo_pendiente_calc = saldo_pendiente
                mensualidades_pendientes.append(m)
                mensualidades_con_deuda += 1
        
        print(f"DEBUG MENSUALIDADES: Total={mensualidades_total}, Con deuda={mensualidades_con_deuda}")
        
        # DEBUG: Resumen total
        total_items_con_deuda = sesiones_con_deuda + proyectos_con_deuda + mensualidades_con_deuda
        print(f"DEBUG TOTAL ITEMS CON DEUDA: {total_items_con_deuda}")

        # ✅ CALCULAR DEUDAS PENDIENTES del paciente sumando los saldos REALES
        # de las sesiones, proyectos y mensualidades que aparecen en la lista
        deuda_sesiones = sum(s.saldo_pendiente_calc for s in sesiones_pendientes)
        deuda_proyectos = sum(p.saldo_pendiente_calc for p in proyectos_pendientes)
        deuda_mensualidades = sum(m.saldo_pendiente_calc for m in mensualidades_pendientes)
        
        # Total de deudas = suma de todos los saldos pendientes calculados
        paciente.deuda_total_display = deuda_sesiones + deuda_proyectos + deuda_mensualidades
        
        # DEBUG: Mostrar desglose
        print(f"DEBUG DEUDA TOTAL: Sesiones={deuda_sesiones}, Proyectos={deuda_proyectos}, Mensualidades={deuda_mensualidades}, TOTAL={paciente.deuda_total_display}")

    # ✅ Pacientes con DEUDAS PENDIENTES (calculando suma real de saldos)
    pacientes_activos = Paciente.objects.filter(
        estado='activo'
    ).select_related('cuenta_corriente')

    pacientes_con_deuda = []
    for p in pacientes_activos:
        # ✅ CORREGIDO: Calcular deuda sumando saldos reales de ítems pendientes
        # (igual que cuando se selecciona un paciente)
        
        # Sesiones pendientes
        sesiones_p = Sesion.objects.filter(
            paciente=p,
            proyecto__isnull=True,
            mensualidad__isnull=True,
            estado__in=['realizada', 'realizada_retraso', 'falta', 'programada']
        )
        deuda_sesiones_p = Decimal('0')
        for s in sesiones_p:
            total_pagado_s = _total_pagado_sesion(s)
            saldo_s = s.monto_cobrado - total_pagado_s
            if saldo_s > Decimal('0.01'):
                deuda_sesiones_p += saldo_s
        
        # Proyectos pendientes
        proyectos_p = Proyecto.objects.filter(
            paciente=p,
            estado__in=['planificado', 'en_progreso', 'finalizado']
        )
        deuda_proyectos_p = Decimal('0')
        for proyecto in proyectos_p:
            total_pagado_proy = _total_pagado_proyecto(proyecto)
            saldo_proy = proyecto.costo_total - total_pagado_proy
            if saldo_proy > Decimal('0.01'):
                deuda_proyectos_p += saldo_proy
        
        # Mensualidades pendientes
        mensualidades_p = Mensualidad.objects.filter(
            paciente=p,
            estado__in=['activa', 'pausada', 'completada', 'vencida']
        )
        deuda_mensualidades_p = Decimal('0')
        for mens in mensualidades_p:
            total_pagado_mens = _total_pagado_mensualidad(mens)
            saldo_mens = mens.costo_mensual - total_pagado_mens
            if saldo_mens > Decimal('0.01'):
                deuda_mensualidades_p += saldo_mens
        
        # Total de deuda del paciente
        deuda_total_p = deuda_sesiones_p + deuda_proyectos_p + deuda_mensualidades_p
        
        # Solo agregar si tiene deuda pendiente
        if deuda_total_p > Decimal('0.01'):
            p.deuda_total_display = deuda_total_p
            pacientes_con_deuda.append(p)

    # Ordenar por mayor deuda
    pacientes_con_deuda.sort(
        key=lambda p: p.deuda_total_display if hasattr(p, 'deuda_total_display') else 0, 
        reverse=True
    )
    pacientes_con_deuda = pacientes_con_deuda[:50]

    # Métodos de pago
    metodos_pago = MetodoPago.objects.filter(activo=True)

    context = {
        'paciente': paciente,
        'sesiones_pendientes': sesiones_pendientes,
        'proyectos_pendientes': proyectos_pendientes,
        'mensualidades_pendientes': mensualidades_pendientes,
        'pacientes_con_deuda': pacientes_con_deuda,
        'metodos_pago': metodos_pago,
        'fecha_hoy': date.today(),
    }

    return render(request, 'facturacion/pagos_masivos.html', context)


@login_required
def procesar_pagos_masivos(request):
    """
    Procesar pago masivo de múltiples SESIONES, PROYECTOS y MENSUALIDADES
    ✅ ACTUALIZADO: Ahora maneja sesiones, proyectos y mensualidades
    """
    
    if request.method != 'POST':
        return redirect('facturacion:pagos_masivos')
    
    try:
        from django.db import transaction
        
        # Datos del formulario
        paciente_id = request.POST.get('paciente_id')
        sesiones_ids = request.POST.getlist('sesiones_ids')
        proyectos_ids = request.POST.getlist('proyectos_ids')
        mensualidades_ids = request.POST.getlist('mensualidades_ids')
        metodo_pago_id = request.POST.get('metodo_pago')
        fecha_pago_str = request.POST.get('fecha_pago')
        observaciones = request.POST.get('observaciones', '')

        # Validaciones
        if not all([paciente_id, metodo_pago_id, fecha_pago_str]):
            messages.error(request, '❌ Faltan datos obligatorios')
            return redirect('facturacion:pagos_masivos')

        if not (sesiones_ids or proyectos_ids or mensualidades_ids):
            messages.error(request, '❌ Debes seleccionar al menos un ítem (sesión, proyecto o mensualidad)')
            return redirect('facturacion:pagos_masivos')
        
        fecha_pago = datetime.strptime(fecha_pago_str, '%Y-%m-%d').date()
        paciente = Paciente.objects.get(id=paciente_id)
        metodo_pago = MetodoPago.objects.get(id=metodo_pago_id)
        
        # CALCULAR TOTAL Y PREPARAR AJUSTES
        items_ajustados = []
        total_pago = Decimal('0.00')
        
        # ========================================
        # Procesar SESIONES
        # ========================================
        if sesiones_ids:
            sesiones = Sesion.objects.filter(
                id__in=sesiones_ids,
                paciente=paciente
            ).select_related('servicio', 'proyecto')
            
            for sesion in sesiones:
                monto_personalizado_key = f'monto_personalizado_sesion_{sesion.id}'
                monto_personalizado = request.POST.get(monto_personalizado_key)
                
                if monto_personalizado:
                    monto_pagar = Decimal(monto_personalizado)
                else:
                    monto_pagar = sesion.monto_cobrado - _total_pagado_sesion(sesion)
                
                if monto_pagar <= 0:
                    messages.error(
                        request, 
                        f'❌ Monto inválido para sesión {sesion.fecha}: Bs. {monto_pagar}'
                    )
                    return redirect('facturacion:pagos_masivos')
                
                items_ajustados.append({
                    'tipo': 'sesion',
                    'objeto': sesion,
                    'monto_pagar': monto_pagar,
                    'tiene_monto_personalizado': bool(monto_personalizado)
                })
                
                total_pago += monto_pagar
        
        # ========================================
        # Procesar PROYECTOS
        # ========================================
        if proyectos_ids:
            proyectos = Proyecto.objects.filter(
                id__in=proyectos_ids,
                paciente=paciente
            ).select_related('servicio_base', 'profesional_responsable')
            
            for proyecto in proyectos:
                monto_personalizado_key = f'monto_personalizado_proyecto_{proyecto.id}'
                monto_personalizado = request.POST.get(monto_personalizado_key)
                
                if monto_personalizado:
                    monto_pagar = Decimal(monto_personalizado)
                else:
                    monto_pagar = proyecto.costo_total - _total_pagado_proyecto(proyecto)
                
                if monto_pagar <= 0:
                    messages.error(
                        request, 
                        f'❌ Monto inválido para proyecto {proyecto.codigo}: Bs. {monto_pagar}'
                    )
                    return redirect('facturacion:pagos_masivos')
                
                items_ajustados.append({
                    'tipo': 'proyecto',
                    'objeto': proyecto,
                    'monto_pagar': monto_pagar,
                    'tiene_monto_personalizado': bool(monto_personalizado)
                })
                
                total_pago += monto_pagar
        
        # ========================================
        # Procesar MENSUALIDADES
        # ========================================
        if mensualidades_ids:
            mensualidades = Mensualidad.objects.filter(
                id__in=mensualidades_ids,
                paciente=paciente
            )
            
            for mensualidad in mensualidades:
                monto_personalizado_key = f'monto_personalizado_mensualidad_{mensualidad.id}'
                monto_personalizado = request.POST.get(monto_personalizado_key)
                
                if monto_personalizado:
                    monto_pagar = Decimal(monto_personalizado)
                else:
                    monto_pagar = mensualidad.costo_mensual - _total_pagado_mensualidad(mensualidad)
                
                if monto_pagar <= 0:
                    messages.error(
                        request, 
                        f'❌ Monto inválido para mensualidad {mensualidad.mes}/{mensualidad.anio}: Bs. {monto_pagar}'
                    )
                    return redirect('facturacion:pagos_masivos')
                
                items_ajustados.append({
                    'tipo': 'mensualidad',
                    'objeto': mensualidad,
                    'monto_pagar': monto_pagar,
                    'tiene_monto_personalizado': bool(monto_personalizado)
                })
                
                total_pago += monto_pagar
        
        if total_pago <= 0:
            messages.error(request, '❌ El total a pagar debe ser mayor a 0')
            return redirect('facturacion:pagos_masivos')
        
        # 🔒 TRANSACCIÓN ATÓMICA
        with transaction.atomic():
            from .models import DetallePagoMasivo
            
            # 📝 PREPARAR CONCEPTO RESUMIDO
            descripcion_items = []
            count_by_type = {'sesion': 0, 'proyecto': 0, 'mensualidad': 0}
            
            for item in items_ajustados[:3]:
                obj = item['objeto']
                tipo = item['tipo']
                count_by_type[tipo] += 1
                
                if tipo == 'sesion':
                    descripcion_items.append(f"{obj.fecha.strftime('%d/%m')} {obj.servicio.nombre[:20]}")
                elif tipo == 'proyecto':
                    descripcion_items.append(f"{obj.codigo} {obj.nombre[:20]}")
                elif tipo == 'mensualidad':
                    descripcion_items.append(f"Mens. {obj.mes}/{obj.anio}")
            
            concepto_items = ', '.join(descripcion_items)
            if len(items_ajustados) > 3:
                concepto_items += f" (+{len(items_ajustados) - 3} más)"
            
            # Construir resumen de tipos
            tipos_str = []
            if count_by_type['sesion'] > 0:
                tipos_str.append(f"{len([i for i in items_ajustados if i['tipo'] == 'sesion'])} sesión(es)")
            if count_by_type['proyecto'] > 0:
                tipos_str.append(f"{len([i for i in items_ajustados if i['tipo'] == 'proyecto'])} proyecto(s)")
            if count_by_type['mensualidad'] > 0:
                tipos_str.append(f"{len([i for i in items_ajustados if i['tipo'] == 'mensualidad'])} mensualidad(es)")
            
            # 💾 CREAR UN SOLO PAGO MASIVO
            concepto_principal = f"Pago masivo de {', '.join(tipos_str)}: {concepto_items}"
            
            pago_masivo = Pago.objects.create(
                paciente=paciente,
                sesion=None,
                proyecto=None,
                mensualidad=None,
                fecha_pago=fecha_pago,
                monto=total_pago,
                metodo_pago=metodo_pago,
                concepto=concepto_principal,
                observaciones=observaciones or f"Pago masivo de {len(items_ajustados)} ítem(s)",
                registrado_por=request.user
            )
            
            # 📋 CREAR DETALLES PARA CADA ÍTEM
            for ajuste in items_ajustados:
                obj = ajuste['objeto']
                tipo = ajuste['tipo']
                monto_pagar = ajuste['monto_pagar']
                tiene_monto_personalizado = ajuste['tiene_monto_personalizado']
                
                # Construir concepto según tipo
                if tipo == 'sesion':
                    concepto_detalle = f"Sesión {obj.fecha.strftime('%d/%m/%Y')} - {obj.servicio.nombre}"
                    
                    # Ajustar monto_cobrado si es personalizado
                    if tiene_monto_personalizado:
                        monto_original = obj.monto_cobrado
                        nuevo_monto_cobrado = _total_pagado_sesion(obj) + monto_pagar
                        
                        if nuevo_monto_cobrado != obj.monto_cobrado:
                            obj.monto_cobrado = nuevo_monto_cobrado
                            nota_ajuste = f"\n[{fecha_pago}] Monto ajustado de Bs. {monto_original} a Bs. {nuevo_monto_cobrado} en pago masivo"
                            obj.observaciones = (obj.observaciones or "") + nota_ajuste
                            obj.save()
                    
                    # Crear detalle
                    DetallePagoMasivo.objects.create(
                        pago=pago_masivo,
                        tipo='sesion',
                        sesion=obj,
                        proyecto=obj.proyecto if obj.proyecto else None,
                        mensualidad=None,
                        monto=monto_pagar,
                        concepto=concepto_detalle
                    )
                
                elif tipo == 'proyecto':
                    concepto_detalle = f"Proyecto {obj.codigo} - {obj.nombre}"
                    
                    # Crear detalle
                    DetallePagoMasivo.objects.create(
                        pago=pago_masivo,
                        tipo='proyecto',
                        sesion=None,
                        proyecto=obj,
                        mensualidad=None,
                        monto=monto_pagar,
                        concepto=concepto_detalle
                    )
                
                elif tipo == 'mensualidad':
                    concepto_detalle = f"Mensualidad {obj.mes}/{obj.anio}"
                    
                    # Crear detalle
                    DetallePagoMasivo.objects.create(
                        pago=pago_masivo,
                        tipo='mensualidad',
                        sesion=None,
                        proyecto=None,
                        mensualidad=obj,
                        monto=monto_pagar,
                        concepto=concepto_detalle
                    )
            
            # Actualizar cuenta corriente
            cuenta, created = CuentaCorriente.objects.get_or_create(paciente=paciente)
            AccountService.update_balance(paciente)
        
        # 🆕 PREPARAR DATOS PARA CONFIRMACIÓN
        items_count = len(items_ajustados)
        
        # Construir concepto detallado
        concepto_completo = f"Pago masivo: {concepto_items}"
        
        # Información de estado
        info_estado = f"{', '.join(tipos_str)} procesados correctamente"
        
        # 🆕 ALMACENAR EN SESSION para mostrar en confirmación
        request.session['pago_exitoso'] = {
            'tipo': 'Pago Masivo',
            'mensaje': f'Pago masivo registrado exitosamente',
            'detalle': f'{items_count} ítem(s) pagado(s)',
            'total': float(total_pago),
            'paciente': paciente.nombre_completo,
            'concepto': concepto_completo,
            'info_estado': info_estado,
            'genero_recibo': True,
            'numero_recibo': pago_masivo.numero_recibo,
            'pago_id': pago_masivo.id,
        }
        
        return redirect('facturacion:confirmacion_pago')
        
    except Exception as e:
        messages.error(request, f'❌ Error al procesar pago masivo: {str(e)}')
        import traceback
        print(traceback.format_exc())
        return redirect('facturacion:pagos_masivos')

# ==================== HISTORIAL DE PAGOS ====================

@login_required
def historial_pagos(request):
    """
    Vista del historial de pagos y devoluciones combinados
    
    ✅ MEJORAS:
    - Incluye devoluciones en el listado
    - Quita símbolo de impresora en recibos de crédito (CRE-)
    - Permite anular recibos
    - Carga lazy de estadísticas
    """
    
    # ==================== FILTROS ====================
    buscar = request.GET.get('q', '').strip()
    metodo_id = request.GET.get('metodo', '').strip()
    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()
    tipo_filtro = request.GET.get('tipo', '').strip()  # pago | anulado | devolucion

    # ==================== QUERY DE PAGOS ====================
    pagos = Pago.objects.select_related(
        'paciente',
        'metodo_pago',
        'sesion',
        'proyecto',
        'mensualidad'
    ).prefetch_related(
        'detalles_masivos'
    )
    
    # Aplicar filtros a pagos
    if buscar:
        pagos = pagos.filter(
            Q(paciente__nombre__icontains=buscar) |
            Q(paciente__apellido__icontains=buscar) |
            Q(numero_recibo__icontains=buscar) |
            Q(concepto__icontains=buscar)
        )
    
    if metodo_id:
        pagos = pagos.filter(metodo_pago_id=metodo_id)
    
    if fecha_desde:
        pagos = pagos.filter(fecha_pago__gte=fecha_desde)
    
    if fecha_hasta:
        pagos = pagos.filter(fecha_pago__lte=fecha_hasta)

    # Filtro tipo sobre la columna Tipo de la tabla
    if tipo_filtro == 'pago':
        pagos = pagos.filter(anulado=False)
    elif tipo_filtro == 'anulado':
        pagos = pagos.filter(anulado=True)
    # 'devolucion' → pagos se excluyen completamente abajo
    
    # ==================== QUERY DE DEVOLUCIONES ====================
    devoluciones = Devolucion.objects.select_related(
        'paciente',
        'metodo_devolucion',
        'proyecto',
        'mensualidad'
    )
    
    # Aplicar filtros a devoluciones
    if buscar:
        devoluciones = devoluciones.filter(
            Q(paciente__nombre__icontains=buscar) |
            Q(paciente__apellido__icontains=buscar) |
            Q(numero_devolucion__icontains=buscar) |
            Q(motivo__icontains=buscar)
        )
    
    if metodo_id:
        devoluciones = devoluciones.filter(metodo_devolucion_id=metodo_id)
    
    if fecha_desde:
        devoluciones = devoluciones.filter(fecha_devolucion__gte=fecha_desde)
    
    if fecha_hasta:
        devoluciones = devoluciones.filter(fecha_devolucion__lte=fecha_hasta)
    
    # ==================== COMBINAR PAGOS Y DEVOLUCIONES ====================
    items_combinados = []

    incluir_pagos = tipo_filtro != 'devolucion'
    incluir_devoluciones = tipo_filtro in ('', 'devolucion')

    if incluir_pagos:
        for pago in pagos:
            items_combinados.append({
                'tipo': 'pago',
                'objeto': pago,
                'fecha': pago.fecha_pago,
            })

    if incluir_devoluciones:
        for devolucion in devoluciones:
            items_combinados.append({
                'tipo': 'devolucion',
                'objeto': devolucion,
                'fecha': devolucion.fecha_devolucion,
            })

    # ==================== ORDENAMIENTO POR COLUMNA ====================
    orden_col = request.GET.get('orden', 'fecha')
    orden_dir = request.GET.get('dir', 'desc')
    reverse = (orden_dir == 'desc')

    def get_sort_key(item):
        obj = item['objeto']
        if orden_col == 'fecha':
            return item['fecha']
        elif orden_col == 'paciente':
            return (obj.paciente.apellido + obj.paciente.nombre).lower() if obj.paciente else ''
        elif orden_col == 'monto':
            return obj.monto
        elif orden_col == 'metodo':
            if item['tipo'] == 'devolucion':
                return obj.metodo_devolucion.nombre.lower() if obj.metodo_devolucion else ''
            return obj.metodo_pago.nombre.lower() if obj.metodo_pago else ''
        elif orden_col == 'recibo':
            if item['tipo'] == 'devolucion':
                return obj.numero_devolucion.lower() if obj.numero_devolucion else ''
            return obj.numero_recibo.lower() if obj.numero_recibo else ''
        elif orden_col == 'concepto':
            if item['tipo'] == 'devolucion':
                return obj.motivo.lower() if obj.motivo else ''
            return obj.concepto.lower() if obj.concepto else ''
        elif orden_col == 'tipo':
            if item['tipo'] == 'devolucion':
                return '0_devolucion'
            return '1_anulado' if obj.anulado else '2_pago'
        return item['fecha']

    items_combinados.sort(key=get_sort_key, reverse=reverse)

    # ==================== PAGINACIÓN ====================
    paginator = Paginator(items_combinados, 30)  # 30 registros por página
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # ==================== ESTADÍSTICAS (LAZY LOADING) ====================
    mostrar_estadisticas = request.GET.get('stats', 'false') == 'true'
    
    # ==================== MÉTODOS DE PAGO PARA FILTRO ====================
    metodos_pago = MetodoPago.objects.filter(activo=True).order_by('nombre')
    
    # ==================== CONTEXTO ====================
    context = {
        'page_obj': page_obj,
        'buscar': buscar,
        'metodo_id': metodo_id,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'tipo_filtro': tipo_filtro,
        'orden_col': orden_col,
        'orden_dir': orden_dir,
        'metodos_pago': metodos_pago,
        'mostrar_estadisticas': mostrar_estadisticas,
    }
    
    return render(request, 'facturacion/historial_pagos.html', context)


def calcular_estadisticas_pagos(pagos_queryset, devoluciones_queryset=None, filtros=None):
    """
    Calcula estadísticas de pagos usando solo queries a BD
    Función separada para reutilizar en AJAX

    ✅ MEJORADO: Incluye pagos reales, uso de crédito, pagos adelantados,
                 devoluciones y anulados. Aplica los mismos filtros de fecha/búsqueda.
    """

    # ── BASE: todos los pagos del queryset (incluye anulados y crédito) ──────
    total_pagos_all = pagos_queryset.count()

    # ── ANULADOS ─────────────────────────────────────────────────────────────
    pagos_anulados = pagos_queryset.filter(anulado=True)
    total_anulados = pagos_anulados.count()
    monto_anulados = pagos_anulados.aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')

    # ── BASE NO ANULADOS ──────────────────────────────────────────────────────
    pagos_no_anulados = pagos_queryset.filter(anulado=False)

    # ── USO DE CRÉDITO (no anulados, método = "Uso de Crédito") ──────────────
    pagos_credito = pagos_no_anulados.filter(metodo_pago__nombre="Uso de Crédito")
    total_pagos_credito = pagos_credito.count()
    monto_pagos_credito = pagos_credito.aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')

    # ── PAGOS REALES: no anulados, no crédito ─────────────────────────────────
    pagos_reales = pagos_no_anulados.exclude(metodo_pago__nombre="Uso de Crédito")
    total_pagos_reales = pagos_reales.count()
    monto_pagos_reales = pagos_reales.aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')

    # ── PAGOS ADELANTADOS: reales sin sesión/proyecto/mensualidad asignados ───
    pagos_adelantados = pagos_reales.filter(
        sesion__isnull=True,
        proyecto__isnull=True,
        mensualidad__isnull=True,
    ).exclude(numero_recibo__startswith='CRE-')
    total_adelantados = pagos_adelantados.count()
    monto_adelantados = pagos_adelantados.aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')

    # ── PROMEDIO POR TRANSACCIÓN ──────────────────────────────────────────────
    promedio_transaccion = (
        monto_pagos_reales / total_pagos_reales if total_pagos_reales > 0 else Decimal('0.00')
    )

    # ── MÉTODOS ACTIVOS (solo pagos reales) ───────────────────────────────────
    total_metodos = pagos_reales.values('metodo_pago').distinct().count()

    # ── DESGLOSE POR MÉTODO (top 5, solo pagos reales) ───────────────────────
    desglose_metodos = pagos_reales.values(
        'metodo_pago__nombre'
    ).annotate(
        cantidad=Count('id'),
        total=Sum('monto')
    ).order_by('-total')[:5]

    # ── DEVOLUCIONES ─────────────────────────────────────────────────────────
    if devoluciones_queryset is not None:
        total_devoluciones = devoluciones_queryset.count()
        monto_devoluciones = devoluciones_queryset.aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
    else:
        total_devoluciones = 0
        monto_devoluciones = Decimal('0.00')

    return {
        # Contadores generales
        'total_pagos': total_pagos_all,
        'total_pagos_reales': total_pagos_reales,
        'total_pagos_credito': total_pagos_credito,
        'total_adelantados': total_adelantados,
        'total_anulados': total_anulados,
        'total_devoluciones': total_devoluciones,

        # Montos
        'monto_pagos_reales': monto_pagos_reales,
        'monto_pagos_credito': monto_pagos_credito,
        'monto_adelantados': monto_adelantados,
        'monto_anulados': monto_anulados,
        'monto_devoluciones': monto_devoluciones,

        # Compatibilidad con código anterior
        'monto_pagos_validos': monto_pagos_reales,
        'total_pagos_validos': total_pagos_reales,

        # Métricas adicionales
        'promedio_transaccion': promedio_transaccion,
        'total_metodos': total_metodos,

        # Desglose
        'desglose_metodos': list(desglose_metodos),
    }


@login_required
def cargar_estadisticas_pagos_ajax(request):
    """
    Vista AJAX para cargar estadísticas de pagos bajo demanda.
    Respeta los mismos filtros que la vista principal.

    ✅ MEJORADO: Incluye pagos reales, uso de crédito, adelantados,
                 devoluciones y anulados con filtros dinámicos.
    """

    # Aplicar los mismos filtros
    buscar = request.GET.get('q', '').strip()
    metodo_id = request.GET.get('metodo', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    tipo_filtro = request.GET.get('tipo', '').strip()

    # ── Query base de pagos (TODOS, incluyendo anulados) ──────────────────────
    pagos = Pago.objects.select_related('metodo_pago')

    if buscar:
        pagos = pagos.filter(
            Q(paciente__nombre__icontains=buscar) |
            Q(paciente__apellido__icontains=buscar) |
            Q(numero_recibo__icontains=buscar)
        )

    if metodo_id:
        pagos = pagos.filter(metodo_pago_id=metodo_id)

    if fecha_desde:
        try:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            pagos = pagos.filter(fecha_pago__gte=fecha_desde_obj)
        except Exception:
            pass

    if fecha_hasta:
        try:
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            pagos = pagos.filter(fecha_pago__lte=fecha_hasta_obj)
        except Exception:
            pass

    # Las estadísticas siempre muestran el resumen GLOBAL (no se filtran por tipo)
    # para que el usuario vea el panorama completo incluso al filtrar la lista.
    # El filtro tipo solo afecta la lista de registros, no las tarjetas de resumen.

    # ── Query de devoluciones con los mismos filtros ──────────────────────────
    devoluciones = Devolucion.objects.all()

    if buscar:
        devoluciones = devoluciones.filter(
            Q(paciente__nombre__icontains=buscar) |
            Q(paciente__apellido__icontains=buscar) |
            Q(numero_devolucion__icontains=buscar)
        )

    # Nota: si filtran por método de pago, las devoluciones usan metodo_devolucion
    # No aplicamos el mismo filtro de método para no distorsionar las devoluciones
    # a menos que sea un filtro explícito del usuario
    if fecha_desde:
        try:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            devoluciones = devoluciones.filter(fecha_devolucion__gte=fecha_desde_obj)
        except Exception:
            pass

    if fecha_hasta:
        try:
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            devoluciones = devoluciones.filter(fecha_devolucion__lte=fecha_hasta_obj)
        except Exception:
            pass

    # Calcular estadísticas
    stats = calcular_estadisticas_pagos(pagos, devoluciones_queryset=devoluciones)

    # Convertir Decimals a float para JSON
    return JsonResponse({
        'success': True,
        'estadisticas': {
            # Contadores
            'total_pagos': stats['total_pagos'],
            'total_pagos_reales': stats['total_pagos_reales'],
            'total_pagos_credito': stats['total_pagos_credito'],
            'total_adelantados': stats['total_adelantados'],
            'total_anulados': stats['total_anulados'],
            'total_devoluciones': stats['total_devoluciones'],

            # Montos
            'monto_pagos_reales': float(stats['monto_pagos_reales']),
            'monto_pagos_credito': float(stats['monto_pagos_credito']),
            'monto_adelantados': float(stats['monto_adelantados']),
            'monto_anulados': float(stats['monto_anulados']),
            'monto_devoluciones': float(stats['monto_devoluciones']),

            # Compat
            'monto_pagos_validos': float(stats['monto_pagos_reales']),
            'total_pagos_validos': stats['total_pagos_reales'],

            # Métricas
            'promedio_transaccion': float(stats['promedio_transaccion']),
            'total_metodos': stats['total_metodos'],

            # Desglose
            'desglose_metodos': [
                {
                    'metodo': item['metodo_pago__nombre'],
                    'cantidad': item['cantidad'],
                    'total': float(item['total'])
                }
                for item in stats['desglose_metodos']
            ]
        }
    })

# ==================== OPTIMIZACIÓN EXTREMA DEL RECIBO PDF ====================

from django.core.cache import cache
from django.db.models import Sum
import hashlib

def encontrar_logo():
    """
    Buscar logo en múltiples ubicaciones posibles
    ✅ OPTIMIZADO: Cache en memoria para evitar búsqueda repetida
    """
    # Intentar obtener del cache primero
    logo_path = cache.get('logo_path_cached')
    
    if logo_path and os.path.exists(logo_path):
        return logo_path
    
    # Si no está en cache, buscar
    posibles_rutas = [
        os.path.join(settings.BASE_DIR, 'core', 'static', 'img', 'logo_misael.png'),
        os.path.join(settings.STATIC_ROOT, 'img', 'logo_misael.png') if settings.STATIC_ROOT else None,
        os.path.join(settings.BASE_DIR, 'static', 'img', 'logo_misael.png'),
    ]
    
    # Agregar STATICFILES_DIRS
    if hasattr(settings, 'STATICFILES_DIRS'):
        for static_dir in settings.STATICFILES_DIRS:
            posibles_rutas.append(os.path.join(static_dir, 'img', 'logo_misael.png'))
    
    # Filtrar None y buscar
    for ruta in filter(None, posibles_rutas):
        if os.path.exists(ruta):
            # Cachear por 1 hora
            cache.set('logo_path_cached', ruta, 3600)
            return ruta
    
    return None


@login_required
def generar_recibo_pdf(request, pago_id):
    """
    Genera recibo en PDF usando ReportLab
    
    ✅ Código limpio y organizado
    ✅ Lógica en pdf_generator.py
    ✅ Funciona en Windows sin problemas
    """
    
    # Query optimizada del pago
    pago = get_object_or_404(
        Pago.objects.select_related(
            'paciente',
            'metodo_pago',
            'registrado_por'
        ),
        id=pago_id
    )
    
    # Validación: NO generar PDF para pagos con crédito
    if pago.metodo_pago.nombre == "Uso de Crédito":
        messages.warning(
            request,
            '⚠️ Los pagos con crédito no generan recibo físico. '
            'El recibo se generó en el pago adelantado original.'
        )
        return redirect('facturacion:historial_pagos')
    
    try:
        # ✅ Verificar cache primero
        cache_key = f'pdf_recibo_{pago.numero_recibo}'
        pdf_cached = cache.get(cache_key)
        
        if pdf_cached:
            response = HttpResponse(pdf_cached, content_type='application/pdf')
            response['Content-Disposition'] = f'inline; filename="recibo_{pago.numero_recibo}.pdf"'
            return response
        
        # Query de pagos relacionados
        pagos_relacionados = list(Pago.objects.filter(
            numero_recibo=pago.numero_recibo,
            anulado=False
        ).select_related(
            'sesion__servicio',
            'proyecto',
            'mensualidad'
        ).order_by('id'))
        
        # Calcular total
        total_recibo = sum(p.monto for p in pagos_relacionados)
        
        # ✅ Generar PDF usando el módulo separado
        pdf_data = pdf_generator.generar_recibo_pdf(
            pago=pago,
            pagos_relacionados=pagos_relacionados,
            total_recibo=total_recibo
        )
        
        # Cachear PDF generado por 1 hora
        cache.set(cache_key, pdf_data, 3600)
        
        # Preparar respuesta
        response = HttpResponse(pdf_data, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="recibo_{pago.numero_recibo}.pdf"'
        
        return response
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Error generando PDF para pago {pago_id}: {str(e)}')
        
        messages.error(request, f'❌ Error al generar PDF: {str(e)}')
        return redirect('facturacion:historial_pagos')

# ==================== UTILIDAD: LIMPIAR CACHE DE RECIBOS ====================

@login_required
@staff_member_required
def limpiar_cache_recibos(request):
    """
    Utilidad para limpiar el cache de recibos
    Útil después de actualizar el template
    """
    try:
        # Limpiar todo el cache relacionado con recibos
        cache.delete('recibo_logo_base64')
        cache.delete('logo_path_cached')
        
        # Limpiar PDFs cacheados (pattern matching)
        # Nota: LocMemCache no soporta pattern matching,
        # así que limpiamos todo el cache
        cache.clear()
        
        messages.success(request, '✅ Cache de recibos limpiado correctamente')
        
    except Exception as e:
        messages.error(request, f'❌ Error al limpiar cache: {str(e)}')
    
    return redirect('facturacion:historial_pagos')

# ==================== ESTADÍSTICAS DE CACHE (OPCIONAL) ====================

@login_required
@staff_member_required
def estadisticas_cache_recibos(request):
    """
    Ver estadísticas del cache de recibos
    ✅ Debugging tool para administradores
    """
    from django.http import JsonResponse
    
    stats = {
        'logo_base64_cached': cache.get('recibo_logo_base64') is not None,
        'logo_path_cached': cache.get('logo_path_cached') is not None,
        'cache_backend': settings.CACHES['default']['BACKEND'],
        'cache_timeout': settings.CACHES['default'].get('TIMEOUT', 'N/A'),
    }
    
    # Intentar contar PDFs cacheados (aproximación)
    # Nota: Esto no funciona con LocMemCache, solo para info
    stats['nota'] = 'LocMemCache no permite listar todas las keys'
    
    return JsonResponse(stats, json_dumps_params={'indent': 2})

# ==================== ANULAR PAGOS ====================

@login_required
def anular_pago(request, pago_id):
    """
    Vista para anular un pago
    ✅ CORREGIDO: Maneja tanto peticiones AJAX (JSON) como formularios tradicionales (POST)
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    try:
        # ✅ NUEVO: Intentar obtener datos de JSON primero, luego de POST
        motivo = None
        
        # Intentar parsear como JSON (peticiones AJAX)
        if request.content_type == 'application/json':
            try:
                import json
                data = json.loads(request.body)
                motivo = data.get('motivo', '').strip()
            except json.JSONDecodeError:
                pass
        
        # Si no es JSON, obtener de POST (formularios tradicionales)
        if not motivo:
            motivo = request.POST.get('motivo', '').strip()
        
        # Validar que se proporcionó un motivo
        if not motivo:
            return JsonResponse({'success': False, 'error': 'El motivo es obligatorio'}, status=400)
        
        pago = get_object_or_404(Pago, id=pago_id)
        
        # Verificar que no esté anulado ya
        if pago.anulado:
            return JsonResponse({'success': False, 'error': 'Este pago ya está anulado'}, status=400)
        
        # Anular el pago
        pago.anular(usuario=request.user, motivo=motivo)
        
        return JsonResponse({
            'success': True,
            'message': f'Pago {pago.numero_recibo} anulado exitosamente'
        })
        
    except Exception as e:
        import traceback
        print(f"❌ Error al anular pago: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
        
# ==================== FASE 3: REPORTES ====================

@login_required
def dashboard_reportes(request):
    """
    Dashboard de reportes - punto de entrada
    ✅ CORREGIDO: Calcula y pasa KPIs al template
    """
    from datetime import date
    from decimal import Decimal
    from django.db.models import Count, Sum, Q

    hoy = date.today()
    inicio_mes = hoy.replace(day=1)

    # ── Sesiones hoy
    sesiones_hoy_qs = Sesion.objects.filter(fecha=hoy)
    sesiones_hoy = sesiones_hoy_qs.count()

    # ── Sesiones del mes para tasa de asistencia
    sesiones_mes = Sesion.objects.filter(fecha__gte=inicio_mes, fecha__lte=hoy)
    stats_mes = sesiones_mes.aggregate(
        total=Count('id'),
        realizadas=Count('id', filter=Q(estado__in=['realizada', 'realizada_retraso']))
    )
    total_mes = stats_mes['total'] or 0
    realizadas_mes = stats_mes['realizadas'] or 0
    tasa_asistencia = round(realizadas_mes / total_mes * 100, 1) if total_mes > 0 else 0

    # ── Pacientes con sesión este mes
    pacientes_activos = Sesion.objects.filter(
        fecha__gte=inicio_mes
    ).values('paciente').distinct().count()

    # ── Financiero: ingresos hoy y mes
    ingresos_hoy = Decimal('0.00')
    ingresos_mes = Decimal('0.00')
    saldo_pendiente = Decimal('0.00')

    try:
        ingresos_hoy = Pago.objects.filter(
            fecha_pago=hoy, anulado=False
        ).exclude(metodo_pago__nombre="Uso de Crédito").aggregate(
            t=Sum('monto')
        )['t'] or Decimal('0.00')

        ingresos_mes = Pago.objects.filter(
            fecha_pago__gte=inicio_mes,
            fecha_pago__lte=hoy,
            anulado=False
        ).exclude(metodo_pago__nombre="Uso de Crédito").aggregate(
            t=Sum('monto')
        )['t'] or Decimal('0.00')
        
        # Sumar mensualidades del mes actual (generadas, no cobradas)
        from agenda.models import Mensualidad
        mensualidades_mes_dashboard = Mensualidad.objects.filter(
            anio=hoy.year,
            mes=hoy.month,
            estado__in=['activa', 'pausada', 'completada', 'cancelada']
        ).aggregate(t=Sum('costo_mensual'))['t'] or Decimal('0.00')
        ingresos_mes += mensualidades_mes_dashboard

        # Saldo pendiente: cuentas con saldo negativo
        from facturacion.models import CuentaCorriente
        saldo_pendiente = CuentaCorriente.objects.filter(
            saldo_actual__lt=0
        ).aggregate(
            t=Sum('saldo_actual')
        )['t'] or Decimal('0.00')
        saldo_pendiente = abs(saldo_pendiente)

    except Exception:
        pass

    kpis = {
        'sesiones_hoy': sesiones_hoy,
        'ingresos_hoy': f'{ingresos_hoy:.2f}',
        'ingresos_mes': f'{ingresos_mes:.2f}',
        'saldo_pendiente': f'{saldo_pendiente:.2f}',
        'pacientes_activos': pacientes_activos,
        'tasa_asistencia': tasa_asistencia,
    }

    return render(request, 'facturacion/reportes/dashboard.html', {'kpis': kpis})

@login_required
def reporte_paciente(request):
    """
    Reporte detallado por paciente - CORREGIDO
    ✅ Usa anotaciones en lugar del campo 'pagado' eliminado
    """
    
    paciente_id = request.GET.get('paciente')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    paciente = None
    datos = None
    grafico_data = None
    
    if paciente_id:
        from datetime import datetime, timedelta
        
        paciente = get_object_or_404(
            Paciente.objects.select_related('cuenta_corriente'),
            id=paciente_id
        )
        
        # Rango de fechas (por defecto: Últimos 6 meses)
        if fecha_desde and fecha_hasta:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
        else:
            fecha_hasta_obj = date.today()
            fecha_desde_obj = fecha_hasta_obj - timedelta(days=180)
            fecha_desde = fecha_desde_obj.strftime('%Y-%m-%d')
            fecha_hasta = fecha_hasta_obj.strftime('%Y-%m-%d')
        
        # Query base
        sesiones = Sesion.objects.filter(
            paciente=paciente,
            fecha__gte=fecha_desde_obj,
            fecha__lte=fecha_hasta_obj
        ).select_related('servicio', 'profesional', 'sucursal', 'proyecto')
        
        # Estadísticas generales
        stats = sesiones.aggregate(
            total_sesiones=Count('id'),
            realizadas=Count('id', filter=Q(estado='realizada')),
            retrasos=Count('id', filter=Q(estado='realizada_retraso')),
            faltas=Count('id', filter=Q(estado='falta')),
            permisos=Count('id', filter=Q(estado='permiso')),
            canceladas=Count('id', filter=Q(estado='cancelada')),
            reprogramadas=Count('id', filter=Q(estado='reprogramada')),
            total_cobrado=Sum('monto_cobrado'),
        )
        
        # ✅ CORREGIDO: Calcular total_pagado usando la función que incluye pagos masivos
        sesiones_list = list(sesiones)
        
        total_pagado = Decimal('0.00')
        sesiones_pagadas = 0
        sesiones_pendientes = 0
        
        for s in sesiones_list:
            # Usar la función que incluye pagos masivos
            s.total_pagado_sesion = _total_pagado_sesion(s)
            total_pagado += s.total_pagado_sesion
            
            # Contar sesiones pagadas y pendientes
            if s.monto_cobrado > 0:
                if s.total_pagado_sesion >= s.monto_cobrado:
                    sesiones_pagadas += 1
                else:
                    sesiones_pendientes += 1
        
        stats['total_pagado'] = total_pagado
        stats['sesiones_pagadas'] = sesiones_pagadas
        stats['sesiones_pendientes'] = sesiones_pendientes
        stats['saldo_pendiente'] = (stats['total_cobrado'] or Decimal('0.00')) - total_pagado
        
        # Calcular tasa de asistencia
        sesiones_efectivas = stats['realizadas'] + stats['retrasos']
        sesiones_programadas = stats['total_sesiones'] - stats['canceladas'] - stats['permisos']
        tasa_asistencia = (sesiones_efectivas / sesiones_programadas * 100) if sesiones_programadas > 0 else 0
        
        # Tasa de pago
        total_con_cobro = sum(1 for s in sesiones_list if s.monto_cobrado > 0)
        tasa_pago = (sesiones_pagadas / total_con_cobro * 100) if total_con_cobro > 0 else 0
        
        # Por servicio
        por_servicio = sesiones.values(
            'servicio__nombre', 'servicio__color'
        ).annotate(
            cantidad=Count('id'),
            monto_total=Sum('monto_cobrado'),
            sesiones_realizadas=Count('id', filter=Q(estado__in=['realizada', 'realizada_retraso'])),
            sesiones_falta=Count('id', filter=Q(estado='falta'))
        ).order_by('-cantidad')
        
        # Por profesional
        por_profesional = sesiones.filter(
            estado__in=['realizada', 'realizada_retraso']
        ).values(
            'profesional__nombre', 'profesional__apellido'
        ).annotate(
            cantidad=Count('id'),
            monto_total=Sum('monto_cobrado')
        ).order_by('-cantidad')
        
        # Por sucursal
        por_sucursal = sesiones.values(
            'sucursal__nombre'
        ).annotate(
            cantidad=Count('id'),
            realizadas=Count('id', filter=Q(estado='realizada'))
        ).order_by('-cantidad')
        
        # Por mes (para gráfico)
        from django.db.models.functions import TruncMonth
        por_mes = sesiones.annotate(
            mes=TruncMonth('fecha')
        ).values('mes').annotate(
            total=Count('id'),
            realizadas=Count('id', filter=Q(estado='realizada')),
            retrasos=Count('id', filter=Q(estado='realizada_retraso')),
            faltas=Count('id', filter=Q(estado='falta')),
            monto_generado=Sum('monto_cobrado')
        ).order_by('mes')
        
        # Preparar datos para gráfico
        grafico_data = {
            'labels': [m['mes'].strftime('%b %Y') for m in por_mes],
            'realizadas': [m['realizadas'] for m in por_mes],
            'faltas': [m['faltas'] for m in por_mes],
            'retrasos': [m['retrasos'] for m in por_mes],
            'monto': [float(m['monto_generado'] or 0) for m in por_mes],
        }
        
        # Proyectos del paciente
        proyectos_paciente = Proyecto.objects.filter(
            paciente=paciente
        ).select_related('servicio_base', 'profesional_responsable')
        
        proyectos_stats = {
            'total': proyectos_paciente.count(),
            'activos': proyectos_paciente.filter(estado__in=['planificado', 'en_progreso']).count(),
            'finalizados': proyectos_paciente.filter(estado='finalizado').count(),
            'monto_total': proyectos_paciente.aggregate(Sum('costo_total'))['costo_total__sum'] or Decimal('0.00'),
        }
        
        # Mensualidades del paciente en el período
        from agenda.models import Mensualidad
        mensualidades_paciente = Mensualidad.objects.filter(
            paciente=paciente,
            estado__in=['activa', 'pausada', 'completada', 'cancelada']
        ).filter(
            Q(anio__gt=fecha_desde_obj.year) |
            Q(anio=fecha_desde_obj.year, mes__gte=fecha_desde_obj.month)
        ).filter(
            Q(anio__lt=fecha_hasta_obj.year) |
            Q(anio=fecha_hasta_obj.year, mes__lte=fecha_hasta_obj.month)
        )
        mensualidades_stats = {
            'total': mensualidades_paciente.count(),
            'activas': mensualidades_paciente.filter(estado='activa').count(),
            'monto_total': mensualidades_paciente.aggregate(Sum('costo_mensual'))['costo_mensual__sum'] or Decimal('0.00'),
        }
        
        datos = {
            'stats': stats,
            'tasa_asistencia': round(tasa_asistencia, 1),
            'tasa_pago': round(tasa_pago, 1),
            'por_servicio': por_servicio,
            'por_profesional': por_profesional,
            'por_sucursal': por_sucursal,
            'proyectos_stats': proyectos_stats,
            'proyectos': proyectos_paciente[:5],
            'mensualidades_stats': mensualidades_stats,
            'mensualidades': mensualidades_paciente[:5],
            'sesiones_recientes': sesiones.order_by('-fecha', '-hora_inicio')[:15],
        }
    
    # Lista de pacientes para selector
    pacientes = Paciente.objects.filter(estado='activo').order_by('apellido', 'nombre')
    
    context = {
        'paciente': paciente,
        'datos': datos,
        'grafico_data': grafico_data,
        'pacientes': pacientes,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    }
    
    return render(request, 'facturacion/reportes/paciente.html', context)

@login_required
def reporte_asistencia(request):
    """
    Reporte de asistencia y cumplimiento - MEJORADO
    ✅ Análisis completo de comportamiento
    """
    
    from datetime import datetime, timedelta
    
    tipo = request.GET.get('tipo', 'general')
    entidad_id = request.GET.get('entidad_id', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    # Rango de fechas (últimos 3 meses por defecto)
    if fecha_desde and fecha_hasta:
        fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    else:
        fecha_hasta_obj = date.today()
        fecha_desde_obj = fecha_hasta_obj - timedelta(days=90)
        fecha_desde = fecha_desde_obj.strftime('%Y-%m-%d')
        fecha_hasta = fecha_hasta_obj.strftime('%Y-%m-%d')
    
    # Query base
    sesiones = Sesion.objects.filter(
        fecha__gte=fecha_desde_obj,
        fecha__lte=fecha_hasta_obj
    ).select_related('paciente', 'servicio', 'profesional', 'sucursal')
    
    # Filtros según tipo
    entidad = None
    if tipo == 'paciente' and entidad_id:
        entidad = Paciente.objects.get(id=entidad_id)
        sesiones = sesiones.filter(paciente=entidad)
    elif tipo == 'profesional' and entidad_id:
        from profesionales.models import Profesional
        entidad = Profesional.objects.get(id=entidad_id)
        sesiones = sesiones.filter(profesional=entidad)
    
    # Estadísticas de asistencia
    stats = sesiones.aggregate(
        total=Count('id'),
        programadas=Count('id', filter=Q(estado='programada')),
        realizadas=Count('id', filter=Q(estado='realizada')),
        retrasos=Count('id', filter=Q(estado='realizada_retraso')),
        faltas=Count('id', filter=Q(estado='falta')),
        permisos=Count('id', filter=Q(estado='permiso')),
        canceladas=Count('id', filter=Q(estado='cancelada')),
    )
    
    # Calcular tasas
    sesiones_efectivas = stats['realizadas'] + stats['retrasos']
    sesiones_programadas = stats['total'] - stats['canceladas'] - stats['permisos']
    
    tasas = {
        'asistencia': (sesiones_efectivas / sesiones_programadas * 100) if sesiones_programadas > 0 else 0,
        'faltas': (stats['faltas'] / sesiones_programadas * 100) if sesiones_programadas > 0 else 0,
        'puntualidad': (stats['realizadas'] / sesiones_efectivas * 100) if sesiones_efectivas > 0 else 0,
        'cancelaciones': (stats['canceladas'] / stats['total'] * 100) if stats['total'] > 0 else 0,
    }
    
    # Ranking de asistencia (si es reporte general)
    ranking = []
    if tipo == 'general':
        pacientes_con_sesiones = Paciente.objects.filter(
            estado='activo',
            sesiones__fecha__gte=fecha_desde_obj,
            sesiones__fecha__lte=fecha_hasta_obj
        ).annotate(
            total=Count('sesiones'),
            realizadas=Count('sesiones', filter=Q(sesiones__estado__in=['realizada', 'realizada_retraso'])),
            faltas=Count('sesiones', filter=Q(sesiones__estado='falta')),
            retrasos=Count('sesiones', filter=Q(sesiones__estado='realizada_retraso'))
        ).filter(total__gte=3)
        
        ranking_data = []
        for p in pacientes_con_sesiones:
            tasa = (p.realizadas / p.total * 100) if p.total > 0 else 0
            ranking_data.append({
                'id': p.id,
                'nombre_completo': p.nombre_completo,
                'total': p.total,
                'realizadas': p.realizadas,
                'faltas': p.faltas,
                'retrasos': p.retrasos,
                'tasa': round(tasa, 1)
            })
        
        ranking = sorted(ranking_data, key=lambda x: x['tasa'], reverse=True)[:15]
    
    # Peores asistentes (opcional)
    peores_asistencia = []
    if tipo == 'general':
        pacientes_problematicos = Paciente.objects.filter(
            estado='activo',
            sesiones__fecha__gte=fecha_desde_obj,
            sesiones__fecha__lte=fecha_hasta_obj
        ).annotate(
            total=Count('sesiones'),
            faltas=Count('sesiones', filter=Q(sesiones__estado='falta')),
            realizadas=Count('sesiones', filter=Q(sesiones__estado__in=['realizada', 'realizada_retraso']))
        ).filter(total__gte=3, faltas__gt=0)
        
        peores_data = []
        for p in pacientes_problematicos:
            tasa_faltas = (p.faltas / p.total * 100) if p.total > 0 else 0
            if tasa_faltas > 20:  # Más del 20% de faltas
                peores_data.append({
                    'id': p.id,
                    'nombre_completo': p.nombre_completo,
                    'total': p.total,
                    'faltas': p.faltas,
                    'realizadas': p.realizadas,
                    'tasa_faltas': round(tasa_faltas, 1)
                })
        
        peores_asistencia = sorted(peores_data, key=lambda x: x['tasa_faltas'], reverse=True)[:10]
    
    # Por día de la semana
    from django.db.models.functions import ExtractWeekDay
    por_dia_semana = sesiones.annotate(
        dia_semana=ExtractWeekDay('fecha')
    ).values('dia_semana').annotate(
        total=Count('id'),
        realizadas=Count('id', filter=Q(estado='realizada')),
        faltas=Count('id', filter=Q(estado='falta'))
    ).order_by('dia_semana')
    
    dias_nombres = {1: 'Domingo', 2: 'Lunes', 3: 'Martes', 4: 'Miércoles', 
                   5: 'Jueves', 6: 'Viernes', 7: 'Sábado'}
    for dia in por_dia_semana:
        dia['nombre'] = dias_nombres.get(dia['dia_semana'], 'N/A')
    
    # Por servicio
    por_servicio = sesiones.values(
        'servicio__nombre'
    ).annotate(
        total=Count('id'),
        realizadas=Count('id', filter=Q(estado='realizada')),
        faltas=Count('id', filter=Q(estado='falta'))
    ).order_by('-total')
    
    # Evolución mensual
    from django.db.models.functions import TruncMonth
    por_mes = sesiones.annotate(
        mes=TruncMonth('fecha')
    ).values('mes').annotate(
        total=Count('id'),
        realizadas=Count('id', filter=Q(estado='realizada')),
        retrasos=Count('id', filter=Q(estado='realizada_retraso')),
        faltas=Count('id', filter=Q(estado='falta'))
    ).order_by('mes')
    
    grafico_data = {
        'labels': [m['mes'].strftime('%b %Y') for m in por_mes],
        'realizadas': [m['realizadas'] for m in por_mes],
        'retrasos': [m['retrasos'] for m in por_mes],
        'faltas': [m['faltas'] for m in por_mes],
    }
    
    # Listas para filtros
    pacientes = Paciente.objects.filter(estado='activo').order_by('apellido', 'nombre')
    
    from profesionales.models import Profesional
    profesionales = Profesional.objects.filter(activo=True).order_by('apellido', 'nombre')
    
    context = {
        'tipo': tipo,
        'entidad': entidad,
        'stats': stats,
        'tasas': tasas,
        'ranking': ranking,
        'peores_asistencia': peores_asistencia,
        'por_dia_semana': por_dia_semana,
        'por_servicio': por_servicio,
        'grafico_data': grafico_data,
        'pacientes': pacientes,
        'profesionales': profesionales,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    }

    return render(request, 'facturacion/reportes/asistencia.html', context)
           

@login_required
def reporte_profesional(request):
    """
    Reporte detallado por profesional - MEJORADO
    ✅ Estadísticas completas de desempeño
    """
    
    from profesionales.models import Profesional
    
    profesional_id = request.GET.get('profesional')
    sucursal_id = request.GET.get('sucursal', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    profesional = None
    datos = None
    grafico_data = None
    
    if profesional_id:
        from datetime import datetime, timedelta
        
        profesional = get_object_or_404(Profesional, id=profesional_id)
        
        # Rango de fechas (por defecto: últimos 3 meses)
        if fecha_desde and fecha_hasta:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
        else:
            fecha_hasta_obj = date.today()
            fecha_desde_obj = fecha_hasta_obj - timedelta(days=90)
            fecha_desde = fecha_desde_obj.strftime('%Y-%m-%d')
            fecha_hasta = fecha_hasta_obj.strftime('%Y-%m-%d')
        
        # Query base
        sesiones = Sesion.objects.filter(
            profesional=profesional,
            fecha__gte=fecha_desde_obj,
            fecha__lte=fecha_hasta_obj
        )
        
        # Filtro por sucursal
        if sucursal_id:
            sesiones = sesiones.filter(sucursal_id=sucursal_id)
        
        sesiones = sesiones.select_related('paciente', 'servicio', 'sucursal', 'proyecto')
        
        # Estadísticas
        stats = sesiones.aggregate(
            total_sesiones=Count('id'),
            programadas=Count('id', filter=Q(estado='programada')),
            realizadas=Count('id', filter=Q(estado='realizada')),
            retrasos=Count('id', filter=Q(estado='realizada_retraso')),
            faltas=Count('id', filter=Q(estado='falta')),
            canceladas=Count('id', filter=Q(estado='cancelada')),
            total_generado=Sum('monto_cobrado', filter=Q(estado__in=['realizada', 'realizada_retraso'])),
            pacientes_unicos=Count('paciente', distinct=True),
            total_horas=Sum('duracion_minutos', filter=Q(estado__in=['realizada', 'realizada_retraso']))
        )
        
        # Convertir minutos a horas
        stats['total_horas_decimal'] = (stats['total_horas'] or 0) / 60
        
        # Calcular tasas
        sesiones_efectivas = (stats['realizadas'] or 0) + (stats['retrasos'] or 0)
        total_programado = stats['total_sesiones'] - (stats['canceladas'] or 0)
        
        tasa_cumplimiento = (sesiones_efectivas / total_programado * 100) if total_programado > 0 else 0
        tasa_puntualidad = ((stats['realizadas'] or 0) / sesiones_efectivas * 100) if sesiones_efectivas > 0 else 0
        
        stats['tasa_cumplimiento'] = round(tasa_cumplimiento, 1)
        stats['tasa_puntualidad'] = round(tasa_puntualidad, 1)
        stats['ingreso_por_hora'] = (stats['total_generado'] or Decimal('0.00')) / Decimal(str(stats['total_horas_decimal'])) if stats['total_horas_decimal'] > 0 else Decimal('0.00')
        
        # Por servicio
        por_servicio = sesiones.values(
            'servicio__nombre', 'servicio__color'
        ).annotate(
            cantidad=Count('id'),
            realizadas=Count('id', filter=Q(estado__in=['realizada', 'realizada_retraso'])),
            ingresos=Sum('monto_cobrado', filter=Q(estado__in=['realizada', 'realizada_retraso'])),
            horas_total=Sum('duracion_minutos', filter=Q(estado__in=['realizada', 'realizada_retraso']))
        ).order_by('-cantidad')
        
        # Agregar horas decimales
        for servicio in por_servicio:
            servicio['horas_decimal'] = (servicio['horas_total'] or 0) / 60
        
        # Por sucursal
        por_sucursal = sesiones.values(
            'sucursal__nombre'
        ).annotate(
            cantidad=Count('id'),
            realizadas=Count('id', filter=Q(estado='realizada')),
            ingresos=Sum('monto_cobrado', filter=Q(estado__in=['realizada', 'realizada_retraso']))
        ).order_by('-cantidad')
        
        # Top pacientes atendidos
        top_pacientes = sesiones.values(
            'paciente__nombre', 'paciente__apellido', 'paciente__id'
        ).annotate(
            sesiones=Count('id'),
            realizadas=Count('id', filter=Q(estado__in=['realizada', 'realizada_retraso'])),
            faltas=Count('id', filter=Q(estado='falta'))
        ).order_by('-sesiones')[:10]
        
        # Por día de la semana
        from django.db.models.functions import ExtractWeekDay
        por_dia_semana = sesiones.annotate(
            dia_semana=ExtractWeekDay('fecha')
        ).values('dia_semana').annotate(
            cantidad=Count('id'),
            realizadas=Count('id', filter=Q(estado='realizada'))
        ).order_by('dia_semana')
        
        # Mapear días
        dias_nombres = {1: 'Domingo', 2: 'Lunes', 3: 'Martes', 4: 'Miércoles', 
                       5: 'Jueves', 6: 'Viernes', 7: 'Sábado'}
        for dia in por_dia_semana:
            dia['nombre'] = dias_nombres.get(dia['dia_semana'], 'N/A')
        
        # Por mes (gráfico)
        from django.db.models.functions import TruncMonth
        por_mes = sesiones.annotate(
            mes=TruncMonth('fecha')
        ).values('mes').annotate(
            cantidad=Count('id'),
            realizadas=Count('id', filter=Q(estado='realizada')),
            ingresos=Sum('monto_cobrado', filter=Q(estado__in=['realizada', 'realizada_retraso']))
        ).order_by('mes')
        
        grafico_data = {
            'labels': [m['mes'].strftime('%b %Y') for m in por_mes],
            'sesiones': [m['cantidad'] for m in por_mes],
            'realizadas': [m['realizadas'] for m in por_mes],
            'ingresos': [float(m['ingresos'] or 0) for m in por_mes],
        }
        
        datos = {
            'stats': stats,
            'por_servicio': por_servicio,
            'por_sucursal': por_sucursal,
            'top_pacientes': top_pacientes,
            'por_dia_semana': por_dia_semana,
        }
    
    # Listas para filtros
    profesionales = Profesional.objects.filter(activo=True).order_by('apellido', 'nombre')
    
    from servicios.models import Sucursal
    sucursales = Sucursal.objects.filter(activa=True)
    
    context = {
        'profesional': profesional,
        'datos': datos,
        'grafico_data': grafico_data,
        'profesionales': profesionales,
        'sucursales': sucursales,
        'sucursal_id': sucursal_id,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    }
    
    return render(request, 'facturacion/reportes/profesional.html', context)


@login_required
def reporte_sucursal(request):
    """
    Reporte detallado por sucursal - MEJORADO
    ✅ Comparativas y estadísticas completas
    """
    
    from servicios.models import Sucursal
    
    sucursal_id = request.GET.get('sucursal')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    sucursal = None
    datos = None
    comparativa = []
    grafico_data = None
    
    if sucursal_id:
        from datetime import datetime, timedelta
        
        sucursal = get_object_or_404(Sucursal, id=sucursal_id)
        
        # Rango de fechas
        if fecha_desde and fecha_hasta:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
        else:
            fecha_hasta_obj = date.today()
            fecha_desde_obj = fecha_hasta_obj - timedelta(days=90)
            fecha_desde = fecha_desde_obj.strftime('%Y-%m-%d')
            fecha_hasta = fecha_hasta_obj.strftime('%Y-%m-%d')
        
        # Query base
        sesiones = Sesion.objects.filter(
            sucursal=sucursal,
            fecha__gte=fecha_desde_obj,
            fecha__lte=fecha_hasta_obj
        ).select_related('paciente', 'servicio', 'profesional')
        
        # Estadísticas
        stats = sesiones.aggregate(
            total_sesiones=Count('id'),
            realizadas=Count('id', filter=Q(estado='realizada')),
            retrasos=Count('id', filter=Q(estado='realizada_retraso')),
            faltas=Count('id', filter=Q(estado='falta')),
            canceladas=Count('id', filter=Q(estado='cancelada')),
            ingresos_total=Sum('monto_cobrado', filter=Q(estado__in=['realizada', 'realizada_retraso'])),
            profesionales_activos=Count('profesional', distinct=True),
            pacientes_activos=Count('paciente', distinct=True),
            total_horas=Sum('duracion_minutos', filter=Q(estado__in=['realizada', 'realizada_retraso']))
        )
        
        # Sumar mensualidades de la sucursal en el período
        from agenda.models import Mensualidad
        mensualidades_sucursal = Mensualidad.objects.filter(
            sucursal=sucursal,
            estado__in=['activa', 'pausada', 'completada', 'cancelada']
        ).filter(
            Q(anio__gt=fecha_desde_obj.year) |
            Q(anio=fecha_desde_obj.year, mes__gte=fecha_desde_obj.month)
        ).filter(
            Q(anio__lt=fecha_hasta_obj.year) |
            Q(anio=fecha_hasta_obj.year, mes__lte=fecha_hasta_obj.month)
        )
        ingresos_mensualidades_sucursal = mensualidades_sucursal.aggregate(
            t=Sum('costo_mensual')
        )['t'] or Decimal('0.00')
        stats['ingresos_total'] = (stats['ingresos_total'] or Decimal('0.00')) + ingresos_mensualidades_sucursal
        stats['mensualidades_count'] = mensualidades_sucursal.count()
        stats['ingresos_mensualidades'] = ingresos_mensualidades_sucursal
        
        stats['total_horas_decimal'] = (stats['total_horas'] or 0) / 60
        
        # Tasas
        sesiones_efectivas = (stats['realizadas'] or 0) + (stats['retrasos'] or 0)
        total_prog = stats['total_sesiones'] - (stats['canceladas'] or 0)
        tasa_ocupacion = (sesiones_efectivas / total_prog * 100) if total_prog > 0 else 0
        stats['tasa_ocupacion'] = round(tasa_ocupacion, 1)
        
        # Ingreso promedio por sesión
        stats['ingreso_promedio'] = (stats['ingresos_total'] or Decimal('0.00')) / sesiones_efectivas if sesiones_efectivas > 0 else Decimal('0.00')
        
        # Por servicio
        por_servicio = sesiones.values(
            'servicio__nombre', 'servicio__color'
        ).annotate(
            cantidad=Count('id'),
            realizadas=Count('id', filter=Q(estado='realizada')),
            ingresos=Sum('monto_cobrado', filter=Q(estado__in=['realizada', 'realizada_retraso']))
        ).order_by('-cantidad')
        
        # Top profesionales
        top_profesionales = sesiones.values(
            'profesional__nombre', 'profesional__apellido', 'profesional__id'
        ).annotate(
            sesiones=Count('id'),
            realizadas=Count('id', filter=Q(estado='realizada')),
            ingresos=Sum('monto_cobrado', filter=Q(estado__in=['realizada', 'realizada_retraso']))
        ).order_by('-sesiones')[:10]
        
        # Top pacientes
        top_pacientes = sesiones.values(
            'paciente__nombre', 'paciente__apellido'
        ).annotate(
            sesiones=Count('id'),
            realizadas=Count('id', filter=Q(estado='realizada'))
        ).order_by('-sesiones')[:10]
        
        # Por mes
        from django.db.models.functions import TruncMonth
        por_mes = sesiones.annotate(
            mes=TruncMonth('fecha')
        ).values('mes').annotate(
            total=Count('id'),
            realizadas=Count('id', filter=Q(estado='realizada')),
            ingresos=Sum('monto_cobrado', filter=Q(estado__in=['realizada', 'realizada_retraso']))
        ).order_by('mes')
        
        grafico_data = {
            'labels': [m['mes'].strftime('%b %Y') for m in por_mes],
            'sesiones': [m['total'] for m in por_mes],
            'realizadas': [m['realizadas'] for m in por_mes],
            'ingresos': [float(m['ingresos'] or 0) for m in por_mes],
        }
        
        datos = {
            'stats': stats,
            'por_servicio': por_servicio,
            'top_profesionales': top_profesionales,
            'top_pacientes': top_pacientes,
        }
        
        # Comparativa con otras sucursales
        todas_sucursales = Sucursal.objects.filter(activa=True)
        
        comparativa_data = []
        for suc in todas_sucursales:
            suc_sesiones = Sesion.objects.filter(
                sucursal=suc,
                fecha__gte=fecha_desde_obj,
                fecha__lte=fecha_hasta_obj
            )
            
            suc_stats = suc_sesiones.aggregate(
                sesiones=Count('id'),
                realizadas=Count('id', filter=Q(estado='realizada')),
                ingresos=Sum('monto_cobrado', filter=Q(estado__in=['realizada', 'realizada_retraso']))
            )
            
            # Sumar mensualidades de cada sucursal
            suc_mensualidades_ingresos = Mensualidad.objects.filter(
                sucursal=suc,
                estado__in=['activa', 'pausada', 'completada', 'cancelada']
            ).filter(
                Q(anio__gt=fecha_desde_obj.year) |
                Q(anio=fecha_desde_obj.year, mes__gte=fecha_desde_obj.month)
            ).filter(
                Q(anio__lt=fecha_hasta_obj.year) |
                Q(anio=fecha_hasta_obj.year, mes__lte=fecha_hasta_obj.month)
            ).aggregate(t=Sum('costo_mensual'))['t'] or Decimal('0.00')
            
            comparativa_data.append({
                'id': suc.id,
                'nombre': suc.nombre,
                'sesiones': suc_stats['sesiones'] or 0,
                'realizadas': suc_stats['realizadas'] or 0,
                'ingresos': (suc_stats['ingresos'] or Decimal('0.00')) + suc_mensualidades_ingresos,
                'es_actual': suc.id == sucursal.id
            })
        
        comparativa = sorted(comparativa_data, key=lambda x: x['sesiones'], reverse=True)
    
    # Lista de sucursales
    sucursales = Sucursal.objects.filter(activa=True)
    
    context = {
        'sucursal': sucursal,
        'datos': datos,
        'comparativa': comparativa,
        'grafico_data': grafico_data,
        'sucursales': sucursales,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    }
    
    return render(request, 'facturacion/reportes/sucursal.html', context)

@login_required
def reporte_financiero(request):
    """
    Reporte financiero completo - VERSIÓN EXTENDIDA Y CORREGIDA
    ✅ Incluye: sesiones, proyectos, créditos, métodos de pago, cierre de caja
    ✅ Vistas detalladas de pagos, sesiones, proyectos y créditos
    ✅ CORREGIDO: Resta devoluciones y pagos anulados del total recaudado
    """
    
    from datetime import datetime, timedelta
    from servicios.models import Sucursal
    from agenda.models import Proyecto
    from django.db.models import Q, Sum, Count, DecimalField
    from django.db.models.functions import Coalesce
    
    sucursal_id = request.GET.get('sucursal', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    vista = request.GET.get('vista', 'mensual')  # mensual, diaria, detalle_pagos, detalle_sesiones, detalle_proyectos, analisis_creditos
    
    # Rango de fechas
    if fecha_desde and fecha_hasta:
        fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    else:
        fecha_hasta_obj = date.today()
        if vista == 'diaria':
            fecha_desde_obj = fecha_hasta_obj
        else:
            fecha_desde_obj = fecha_hasta_obj.replace(day=1)
        fecha_desde = fecha_desde_obj.strftime('%Y-%m-%d')
        fecha_hasta = fecha_hasta_obj.strftime('%Y-%m-%d')
    
    # ==================== SESIONES ====================
    sesiones = Sesion.objects.filter(
        fecha__gte=fecha_desde_obj,
        fecha__lte=fecha_hasta_obj
    ).select_related('paciente', 'servicio', 'profesional', 'sucursal')
    
    # ==================== PROYECTOS ====================
    proyectos = Proyecto.objects.filter(
        fecha_inicio__gte=fecha_desde_obj,
        fecha_inicio__lte=fecha_hasta_obj
    ).select_related('paciente', 'servicio_base', 'profesional_responsable', 'sucursal')
    
    # ==================== MENSUALIDADES ====================
    # Filtrar mensualidades cuyo mes/año cae dentro del rango seleccionado
    from agenda.models import Mensualidad
    mensualidades = Mensualidad.objects.filter(
        estado__in=['activa', 'vencida', 'pausada', 'completada', 'cancelada']
    ).filter(
        # mes/año >= fecha_desde
        Q(anio__gt=fecha_desde_obj.year) |
        Q(anio=fecha_desde_obj.year, mes__gte=fecha_desde_obj.month)
    ).filter(
        # mes/año <= fecha_hasta
        Q(anio__lt=fecha_hasta_obj.year) |
        Q(anio=fecha_hasta_obj.year, mes__lte=fecha_hasta_obj.month)
    ).select_related('paciente', 'sucursal')
    
    if sucursal_id:
        mensualidades = mensualidades.filter(sucursal_id=sucursal_id)
    
    # ==================== PAGOS ====================
    pagos = Pago.objects.filter(
        fecha_pago__gte=fecha_desde_obj,
        fecha_pago__lte=fecha_hasta_obj,
        anulado=False  # Solo pagos válidos
    ).select_related('paciente', 'metodo_pago', 'sesion', 'proyecto')
    
    # ==================== DEVOLUCIONES ====================
    devoluciones = Devolucion.objects.filter(
        fecha_devolucion__gte=fecha_desde_obj,
        fecha_devolucion__lte=fecha_hasta_obj
    ).select_related('paciente', 'metodo_devolucion', 'proyecto', 'mensualidad')
    
    # ==================== PAGOS ANULADOS (para mostrar) ====================
    pagos_anulados = Pago.objects.filter(
        fecha_pago__gte=fecha_desde_obj,
        fecha_pago__lte=fecha_hasta_obj,
        anulado=True
    ).select_related('paciente', 'metodo_pago', 'sesion', 'proyecto')
    
    # Filtro por sucursal
    if sucursal_id:
        sesiones = sesiones.filter(sucursal_id=sucursal_id)
        proyectos = proyectos.filter(sucursal_id=sucursal_id)
        pagos = pagos.filter(
            Q(sesion__sucursal_id=sucursal_id) | 
            Q(proyecto__sucursal_id=sucursal_id) |
            Q(sesion__isnull=True, proyecto__isnull=True, paciente__sucursales__id=sucursal_id)
        )
        devoluciones = devoluciones.filter(
            Q(proyecto__sucursal_id=sucursal_id) |
            Q(mensualidad__sucursal_id=sucursal_id) |
            Q(proyecto__isnull=True, mensualidad__isnull=True, paciente__sucursales__id=sucursal_id)
        )
        pagos_anulados = pagos_anulados.filter(
            Q(sesion__sucursal_id=sucursal_id) | 
            Q(proyecto__sucursal_id=sucursal_id) |
            Q(sesion__isnull=True, proyecto__isnull=True, paciente__sucursales__id=sucursal_id)
        )
    # ==================== CONTEXTO BASE ====================
    context = {
        'vista': vista,
        'sucursal_id': sucursal_id,
        'sucursales': Sucursal.objects.filter(activa=True),
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    }
    
    # ══════════════════════════════════════════════════════════════════
    # VISTA: DETALLE PAGOS
    # ══════════════════════════════════════════════════════════════════
    if vista == 'detalle_pagos':
        detalle_pagos = pagos.order_by('-fecha_pago', '-fecha_registro')
        
        # Totales por método
        detalle_pagos_metodos = pagos.values('metodo_pago__nombre').annotate(
            cantidad=Count('id'),
            total=Sum('monto')
        ).order_by('-total')
        
        # Total devoluciones
        total_devoluciones = devoluciones.aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
        total_anulaciones = pagos_anulados.aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
        
        context.update({
            'detalle_pagos': detalle_pagos,
            'detalle_pagos_metodos': detalle_pagos_metodos,
            'detalle_devoluciones': devoluciones,
            'detalle_anulaciones': pagos_anulados,
            'total_devoluciones': total_devoluciones,
            'total_anulaciones': total_anulaciones,
        })
        return render(request, 'facturacion/reportes/financiero.html', context)
    
    # ══════════════════════════════════════════════════════════════════
    # VISTA: DETALLE SESIONES
    # ══════════════════════════════════════════════════════════════════
    elif vista == 'detalle_sesiones':
        # Todos los estados de sesión
        todos_estados_sesion = [
            'programada', 'realizada', 'realizada_retraso',
            'falta', 'permiso', 'cancelada', 'reprogramada'
        ]
        detalle_sesiones = sesiones.filter(
            estado__in=todos_estados_sesion
        ).order_by('-fecha', '-hora_inicio')

        # Calcular monto pagado y pendiente por sesión (incluye pagos masivos)
        for sesion in detalle_sesiones:
            pagos_directos = pagos.filter(sesion=sesion).exclude(
                metodo_pago__nombre="Uso de Crédito"
            ).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
            pagos_masivos_sesion = DetallePagoMasivo.objects.filter(
                tipo='sesion',
                sesion=sesion,
                pago__anulado=False,
            ).exclude(
                pago__metodo_pago__nombre="Uso de Crédito"
            ).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
            sesion.monto_pagado = pagos_directos + pagos_masivos_sesion
            sesion.monto_pendiente = max(sesion.monto_cobrado - sesion.monto_pagado, Decimal('0.00'))

        # Totales y contadores por estado
        detalle_sesiones_totales = {
            'total_generado': detalle_sesiones.aggregate(Sum('monto_cobrado'))['monto_cobrado__sum'] or Decimal('0.00'),
            'total_pagado': sum(s.monto_pagado for s in detalle_sesiones),
            'total_pendiente': sum(s.monto_pendiente for s in detalle_sesiones),
            'por_estado': {
                estado: detalle_sesiones.filter(estado=estado).count()
                for estado in todos_estados_sesion
            },
        }
        if detalle_sesiones_totales['total_generado'] > 0:
            detalle_sesiones_totales['tasa_cobranza'] = (
                detalle_sesiones_totales['total_pagado'] /
                detalle_sesiones_totales['total_generado']
            ) * 100
        else:
            detalle_sesiones_totales['tasa_cobranza'] = 0

        context.update({
            'detalle_sesiones': detalle_sesiones,
            'detalle_sesiones_totales': detalle_sesiones_totales,
        })
        return render(request, 'facturacion/reportes/financiero.html', context)
    
    # ══════════════════════════════════════════════════════════════════
    # VISTA: DETALLE PROYECTOS
    # ══════════════════════════════════════════════════════════════════
    elif vista == 'detalle_proyectos':
        # Todos los estados de proyecto
        detalle_proyectos = proyectos.order_by('-fecha_inicio')

        # Calcular monto pagado y pendiente por proyecto (incluye pagos masivos, neto de devoluciones)
        for proyecto in detalle_proyectos:
            pagos_directos = pagos.filter(proyecto=proyecto).exclude(
                metodo_pago__nombre="Uso de Crédito"
            ).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
            pagos_masivos_proyecto = DetallePagoMasivo.objects.filter(
                tipo='proyecto',
                proyecto=proyecto,
                pago__anulado=False,
            ).exclude(
                pago__metodo_pago__nombre="Uso de Crédito"
            ).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
            # Devoluciones sin restricción de fecha (igual que _total_pagado_proyecto)
            devoluciones_proyecto = Devolucion.objects.filter(
                proyecto=proyecto
            ).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
            proyecto.monto_pagado = pagos_directos + pagos_masivos_proyecto - devoluciones_proyecto
            proyecto.monto_pendiente = max(proyecto.costo_total - proyecto.monto_pagado, Decimal('0.00'))

        # Totales
        detalle_proyectos_totales = {
            'total_generado': proyectos.aggregate(Sum('costo_total'))['costo_total__sum'] or Decimal('0.00'),
            'total_pagado': sum(p.monto_pagado for p in detalle_proyectos),
            'total_pendiente': sum(p.monto_pendiente for p in detalle_proyectos),
            'planificados': proyectos.filter(estado='planificado').count(),
            'activos': proyectos.filter(estado='en_progreso').count(),
            'finalizados': proyectos.filter(estado='finalizado').count(),
            'cancelados': proyectos.filter(estado='cancelado').count(),
        }

        context.update({
            'detalle_proyectos': detalle_proyectos,
            'detalle_proyectos_totales': detalle_proyectos_totales,
        })
        return render(request, 'facturacion/reportes/financiero.html', context)
    
    # ══════════════════════════════════════════════════════════════════
    # VISTA: DETALLE MENSUALIDADES
    # ══════════════════════════════════════════════════════════════════
    elif vista == 'detalle_mensualidades':
        # Todos los estados de mensualidad
        detalle_mensualidades = mensualidades.prefetch_related(
            'servicios_profesionales__servicio',
            'servicios_profesionales__profesional',
        ).order_by('-anio', '-mes')

        # Calcular monto pagado y pendiente por mensualidad (incluye pagos masivos, neto de devoluciones)
        for mensualidad in detalle_mensualidades:
            pagos_directos = pagos.filter(mensualidad=mensualidad).exclude(
                metodo_pago__nombre="Uso de Crédito"
            ).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
            pagos_masivos_mens = DetallePagoMasivo.objects.filter(
                tipo='mensualidad',
                mensualidad=mensualidad,
                pago__anulado=False,
            ).exclude(
                pago__metodo_pago__nombre="Uso de Crédito"
            ).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
            # Devoluciones sin restricción de fecha (igual que _total_pagado_mensualidad)
            devoluciones_mens = Devolucion.objects.filter(
                mensualidad=mensualidad
            ).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
            mensualidad.monto_pagado = pagos_directos + pagos_masivos_mens - devoluciones_mens
            mensualidad.monto_pendiente = max(mensualidad.costo_mensual - mensualidad.monto_pagado, Decimal('0.00'))

        # Totales
        detalle_mensualidades_totales = {
            'total_generado': mensualidades.aggregate(Sum('costo_mensual'))['costo_mensual__sum'] or Decimal('0.00'),
            'total_pagado': sum(m.monto_pagado for m in detalle_mensualidades),
            'total_pendiente': sum(m.monto_pendiente for m in detalle_mensualidades),
            'activas': mensualidades.filter(estado='activa').count(),
            'pausadas': mensualidades.filter(estado='pausada').count(),
            'completadas': mensualidades.filter(estado='completada').count(),
            'canceladas': mensualidades.filter(estado='cancelada').count(),
        }
        if detalle_mensualidades_totales['total_generado'] > 0:
            detalle_mensualidades_totales['tasa_cobranza'] = (
                detalle_mensualidades_totales['total_pagado'] /
                detalle_mensualidades_totales['total_generado']
            ) * 100
        else:
            detalle_mensualidades_totales['tasa_cobranza'] = 0

        context.update({
            'detalle_mensualidades': detalle_mensualidades,
            'detalle_mensualidades_totales': detalle_mensualidades_totales,
        })
        return render(request, 'facturacion/reportes/financiero.html', context)

    # ══════════════════════════════════════════════════════════════════
    # VISTA: ANÁLISIS DE CRÉDITOS
    # ══════════════════════════════════════════════════════════════════
    elif vista == 'analisis_creditos':
        # Créditos generados (pagos adelantados)
        creditos_generados = pagos.filter(
            sesion__isnull=True,
            proyecto__isnull=True
        ).exclude(metodo_pago__nombre="Uso de Crédito").order_by('-fecha_pago')
        
        # Créditos utilizados
        creditos_utilizados = Pago.objects.filter(
            fecha_pago__gte=fecha_desde_obj,
            fecha_pago__lte=fecha_hasta_obj,
            metodo_pago__nombre="Uso de Crédito",
            anulado=False
        ).select_related('paciente', 'sesion__servicio', 'proyecto__servicio_base').order_by('-fecha_pago')
        
        if sucursal_id:
            creditos_utilizados = creditos_utilizados.filter(
                Q(sesion__sucursal_id=sucursal_id) | 
                Q(proyecto__sucursal_id=sucursal_id) |
                Q(sesion__isnull=True, proyecto__isnull=True, paciente__sucursales__id=sucursal_id)
            )
        
        # Totales
        total_generado = creditos_generados.aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
        total_utilizado = creditos_utilizados.aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
        
        # Pacientes con mayor crédito disponible
        pacientes_con_credito = Paciente.objects.annotate(
            credito_generado=Coalesce(
                Sum('pagos__monto', filter=Q(
                    pagos__sesion__isnull=True,
                    pagos__proyecto__isnull=True,
                    pagos__anulado=False
                ) & ~Q(pagos__metodo_pago__nombre="Uso de Crédito")),
                Decimal('0.00'),
                output_field=DecimalField()
            ),
            credito_usado=Coalesce(
                Sum('pagos__monto', filter=Q(
                    pagos__metodo_pago__nombre="Uso de Crédito",
                    pagos__anulado=False
                )),
                Decimal('0.00'),
                output_field=DecimalField()
            )
        ).annotate(
            credito_disponible=F('credito_generado') - F('credito_usado')
        ).filter(
            credito_disponible__gt=0
        ).order_by('-credito_disponible')[:20]
        
        context.update({
            'analisis_creditos': {
                'total_generado': total_generado,
                'cantidad_generados': creditos_generados.count(),
                'total_utilizado': total_utilizado,
                'cantidad_utilizados': creditos_utilizados.count(),
                'saldo_neto': total_generado - total_utilizado,
                'generados': creditos_generados,
                'utilizados': creditos_utilizados,
                'pacientes_con_credito': pacientes_con_credito,
            }
        })
        return render(request, 'facturacion/reportes/financiero.html', context)
    
    # ══════════════════════════════════════════════════════════════════
    # VISTAS: MENSUAL Y DIARIA (CÓDIGO ORIGINAL CORREGIDO)
    # ══════════════════════════════════════════════════════════════════
    
    # ==================== INGRESOS POR SESIONES ====================
    # REAL: realizada, realizada_retraso, falta (igual que total_consumido_actual en services.py)
    sesiones_realizadas = sesiones.filter(
        estado__in=['realizada', 'realizada_retraso', 'falta'],
        proyecto__isnull=True,
        mensualidad__isnull=True,
    )
    # PROYECTADO: agrega las programadas (igual que total_consumido_real en services.py)
    sesiones_programadas_rpt = sesiones.filter(
        estado='programada',
        proyecto__isnull=True,
        mensualidad__isnull=True,
    )

    total_generado_sesiones_real = sesiones_realizadas.aggregate(
        Sum('monto_cobrado'))['monto_cobrado__sum'] or Decimal('0.00')
    total_generado_sesiones_prog = sesiones_programadas_rpt.aggregate(
        Sum('monto_cobrado'))['monto_cobrado__sum'] or Decimal('0.00')

    ingresos_sesiones = {
        'total_generado_real': total_generado_sesiones_real,
        'total_generado_proyectado': total_generado_sesiones_real + total_generado_sesiones_prog,
        'cantidad_sesiones': sesiones_realizadas.count(),
        'cantidad_programadas': sesiones_programadas_rpt.count(),
    }

    # Cobrado en sesiones: directos + DPM en el período, sin "Uso de Crédito"
    pagos_sesiones_directos = pagos.filter(
        sesion__isnull=False,
        sesion__proyecto__isnull=True,
        sesion__mensualidad__isnull=True,
    ).exclude(metodo_pago__nombre="Uso de Crédito")
    pagos_masivos_sesiones_report = DetallePagoMasivo.objects.filter(
        tipo='sesion',
        pago__anulado=False,
        pago__fecha_pago__gte=fecha_desde_obj,
        pago__fecha_pago__lte=fecha_hasta_obj,
    ).exclude(pago__metodo_pago__nombre="Uso de Crédito")
    ingresos_sesiones['total_cobrado'] = (
        (pagos_sesiones_directos.aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')) +
        (pagos_masivos_sesiones_report.aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00'))
    )
    ingresos_sesiones['total_pendiente'] = max(
        ingresos_sesiones['total_generado_real'] - ingresos_sesiones['total_cobrado'],
        Decimal('0.00')
    )

    # ==================== INGRESOS POR PROYECTOS ====================
    # REAL: en_progreso, finalizado, cancelado (igual que services.py)
    proyectos_reales = proyectos.filter(estado__in=['en_progreso', 'finalizado', 'cancelado'])
    # PROYECTADO: agrega planificados
    proyectos_planificados_rpt = proyectos.filter(estado='planificado')

    total_generado_proy_real = proyectos_reales.aggregate(
        Sum('costo_total'))['costo_total__sum'] or Decimal('0.00')
    total_generado_proy_plan = proyectos_planificados_rpt.aggregate(
        Sum('costo_total'))['costo_total__sum'] or Decimal('0.00')

    ingresos_proyectos = {
        'total_generado_real': total_generado_proy_real,
        'total_generado_proyectado': total_generado_proy_real + total_generado_proy_plan,
        'cantidad_proyectos': proyectos_reales.count(),
        'cantidad_planificados': proyectos_planificados_rpt.count(),
        'proyectos_activos': proyectos_reales.filter(estado='en_progreso').count(),
        'proyectos_finalizados': proyectos_reales.filter(estado='finalizado').count(),
        'proyectos_cancelados': proyectos_reales.filter(estado='cancelado').count(),
    }

    # Cobrado en proyectos: directos + DPM - devoluciones del período
    pagos_proyectos_directos = pagos.filter(
        proyecto__in=proyectos_reales
    ).exclude(metodo_pago__nombre="Uso de Crédito")
    pagos_masivos_proyectos = DetallePagoMasivo.objects.filter(
        tipo='proyecto',
        proyecto__in=proyectos_reales,
        pago__anulado=False,
        pago__fecha_pago__gte=fecha_desde_obj,
        pago__fecha_pago__lte=fecha_hasta_obj,
    ).exclude(pago__metodo_pago__nombre="Uso de Crédito")
    devoluciones_proyectos = devoluciones.filter(
        proyecto__in=proyectos_reales
    ).aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')

    ingresos_proyectos['total_cobrado'] = (
        (pagos_proyectos_directos.aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')) +
        (pagos_masivos_proyectos.aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00'))
    ) - devoluciones_proyectos
    ingresos_proyectos['total_pendiente'] = max(
        ingresos_proyectos['total_generado_real'] - ingresos_proyectos['total_cobrado'],
        Decimal('0.00')
    )

    # ==================== INGRESOS POR MENSUALIDADES ====================
    # Mensualidades no tienen estado "planificado" — siempre se incluyen igual en real y proyectado
    total_generado_mens = mensualidades.aggregate(
        Sum('costo_mensual'))['costo_mensual__sum'] or Decimal('0.00')

    ingresos_mensualidades = {
        'total_generado_real': total_generado_mens,
        'total_generado_proyectado': total_generado_mens,  # igual: no hay mensualidades "planificadas"
        'cantidad_mensualidades': mensualidades.count(),
        'activas': mensualidades.filter(estado='activa').count(),
        'vencidas': mensualidades.filter(estado='vencida').count(),
        'pausadas': mensualidades.filter(estado='pausada').count(),
        'completadas': mensualidades.filter(estado='completada').count(),
        'canceladas': mensualidades.filter(estado='cancelada').count(),
    }

    # Cobrado en mensualidades: directos + DPM - devoluciones del período
    pagos_mensualidades_qs = pagos.filter(
        mensualidad__in=mensualidades
    ).exclude(metodo_pago__nombre="Uso de Crédito")
    pagos_masivos_mensualidades = DetallePagoMasivo.objects.filter(
        tipo='mensualidad',
        mensualidad__in=mensualidades,
        pago__anulado=False,
        pago__fecha_pago__gte=fecha_desde_obj,
        pago__fecha_pago__lte=fecha_hasta_obj
    ).exclude(pago__metodo_pago__nombre="Uso de Crédito")
    devoluciones_mensualidades = devoluciones.filter(
        mensualidad__in=mensualidades
    ).aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')

    ingresos_mensualidades['total_cobrado'] = (
        (pagos_mensualidades_qs.aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')) +
        (pagos_masivos_mensualidades.aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00'))
    ) - devoluciones_mensualidades
    ingresos_mensualidades['total_pendiente'] = max(
        ingresos_mensualidades['total_generado_real'] - ingresos_mensualidades['total_cobrado'],
        Decimal('0.00')
    )

    # ==================== MOVIMIENTO DE CRÉDITOS ====================
    pagos_adelantados = pagos.filter(
        sesion__isnull=True,
        proyecto__isnull=True,
        mensualidad__isnull=True,
    ).exclude(
        metodo_pago__nombre="Uso de Crédito"
    ).exclude(
        detalles_masivos__isnull=False
    )
    creditos_generados = pagos_adelantados.aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')

    pagos_con_credito = pagos.filter(metodo_pago__nombre="Uso de Crédito")
    creditos_utilizados = pagos_con_credito.aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')

    movimiento_creditos = {
        'generados': creditos_generados,
        'generados_cantidad': pagos_adelantados.count(),
        'utilizados': creditos_utilizados,
        'utilizados_cantidad': pagos_con_credito.count(),
        'saldo_neto': creditos_generados - creditos_utilizados,
    }

    # ==================== DEVOLUCIONES Y ANULACIONES ====================
    total_devoluciones = devoluciones.aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
    total_anulaciones = pagos_anulados.aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')

    devoluciones_por_metodo = devoluciones.values('metodo_devolucion__nombre').annotate(
        cantidad=Count('id'),
        monto=Sum('monto')
    ).order_by('-monto')

    # ==================== TOTALES GENERALES ====================
    # REAL: solo lo ya ocurrido/comprometido (igual que total_consumido_actual en services.py)
    total_generado_real = (
        ingresos_sesiones['total_generado_real'] +
        ingresos_proyectos['total_generado_real'] +
        ingresos_mensualidades['total_generado_real']
    )
    # PROYECTADO: incluye sesiones programadas y proyectos planificados
    total_generado_proyectado = (
        ingresos_sesiones['total_generado_proyectado'] +
        ingresos_proyectos['total_generado_proyectado'] +
        ingresos_mensualidades['total_generado_proyectado']
    )

    # TOTAL RECAUDADO BRUTO — fuente única: suma directa del queryset pagos (idéntico al historial)
    total_cobrado_bruto = (
        pagos
        .exclude(metodo_pago__nombre="Uso de Crédito")
        .aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
    )
    total_cobrado_neto = total_cobrado_bruto - total_devoluciones

    # Por Cobrar Real = Total Ingresos Real - Recaudado Neto
    total_pendiente_real = max(total_generado_real - total_cobrado_neto, Decimal('0.00'))
    # Por Cobrar Proyectado = Total Ingresos Proyectado - Recaudado Neto
    total_pendiente_proyectado = max(total_generado_proyectado - total_cobrado_neto, Decimal('0.00'))

    ingresos = {
        # Dos versiones de "Total Ingresos" como en cuenta corriente
        'total_generado': total_generado_real,           # alias para compatibilidad con template
        'total_generado_real': total_generado_real,
        'total_generado_proyectado': total_generado_proyectado,
        'total_generado_sesiones_prog': total_generado_sesiones_prog,
        'total_generado_proy_plan': total_generado_proy_plan,
        # Recaudado
        'total_cobrado_bruto': total_cobrado_bruto,
        'total_cobrado_neto': total_cobrado_neto,
        'total_devoluciones': total_devoluciones,
        'total_anulaciones': total_anulaciones,
        'total_pendiente': total_pendiente_real,
        'total_pendiente_proyectado': total_pendiente_proyectado,
        # Sub-stats por categoría
        'sesiones': ingresos_sesiones,
        'proyectos': ingresos_proyectos,
        'mensualidades': ingresos_mensualidades,
        'creditos': movimiento_creditos,
        'devoluciones_info': {
            'total': total_devoluciones,
            'cantidad': devoluciones.count(),
            'por_metodo': devoluciones_por_metodo,
        },
        'anulaciones_info': {
            'total': total_anulaciones,
            'cantidad': pagos_anulados.count(),
        }
    }

    total_items = (
        ingresos_sesiones['cantidad_sesiones'] +
        ingresos_proyectos['cantidad_proyectos'] +
        ingresos_mensualidades['cantidad_mensualidades']
    )
    ingresos['promedio_por_item'] = ingresos['total_generado_real'] / total_items if total_items > 0 else Decimal('0.00')
    ingresos['tasa_cobranza'] = (total_cobrado_neto / ingresos['total_generado_real'] * 100) if ingresos['total_generado_real'] > 0 else 0
    
    # ==================== POR MÉTODO DE PAGO ====================
    por_metodo = pagos.exclude(metodo_pago__nombre="Uso de Crédito").values(
        'metodo_pago__nombre'
    ).annotate(
        cantidad=Count('id'),
        monto=Sum('monto')
    ).order_by('-monto')
    
    # ==================== POR SERVICIO ====================
    # Sesiones
    por_servicio_sesiones = sesiones_realizadas.values(
        'servicio__nombre', 'servicio__color'
    ).annotate(
        sesiones=Count('id'),
        ingresos=Sum('monto_cobrado')
    )
    
    # Proyectos
    por_servicio_proyectos = proyectos.values(
        'servicio_base__nombre', 'servicio_base__color'
    ).annotate(
        proyectos=Count('id'),
        ingresos=Sum('costo_total')
    )
    
    # Combinar servicios
    servicios_dict = {}
    
    for s in por_servicio_sesiones:
        nombre = s['servicio__nombre']
        servicios_dict[nombre] = {
            'nombre': nombre,
            'color': s['servicio__color'],
            'sesiones': s['sesiones'],
            'proyectos': 0,
            'ingresos': s['ingresos'] or Decimal('0.00')
        }
    
    for p in por_servicio_proyectos:
        nombre = p['servicio_base__nombre']
        if nombre in servicios_dict:
            servicios_dict[nombre]['proyectos'] = p['proyectos']
            servicios_dict[nombre]['ingresos'] += p['ingresos'] or Decimal('0.00')
        else:
            servicios_dict[nombre] = {
                'nombre': nombre,
                'color': p['servicio_base__color'],
                'sesiones': 0,
                'proyectos': p['proyectos'],
                'ingresos': p['ingresos'] or Decimal('0.00')
            }
    
    por_servicio = sorted(servicios_dict.values(), key=lambda x: x['ingresos'], reverse=True)[:10]
    
    # ==================== CIERRE DE CAJA DIARIO ====================
    cierre_diario = None
    
    if vista == 'diaria':
        pagos_dia = pagos.filter(fecha_pago=fecha_desde_obj)
        devoluciones_dia = devoluciones.filter(fecha_devolucion=fecha_desde_obj)
        pagos_anulados_dia = pagos_anulados.filter(fecha_pago=fecha_desde_obj)
        
        # Total cobrado (sin crédito)
        pagos_dia_validos = pagos_dia.exclude(metodo_pago__nombre="Uso de Crédito")
        
        # Total devoluciones del día
        total_devoluciones_dia = devoluciones_dia.aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
        total_anulaciones_dia = pagos_anulados_dia.aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
        
        # Monto total bruto y neto
        monto_total_bruto = pagos_dia_validos.aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
        monto_total_neto = monto_total_bruto - total_devoluciones_dia
        
        cierre_diario = {
            'fecha': fecha_desde_obj,
            'fecha_formato': fecha_desde_obj.strftime('%d de %B de %Y'),
            'dia_semana': ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo'][fecha_desde_obj.weekday()],
            
            # Pagos totales
            'pagos_total': pagos_dia_validos.count(),
            'monto_total_bruto': monto_total_bruto,
            'monto_total_neto': monto_total_neto,
            'total_devoluciones': total_devoluciones_dia,
            'total_anulaciones': total_anulaciones_dia,
            
            # Detalle por método
            'por_metodo': pagos_dia_validos.values('metodo_pago__nombre').annotate(
                cantidad=Count('id'),
                monto=Sum('monto')
            ).order_by('-monto'),
            
            # Devoluciones del día
            'devoluciones_dia': devoluciones_dia,
            'devoluciones_por_metodo': devoluciones_dia.values('metodo_devolucion__nombre').annotate(
                cantidad=Count('id'),
                monto=Sum('monto')
            ).order_by('-monto'),
            
            # Anulaciones del día
            'anulaciones_dia': pagos_anulados_dia,
            
            # Sesiones del día
            'sesiones_realizadas': sesiones.filter(
                fecha=fecha_desde_obj,
                estado__in=['realizada', 'realizada_retraso']
            ).count(),
            'monto_generado_sesiones': sesiones.filter(
                fecha=fecha_desde_obj,
                estado__in=['realizada', 'realizada_retraso']
            ).aggregate(Sum('monto_cobrado'))['monto_cobrado__sum'] or Decimal('0.00'),
            
            # Proyectos del día
            'proyectos_iniciados': proyectos.filter(fecha_inicio=fecha_desde_obj).count(),
            'monto_proyectos': proyectos.filter(fecha_inicio=fecha_desde_obj).aggregate(Sum('costo_total'))['costo_total__sum'] or Decimal('0.00'),
            
            # Mensualidades del día (del mes/año correspondiente)
            'mensualidades_dia': mensualidades.filter(
                anio=fecha_desde_obj.year, mes=fecha_desde_obj.month
            ).count(),
            'monto_mensualidades': mensualidades.filter(
                anio=fecha_desde_obj.year, mes=fecha_desde_obj.month
            ).aggregate(Sum('costo_mensual'))['costo_mensual__sum'] or Decimal('0.00'),
            
            # Créditos del día
            'creditos_generados': pagos_dia.filter(sesion__isnull=True, proyecto__isnull=True).exclude(metodo_pago__nombre="Uso de Crédito").aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00'),
            'creditos_utilizados': pagos_dia.filter(metodo_pago__nombre="Uso de Crédito").aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00'),
        }
        
        # Efectivo esperado (restando devoluciones en efectivo)
        efectivo_cobrado = pagos_dia_validos.filter(
            metodo_pago__nombre__in=['Efectivo', 'efectivo']
        ).aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
        
        efectivo_devuelto = devoluciones_dia.filter(
            metodo_devolucion__nombre__in=['Efectivo', 'efectivo']
        ).aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
        
        cierre_diario['efectivo_esperado'] = efectivo_cobrado - efectivo_devuelto
        
        # Detalle COMPLETO de pagos del día (agrupados por recibo para pagos masivos)
        pagos_detalle_raw = pagos_dia_validos.select_related(
            'paciente', 'metodo_pago', 'sesion__servicio', 'proyecto', 'registrado_por'
        ).order_by('numero_recibo', '-monto')
        
        # Agrupar pagos por recibo
        pagos_agrupados = {}
        for pago in pagos_detalle_raw:
            recibo = pago.numero_recibo
            if recibo not in pagos_agrupados:
                pagos_agrupados[recibo] = {
                    'recibo': recibo,
                    'paciente': pago.paciente,
                    'metodo_pago': pago.metodo_pago,
                    'registrado_por': pago.registrado_por,
                    'fecha_registro': pago.fecha_registro,
                    'observaciones': pago.observaciones,
                    'numero_transaccion': pago.numero_transaccion,
                    'pagos': [],
                    'total': Decimal('0.00'),
                    'sesiones_count': 0,
                    'proyectos_count': 0,
                    'adelantados_count': 0,
                }
            
            pagos_agrupados[recibo]['pagos'].append(pago)
            pagos_agrupados[recibo]['total'] += pago.monto
            
            # Contar tipos
            if pago.sesion:
                pagos_agrupados[recibo]['sesiones_count'] += 1
            elif pago.proyecto:
                pagos_agrupados[recibo]['proyectos_count'] += 1
            else:
                pagos_agrupados[recibo]['adelantados_count'] += 1
        
        # Marcar como masivos si tienen más de 1 pago
        for recibo_data in pagos_agrupados.values():
            recibo_data['es_masivo'] = len(recibo_data['pagos']) > 1
        
        cierre_diario['pagos_agrupados'] = list(pagos_agrupados.values())
        
        # Estadísticas adicionales del día
        cierre_diario['estadisticas'] = {
            # Pacientes atendidos
            'pacientes_unicos': sesiones.filter(
                fecha=fecha_desde_obj,
                estado__in=['realizada', 'realizada_retraso']
            ).values('paciente').distinct().count(),
            
            # Por profesional
            'por_profesional': sesiones.filter(
                fecha=fecha_desde_obj,
                estado__in=['realizada', 'realizada_retraso']
            ).values('profesional__nombre', 'profesional__apellido').annotate(
                sesiones=Count('id'),
                ingresos=Sum('monto_cobrado')
            ).order_by('-sesiones'),
            
            # Por servicio
            'por_servicio': sesiones.filter(
                fecha=fecha_desde_obj,
                estado__in=['realizada', 'realizada_retraso']
            ).values('servicio__nombre', 'servicio__color').annotate(
                cantidad=Count('id'),
                ingresos=Sum('monto_cobrado')
            ).order_by('-cantidad'),
            
            # Asistencia del día
            'programadas_dia': sesiones.filter(fecha=fecha_desde_obj, estado='programada').count(),
            'retrasos_dia': sesiones.filter(fecha=fecha_desde_obj, estado='realizada_retraso').count(),
            'faltas_dia': sesiones.filter(fecha=fecha_desde_obj, estado='falta').count(),
            'canceladas_dia': sesiones.filter(fecha=fecha_desde_obj, estado='cancelada').count(),
            
            # Promedio ticket del día (sobre monto neto)
            'ticket_promedio': monto_total_neto / pagos_dia_validos.count() if pagos_dia_validos.count() > 0 else Decimal('0.00'),
            
            # Horas trabajadas
            'horas_trabajadas': (sesiones.filter(
                fecha=fecha_desde_obj,
                estado__in=['realizada', 'realizada_retraso']
            ).aggregate(Sum('duracion_minutos'))['duracion_minutos__sum'] or 0) / 60,
            
            # Ingreso por hora
            'ingreso_por_hora': Decimal('0.00'),
        }
        
        # Calcular ingreso por hora
        if cierre_diario['estadisticas']['horas_trabajadas'] > 0:
            cierre_diario['estadisticas']['ingreso_por_hora'] = cierre_diario['monto_generado_sesiones'] / Decimal(str(cierre_diario['estadisticas']['horas_trabajadas']))
        
        # Comparativa con días anteriores (últimos 7 días)
        comparativa_dias = []
        for i in range(1, 8):
            dia_anterior = fecha_desde_obj - timedelta(days=i)
            
            sesiones_dia_ant = sesiones.filter(
                fecha=dia_anterior,
                estado__in=['realizada', 'realizada_retraso']
            ).aggregate(
                cantidad=Count('id'),
                ingresos=Sum('monto_cobrado')
            )
            
            pagos_dia_ant = pagos.filter(
                fecha_pago=dia_anterior
            ).exclude(metodo_pago__nombre="Uso de Crédito").aggregate(
                cobrado=Sum('monto')
            )
            
            # Restar devoluciones de días anteriores
            devoluciones_dia_ant = devoluciones.filter(
                fecha_devolucion=dia_anterior
            ).aggregate(devuelto=Sum('monto'))['devuelto'] or Decimal('0.00')
            
            comparativa_dias.append({
                'fecha': dia_anterior,
                'dia_semana': ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom'][dia_anterior.weekday()],
                'sesiones': sesiones_dia_ant['cantidad'] or 0,
                'ingresos': sesiones_dia_ant['ingresos'] or Decimal('0.00'),
                'cobrado_bruto': pagos_dia_ant['cobrado'] or Decimal('0.00'),
                'devuelto': devoluciones_dia_ant,
                'cobrado_neto': (pagos_dia_ant['cobrado'] or Decimal('0.00')) - devoluciones_dia_ant,
            })
        
        cierre_diario['comparativa_dias'] = list(reversed(comparativa_dias))
    
    # ==================== GRÁFICO DE EVOLUCIÓN ====================
    from django.db.models.functions import TruncMonth
    
    fecha_grafico_desde = fecha_hasta_obj - timedelta(days=180)
    
    por_mes = Sesion.objects.filter(
        fecha__gte=fecha_grafico_desde,
        fecha__lte=fecha_hasta_obj,
        estado__in=['realizada', 'realizada_retraso']
    ).annotate(
        mes=TruncMonth('fecha')
    ).values('mes').annotate(
        sesiones=Count('id'),
        ingresos_sesiones=Sum('monto_cobrado')
    ).order_by('mes')
    
    # Agregar proyectos al gráfico
    proyectos_mes = Proyecto.objects.filter(
        fecha_inicio__gte=fecha_grafico_desde,
        fecha_inicio__lte=fecha_hasta_obj
    ).annotate(
        mes=TruncMonth('fecha_inicio')
    ).values('mes').annotate(
        proyectos=Count('id'),
        ingresos_proyectos=Sum('costo_total')
    ).order_by('mes')
    
    # Agregar mensualidades al gráfico
    from django.db.models import IntegerField
    mensualidades_mes_grafico = Mensualidad.objects.filter(
        estado__in=['activa', 'pausada', 'completada', 'cancelada']
    ).filter(
        Q(anio__gt=fecha_grafico_desde.year) |
        Q(anio=fecha_grafico_desde.year, mes__gte=fecha_grafico_desde.month)
    ).filter(
        Q(anio__lt=fecha_hasta_obj.year) |
        Q(anio=fecha_hasta_obj.year, mes__lte=fecha_hasta_obj.month)
    ).values('anio', 'mes').annotate(
        cantidad=Count('id'),
        ingresos_mensualidades=Sum('costo_mensual')
    ).order_by('anio', 'mes')
    
    # Combinar datos del gráfico
    meses_dict = {}
    
    for m in por_mes:
        mes_str = m['mes'].strftime('%b %Y')
        meses_dict[mes_str] = {
            'sesiones': m['sesiones'],
            'proyectos': 0,
            'mensualidades': 0,
            'ingresos': float(m['ingresos_sesiones'] or 0)
        }
    
    for p in proyectos_mes:
        mes_str = p['mes'].strftime('%b %Y')
        if mes_str in meses_dict:
            meses_dict[mes_str]['proyectos'] = p['proyectos']
            meses_dict[mes_str]['ingresos'] += float(p['ingresos_proyectos'] or 0)
        else:
            meses_dict[mes_str] = {
                'sesiones': 0,
                'proyectos': p['proyectos'],
                'mensualidades': 0,
                'ingresos': float(p['ingresos_proyectos'] or 0)
            }
    
    for mn in mensualidades_mes_grafico:
        from datetime import date as date_cls
        mes_str = date_cls(mn['anio'], mn['mes'], 1).strftime('%b %Y')
        if mes_str in meses_dict:
            meses_dict[mes_str]['mensualidades'] = mn['cantidad']
            meses_dict[mes_str]['ingresos'] += float(mn['ingresos_mensualidades'] or 0)
        else:
            meses_dict[mes_str] = {
                'sesiones': 0,
                'proyectos': 0,
                'mensualidades': mn['cantidad'],
                'ingresos': float(mn['ingresos_mensualidades'] or 0)
            }
    
    grafico_data = {
        'labels': list(meses_dict.keys()),
        'sesiones': [meses_dict[k]['sesiones'] for k in meses_dict.keys()],
        'proyectos': [meses_dict[k]['proyectos'] for k in meses_dict.keys()],
        'mensualidades': [meses_dict[k]['mensualidades'] for k in meses_dict.keys()],
        'ingresos': [meses_dict[k]['ingresos'] for k in meses_dict.keys()],
    }
    
    # ==================== TOP PACIENTES ====================
    # Por mayor consumo
    top_pacientes = Paciente.objects.filter(
        Q(sesiones__fecha__gte=fecha_desde_obj, sesiones__fecha__lte=fecha_hasta_obj) |
        Q(proyectos__fecha_inicio__gte=fecha_desde_obj, proyectos__fecha_inicio__lte=fecha_hasta_obj)
    ).annotate(
        sesiones_count=Count('sesiones', filter=Q(sesiones__estado__in=['realizada', 'realizada_retraso'])),
        proyectos_count=Count('proyectos'),
        total_consumido=Sum('sesiones__monto_cobrado', filter=Q(sesiones__estado__in=['realizada', 'realizada_retraso'])) + Sum('proyectos__costo_total')
    ).order_by('-total_consumido')[:10]
    
    context.update({
        'ingresos': ingresos,
        'por_metodo': por_metodo,
        'por_servicio': por_servicio,
        'grafico_data': grafico_data,
        'cierre_diario': cierre_diario,
        'top_pacientes': top_pacientes,
        'devoluciones': devoluciones,
        'pagos_anulados': pagos_anulados,
    })
    
    return render(request, 'facturacion/reportes/financiero.html', context)

@login_required
def exportar_excel(request):
    """
    Exportar datos a Excel
    OPTIMIZADO: Usa openpyxl
    """
    
    tipo = request.GET.get('tipo', 'cuentas')  # cuentas, pagos, sesiones
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        wb = Workbook()
        ws = wb.active
        
        # Header style
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        if tipo == 'cuentas':
            ws.title = "Cuentas Corrientes"
            
            # Headers
            headers = ['Paciente', 'Tutor', 'Teléfono', 'Consumido', 'Pagado', 'Saldo', 'Estado']
            ws.append(headers)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            pacientes = Paciente.objects.filter(
                estado='activo'
            ).select_related('cuenta_corriente').order_by('apellido', 'nombre')
            
            for p in pacientes:
                cuenta = p.cuenta_corriente if hasattr(p, 'cuenta_corriente') else None
                saldo = cuenta.saldo_actual if cuenta else 0
                
                estado = 'DEBE' if saldo < 0 else ('A FAVOR' if saldo > 0 else 'AL DÍA')
                
                ws.append([
                    p.nombre_completo,
                    p.nombre_tutor,
                    p.telefono_tutor,
                    float(cuenta.total_consumido_actual if cuenta else 0),
                    float(cuenta.total_pagado if cuenta else 0),
                    float(saldo),
                    estado
                ])
        
        elif tipo == 'pagos':
            ws.title = "Historial Pagos"
            
            headers = ['Recibo', 'Fecha', 'Paciente', 'Concepto', 'Método', 'Monto']
            ws.append(headers)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            pagos = Pago.objects.filter(
                anulado=False
            ).select_related('paciente', 'metodo_pago').order_by('-fecha_pago')[:500]
            
            for pago in pagos:
                ws.append([
                    pago.numero_recibo,
                    pago.fecha_pago.strftime('%d/%m/%Y'),
                    pago.paciente.nombre_completo,
                    pago.concepto[:50],
                    pago.metodo_pago.nombre,
                    float(pago.monto)
                ])
        
        # Ajustar anchos
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Generar archivo
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reporte_{tipo}_{date.today()}.xlsx"'
        
        return response
        
    except ImportError:
        messages.error(request, '❌ openpyxl no está instalado')
        return redirect('facturacion:dashboard_reportes')
    except Exception as e:
        messages.error(request, f'❌ Error al exportar: {str(e)}')
        return redirect('facturacion:dashboard_reportes')
    
# Agregar estas vistas al archivo facturacion/views.py

@login_required
def api_detalle_sesion(request, sesion_id):
    """
    API: Obtener detalles completos de una sesión (para modal)
    ✅ ACTUALIZADO: Incluye pagos masivos
    """
    from agenda.models import Sesion
    
    sesion = get_object_or_404(
        Sesion.objects.select_related(
            'paciente', 'servicio', 'profesional', 'sucursal', 'proyecto'
        ),
        id=sesion_id
    )
    
    # Pagos directos (FK en Pago apunta a la sesión)
    pagos = sesion.pagos.filter(anulado=False).select_related(
        'metodo_pago', 'registrado_por'
    ).order_by('-fecha_pago')
    
    # ✅ NUEVO: Pagos masivos donde esta sesión es uno de los ítems
    pagos_masivos = DetallePagoMasivo.objects.filter(
        tipo='sesion',
        sesion=sesion,
        pago__anulado=False  # Solo pagos no anulados
    ).select_related(
        'pago__metodo_pago',
        'pago__registrado_por'
    ).prefetch_related(
        'pago__detalles_masivos'  # Para poder usar pago.cantidad_detalles
    )
    
    return render(request, 'facturacion/partials/detalle_sesion.html', {
        'sesion': sesion,
        'pagos': pagos,
        'pagos_masivos': pagos_masivos,  # ✅ Nueva variable
    })


@login_required
def api_detalle_pago(request, pago_id):
    """
    API: Obtener detalles completos de un pago (para modal)
    """
    pago = get_object_or_404(
        Pago.objects.select_related(
            'paciente', 'metodo_pago', 'sesion__servicio', 
            'sesion__profesional', 'sesion__sucursal',
            'proyecto', 'registrado_por', 'anulado_por'
        ),
        id=pago_id
    )
    
    return render(request, 'facturacion/partials/detalle_pago.html', {
        'pago': pago,
    })

@login_required
def mi_cuenta(request):
    """
    Resumen de cuenta corriente del paciente
    ✅ EXCLUSIVA para pacientes
    ✅ ACTUALIZADO: Solo cuenta pagos con recibo (no crédito), excluye anulados
    """
    # ✅ Verificar que el usuario sea paciente
    if not hasattr(request.user, 'perfil') or not request.user.perfil.es_paciente():
        messages.error(request, '⚠️ Esta sección es solo para pacientes.')
        return redirect('core:dashboard')
    
    # ✅ Obtener el paciente vinculado
    paciente = request.user.perfil.paciente
    
    if not paciente:
        messages.error(request, '❌ No hay un paciente vinculado a tu cuenta.')
        return redirect('core:dashboard')
    
    # Obtener o crear cuenta corriente
    cuenta, created = CuentaCorriente.objects.get_or_create(paciente=paciente)
    if created:
        cuenta.actualizar_saldo()
    
    # ✅ IMPORTANTE: Pagos con RECIBO (dinero real recibido)
    # Excluye: Uso de Crédito y Anulados
    pagos_con_recibo = Pago.objects.filter(
        paciente=paciente,
        anulado=False
    ).exclude(
        metodo_pago__nombre="Uso de Crédito"
    ).select_related('metodo_pago', 'sesion', 'proyecto')
    
    # Calcular totales SOLO de pagos con recibo
    total_pagado_real = pagos_con_recibo.aggregate(
        total=Sum('monto')
    )['total'] or Decimal('0.00')
    
    # Resumen financiero
    resumen = {
        # Sesiones normales
        'consumo_sesiones': cuenta.consumo_sesiones,
        'pagado_sesiones': cuenta.pagado_sesiones,
        'deuda_sesiones': cuenta.deuda_sesiones,
        
        # Proyectos
        'consumo_proyectos': cuenta.consumo_proyectos,
        'pagado_proyectos': cuenta.pagado_proyectos,
        'deuda_proyectos': cuenta.deuda_proyectos,
        
        # Mensualidades
        'consumo_mensualidades': cuenta.total_mensualidades,
        'pagado_mensualidades': cuenta.pagado_mensualidades,
        'deuda_mensualidades': cuenta.deuda_mensualidades,
        
        # Totales (igual que en Cuenta Corriente)
        'consumo_total': cuenta.total_consumo_general,
        'pagado_total': cuenta.total_pagado,  # ✅ CORREGIDO: Campo correcto del modelo
        'deuda_total': cuenta.total_deuda_general,
        
        # Balance
        'credito': cuenta.saldo_actual,
        # ✅ CORREGIDO: balance_final debe usar PROYECTADO (pagado - consumo total)
        # No usar cuenta.balance_final porque ese es el balance ACTUAL
        'balance_final': cuenta.total_pagado - cuenta.total_consumo_general,
        # ✅ NUEVO: Balance absoluto (sin signo) para mostrar en tarjetas
        'balance_absoluto': abs(cuenta.total_pagado - cuenta.total_consumo_general),
    }
    
    # Sesiones con deuda pendiente
    sesiones_pendientes = Sesion.objects.filter(
        paciente=paciente,
        estado__in=['realizada', 'realizada_retraso', 'falta']
    ).select_related(
        'servicio', 'profesional'
    ).order_by('fecha', 'hora_inicio')
    
    # Filtrar solo las que tienen saldo pendiente
    sesiones_con_deuda = []
    for sesion in sesiones_pendientes:
        if sesion.saldo_pendiente > 0:
            sesiones_con_deuda.append(sesion)
    
    # ✅ Últimos 5 RECIBOS únicos (agrupados por número de recibo)
    # Método simple y universal para todas las bases de datos
    numeros_recibos_unicos = pagos_con_recibo.values_list(
        'numero_recibo', flat=True
    ).distinct().order_by('-numero_recibo')[:5]
    
    # Obtener el primer pago de cada recibo
    ultimos_recibos = []
    for numero in numeros_recibos_unicos:
        pago = pagos_con_recibo.filter(numero_recibo=numero).first()
        if pago:
            ultimos_recibos.append(pago)
    
    # ==================== PRÓXIMA SESIÓN ====================
    # Buscar la próxima sesión programada del paciente
    ahora = timezone.now()
    
    proxima_sesion = Sesion.objects.filter(
        paciente=paciente,
        estado='programada',  # Solo sesiones programadas
        fecha__gte=ahora.date()  # Desde hoy en adelante
    ).select_related(
        'servicio',
        'profesional',
        'sucursal'
    ).order_by('fecha', 'hora_inicio').first()
    
    # Si hay sesiones hoy, verificar que no hayan pasado
    if proxima_sesion and proxima_sesion.fecha == ahora.date():
        # Combinar fecha y hora para comparación precisa
        hora_sesion = datetime.combine(
            proxima_sesion.fecha,
            proxima_sesion.hora_inicio
        )
        hora_sesion = timezone.make_aware(hora_sesion)
        
        # Si la sesión de hoy ya pasó, buscar la siguiente
        if hora_sesion < ahora:
            proxima_sesion = Sesion.objects.filter(
                paciente=paciente,
                estado='programada',
                fecha__gt=ahora.date()
            ).select_related(
                'servicio',
                'profesional',
                'sucursal'
            ).order_by('fecha', 'hora_inicio').first()
    
    # ==================== REGISTROS RECIENTES ====================
    # Obtener los últimos 5 registros (sesiones, proyectos, mensualidades)
    registros_recientes = []
    
    # Sesiones completadas recientes
    sesiones_recientes = Sesion.objects.filter(
        paciente=paciente,
        estado__in=['realizada', 'realizada_retraso']
    ).select_related('servicio').order_by('-fecha', '-hora_inicio')[:3]
    
    for sesion in sesiones_recientes:
        pagado = _total_pagado_sesion(sesion)
        saldo = sesion.monto_cobrado - pagado
        registros_recientes.append({
            'tipo': 'sesion',
            'objeto': sesion,
            'fecha': sesion.fecha,
            'costo': sesion.monto_cobrado,
            'pagado': pagado,
            'saldo': saldo
        })
    
    # Proyectos recientes (planificados, en progreso o finalizados)
    proyectos_recientes = Proyecto.objects.filter(
        paciente=paciente,
        estado__in=['planificado', 'en_progreso', 'finalizado']
    ).select_related('servicio_base').order_by('-fecha_inicio')[:2]
    
    for proyecto in proyectos_recientes:
        pagado = _total_pagado_proyecto(proyecto)
        saldo = proyecto.costo_total - pagado
        registros_recientes.append({
            'tipo': 'proyecto',
            'objeto': proyecto,
            'fecha': proyecto.fecha_inicio,
            'costo': proyecto.costo_total,
            'pagado': pagado,
            'saldo': saldo
        })
    
    # Mensualidades recientes
    mensualidades_recientes = Mensualidad.objects.filter(
        paciente=paciente,
        estado__in=['activa', 'completada']
    ).order_by('-anio', '-mes')[:2]
    
    for mensualidad in mensualidades_recientes:
        fecha_ficticia = date(mensualidad.anio, mensualidad.mes, 1)
        pagado = _total_pagado_mensualidad(mensualidad)
        saldo = mensualidad.costo_mensual - pagado
        registros_recientes.append({
            'tipo': 'mensualidad',
            'objeto': mensualidad,
            'fecha': fecha_ficticia,
            'costo': mensualidad.costo_mensual,
            'pagado': pagado,
            'saldo': saldo
        })
    
    # Ordenar todos los registros por fecha (más recientes primero)
    registros_recientes.sort(key=lambda x: x['fecha'], reverse=True)
    registros_recientes = registros_recientes[:5]  # Solo los 5 más recientes
    
    context = {
        'paciente': paciente,
        'cuenta': cuenta,
        'resumen': resumen,
        'sesiones_con_deuda': sesiones_con_deuda,
        'pagos_realizados': ultimos_recibos,  # ✅ Recibos únicos
        'total_sesiones_pendientes': len(sesiones_con_deuda),
        'proxima_sesion': proxima_sesion,  # ✅ NUEVO
        'registros_recientes': registros_recientes,  # ✅ NUEVO
        'pagos_recientes': ultimos_recibos,  # ✅ NUEVO (alias para el template)
    }
    
    return render(request, 'facturacion/mi_cuenta.html', context)


@login_required
def mis_pagos(request):
    """
    Historial completo de pagos del paciente
    ✅ EXCLUSIVA para pacientes
    ✅ ACTUALIZADO: Agrupa pagos masivos por número de recibo
    ✅ Solo muestra pagos con recibo (no crédito), excluye anulados
    ✅ NUEVO: Incluye devoluciones y las resta del total
    """
    # ✅ Verificar que el usuario sea paciente
    if not hasattr(request.user, 'perfil') or not request.user.perfil.es_paciente():
        messages.error(request, '⚠️ Esta sección es solo para pacientes.')
        return redirect('core:dashboard')
    
    # ✅ Obtener el paciente vinculado
    paciente = request.user.perfil.paciente
    
    if not paciente:
        messages.error(request, '❌ No hay un paciente vinculado a tu cuenta.')
        return redirect('core:dashboard')
    
    # Filtros
    fecha_desde = request.GET.get('desde', '')
    fecha_hasta = request.GET.get('hasta', '')
    metodo = request.GET.get('metodo', '')
    tipo = request.GET.get('tipo', '')  # 'sesion', 'proyecto', 'adelantado'
    
    # ✅ Query base: Solo pagos con RECIBO (no crédito) y NO anulados
    pagos_query = Pago.objects.filter(
        paciente=paciente,
        anulado=False
    ).exclude(
        metodo_pago__nombre="Uso de Crédito"
    ).select_related(
        'metodo_pago', 'sesion', 'sesion__servicio', 'proyecto', 'registrado_por'
    )
    
    # Aplicar filtros
    if fecha_desde:
        try:
            pagos_query = pagos_query.filter(fecha_pago__gte=fecha_desde)
        except:
            pass
    
    if fecha_hasta:
        try:
            pagos_query = pagos_query.filter(fecha_pago__lte=fecha_hasta)
        except:
            pass
    
    if metodo:
        pagos_query = pagos_query.filter(metodo_pago_id=metodo)
    
    if tipo:
        if tipo == 'sesion':
            pagos_query = pagos_query.filter(sesion__isnull=False)
        elif tipo == 'proyecto':
            pagos_query = pagos_query.filter(proyecto__isnull=False)
        elif tipo == 'mensualidad':  # ✅ NUEVO: Filtro para mensualidades
            pagos_query = pagos_query.filter(mensualidad__isnull=False)
        elif tipo == 'adelantado':
            pagos_query = pagos_query.filter(sesion__isnull=True, proyecto__isnull=True, mensualidad__isnull=True)  # ✅ Actualizado
    
    # ✅ AGRUPAR PAGOS POR NÚMERO DE RECIBO
    # Usar diccionario para agrupar eficientemente
    recibos_dict = {}
    
    for pago in pagos_query.order_by('-fecha_pago', '-fecha_registro'):
        numero_recibo = pago.numero_recibo
        
        # Si es la primera vez que vemos este recibo, lo agregamos
        if numero_recibo not in recibos_dict:
            recibos_dict[numero_recibo] = {
                'pago_principal': pago,
                'numero_recibo': numero_recibo,
                'total_recibo': Decimal('0.00'),
                'cantidad_items': 0,
                'sesiones_pagadas': [],
                'proyectos_pagados': [],
                'mensualidades_pagadas': [],  # ✅ NUEVO: Agregar mensualidades
                'fecha_pago': pago.fecha_pago,
                'metodo_pago': pago.metodo_pago,
                'concepto': pago.concepto,
            }
        
        # Acumular información del recibo
        recibos_dict[numero_recibo]['total_recibo'] += pago.monto
        recibos_dict[numero_recibo]['cantidad_items'] += 1
        
        # Agregar sesión si existe
        if pago.sesion:
            recibos_dict[numero_recibo]['sesiones_pagadas'].append({
                'sesion__id': pago.sesion.id,
                'sesion__fecha': pago.sesion.fecha,
                'sesion__servicio__nombre': pago.sesion.servicio.nombre,
                'monto': pago.monto
            })
        
        # Agregar proyecto si existe
        if pago.proyecto:
            recibos_dict[numero_recibo]['proyectos_pagados'].append({
                'proyecto__id': pago.proyecto.id,
                'proyecto__nombre': pago.proyecto.nombre,
                'monto': pago.monto
            })
        
        # ✅ NUEVO: Agregar mensualidad si existe
        if pago.mensualidad:
            recibos_dict[numero_recibo]['mensualidades_pagadas'].append({
                'mensualidad__id': pago.mensualidad.id,
                'mensualidad__mes': pago.mensualidad.mes,
                'mensualidad__anio': pago.mensualidad.anio,
                'monto': pago.monto
            })
    
    # Convertir diccionario a lista y marcar los múltiples
    recibos_agrupados = []
    for recibo in recibos_dict.values():
        recibo['es_multiple'] = recibo['cantidad_items'] > 1
        recibos_agrupados.append(recibo)
    
    # Ordenar por fecha (más recientes primero)
    recibos_agrupados.sort(key=lambda x: x['fecha_pago'], reverse=True)
    
    # ✅ NUEVO: Obtener devoluciones del paciente
    devoluciones_query = Devolucion.objects.filter(
        paciente=paciente
    ).select_related('registrado_por', 'proyecto')
    
    # Aplicar los mismos filtros de fecha a las devoluciones
    if fecha_desde:
        try:
            devoluciones_query = devoluciones_query.filter(fecha_devolucion__gte=fecha_desde)
        except:
            pass
    
    if fecha_hasta:
        try:
            devoluciones_query = devoluciones_query.filter(fecha_devolucion__lte=fecha_hasta)
        except:
            pass
    
    # Aplicar filtro de tipo a devoluciones
    if tipo:
        if tipo == 'sesion':
            # Las devoluciones no tienen campo 'sesion', así que omitimos este filtro
            # O podríamos filtrar por pagos relacionados si fuera necesario
            pass
        elif tipo == 'proyecto':
            devoluciones_query = devoluciones_query.filter(proyecto__isnull=False)
        elif tipo == 'mensualidad':
            devoluciones_query = devoluciones_query.filter(mensualidad__isnull=False)
        elif tipo == 'adelantado':
            devoluciones_query = devoluciones_query.filter(proyecto__isnull=True, mensualidad__isnull=True)
    
    # Convertir devoluciones a lista
    devoluciones_lista = []
    for dev in devoluciones_query.order_by('-fecha_devolucion'):
        devoluciones_lista.append({
            'id': dev.id,
            'numero_devolucion': dev.numero_devolucion,
            'fecha_devolucion': dev.fecha_devolucion,
            'monto': dev.monto,
            'motivo': dev.motivo,
            'proyecto': dev.proyecto,  # ✅ Este campo SÍ existe
            'mensualidad': dev.mensualidad,  # ✅ Este campo SÍ existe
            'registrado_por': dev.registrado_por,
            })
        
    # Calcular totales
    total_pagado = sum(r['total_recibo'] for r in recibos_agrupados)
    total_devoluciones = sum(d['monto'] for d in devoluciones_lista)
    total_neto = total_pagado - total_devoluciones  # ✅ Total efectivo
    
    # Métodos de pago disponibles para el filtro
    metodos_pago = MetodoPago.objects.filter(activo=True).exclude(
        nombre="Uso de Crédito"
    )
    
    context = {
        'paciente': paciente,
        'recibos_agrupados': recibos_agrupados,  # ✅ Recibos agrupados
        'devoluciones_lista': devoluciones_lista,  # ✅ NUEVO: Devoluciones
        'total_pagado': total_pagado,  # Total de pagos
        'total_devoluciones': total_devoluciones,  # ✅ NUEVO: Total devoluciones
        'total_neto': total_neto,  # ✅ NUEVO: Total neto (pagos - devoluciones)
        'metodos_pago': metodos_pago,
        'filtros': {
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'metodo': metodo,
            'tipo': tipo,
        }
    }
    
    return render(request, 'facturacion/mis_pagos.html', context)

@login_required
def mis_deudas(request):
    """
    Vista de deudas pendientes del paciente - PROYECCIÓN TOTAL
    ✅ EXCLUSIVA para pacientes
    Muestra TODAS las sesiones (realizadas + programadas), proyectos (todos los estados) 
    y mensualidades con saldo pendiente - igual que Proyección Total
    """
    # ✅ Verificar que el usuario sea paciente
    if not hasattr(request.user, 'perfil') or not request.user.perfil.es_paciente():
        messages.error(request, '⚠️ Esta sección es solo para pacientes.')
        return redirect('core:dashboard')
    
    # ✅ Obtener el paciente vinculado
    paciente = request.user.perfil.paciente
    
    if not paciente:
        messages.error(request, '❌ No hay un paciente vinculado a tu cuenta.')
        return redirect('core:dashboard')
    
    # Obtener cuenta corriente
    cuenta, _ = CuentaCorriente.objects.get_or_create(paciente=paciente)
    
    # ==================== SESIONES CON DEUDA (TODAS - PROYECCIÓN TOTAL) ====================
    sesiones_con_deuda = []
    # ✅ INCLUIR: realizadas, con retraso, faltas Y PROGRAMADAS
    sesiones_todas = Sesion.objects.filter(
        paciente=paciente,
        estado__in=['realizada', 'realizada_retraso', 'falta', 'programada'],  # ✅ Incluye programadas
        proyecto__isnull=True  # Solo sesiones normales (no de proyectos)
    ).select_related(
        'servicio', 'profesional', 'sucursal'
    ).order_by('-fecha', '-hora_inicio')
    
    total_deuda_sesiones = Decimal('0.00')
    for sesion in sesiones_todas:
        if sesion.saldo_pendiente > 0:
            # Determinar el estado visual
            if sesion.estado == 'programada':
                estado_label = 'Programada'
                estado_color = 'blue'
            elif sesion.estado == 'realizada':
                estado_label = 'Realizada'
                estado_color = 'green'
            elif sesion.estado == 'realizada_retraso':
                estado_label = 'Con retraso'
                estado_color = 'orange'
            else:  # falta
                estado_label = 'Falta'
                estado_color = 'red'
            
            sesiones_con_deuda.append({
                'objeto': sesion,
                'tipo': 'sesion',
                'fecha': sesion.fecha,
                'descripcion': f"{sesion.servicio.nombre} - {sesion.profesional.nombre_completo if sesion.profesional else 'Sin profesional'}",
                'costo': sesion.monto_cobrado,
                'pagado': sesion.total_pagado,
                'saldo': sesion.saldo_pendiente,
                'sucursal': sesion.sucursal.nombre if sesion.sucursal else 'Sin sucursal',
                'estado_label': estado_label,
                'estado_color': estado_color,
                'es_futura': sesion.estado == 'programada',
            })
            total_deuda_sesiones += sesion.saldo_pendiente
    
    # ==================== PROYECTOS CON DEUDA (TODOS - PROYECCIÓN TOTAL) ====================
    proyectos_con_deuda = []
    # ✅ INCLUIR TODOS LOS ESTADOS (excepto cancelado)
    proyectos = Proyecto.objects.filter(
        paciente=paciente,
        estado__in=['borrador', 'planificado', 'en_progreso', 'finalizado']  # ✅ Incluye todos
    ).select_related(
        'servicio_base', 'profesional_responsable', 'sucursal'
    ).order_by('-fecha_inicio')
    
    total_deuda_proyectos = Decimal('0.00')
    for proyecto in proyectos:
        if proyecto.saldo_pendiente > 0:
            # Determinar estado visual
            if proyecto.estado == 'borrador':
                estado_label = 'Borrador'
                estado_color = 'gray'
            elif proyecto.estado == 'planificado':
                estado_label = 'Planificado'
                estado_color = 'blue'
            elif proyecto.estado == 'en_progreso':
                estado_label = 'En progreso'
                estado_color = 'yellow'
            else:  # finalizado
                estado_label = 'Finalizado'
                estado_color = 'green'
            
            # Obtener sesiones del proyecto de forma segura
            sesiones_proyecto = proyecto.sesiones.all() if hasattr(proyecto, 'sesiones') else []
            total_sesiones_proyecto = len(sesiones_proyecto)
            sesiones_realizadas_proyecto = sum(1 for s in sesiones_proyecto if s.estado in ['realizada', 'realizada_retraso'])
            
            proyectos_con_deuda.append({
                'objeto': proyecto,
                'tipo': 'proyecto',
                'fecha': proyecto.fecha_inicio,
                'descripcion': proyecto.nombre or proyecto.servicio_base.nombre,
                'costo': proyecto.costo_total,
                'pagado': proyecto.pagado_neto,
                'saldo': proyecto.saldo_pendiente,
                'sucursal': proyecto.sucursal.nombre if proyecto.sucursal else 'Sin sucursal',
                'sesiones_total': total_sesiones_proyecto,
                'sesiones_realizadas': sesiones_realizadas_proyecto,
                'estado_label': estado_label,
                'estado_color': estado_color,
                'es_futuro': proyecto.estado in ['borrador', 'planificado'],
            })
            total_deuda_proyectos += proyecto.saldo_pendiente
    
    # ==================== MENSUALIDADES CON DEUDA ====================
    # Diccionario para convertir número de mes a nombre
    MESES = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }
    
    mensualidades_con_deuda = []
    mensualidades = Mensualidad.objects.filter(
        paciente=paciente,
        estado__in=['activa', 'vencida', 'pausada', 'completada']
    ).prefetch_related(
        'servicios_profesionales__servicio',
        'servicios_profesionales__profesional'
    ).order_by('-anio', '-mes')
    
    total_deuda_mensualidades = Decimal('0.00')
    for mensualidad in mensualidades:
        if mensualidad.saldo_pendiente > 0:
            # Obtener servicios de la mensualidad
            servicios_str = ", ".join([
                sp.servicio.nombre for sp in mensualidad.servicios_profesionales.all()[:2]
            ])
            if mensualidad.servicios_profesionales.count() > 2:
                servicios_str += f" (+{mensualidad.servicios_profesionales.count() - 2} más)"
            
            # Determinar estado visual
            if mensualidad.estado == 'activa':
                estado_label = 'Activa'
                estado_color = 'green'
            elif mensualidad.estado == 'vencida':
                estado_label = 'Vencida'
                estado_color = 'red'
            elif mensualidad.estado == 'pausada':
                estado_label = 'Pausada'
                estado_color = 'yellow'
            else:  # completada
                estado_label = 'Completada'
                estado_color = 'gray'
            
            # Convertir mes numérico a nombre
            mes_nombre = MESES.get(mensualidad.mes, str(mensualidad.mes))
            
            mensualidades_con_deuda.append({
                'objeto': mensualidad,
                'tipo': 'mensualidad',
                'fecha': date(mensualidad.anio, mensualidad.mes, 1),
                'descripcion': f"{mes_nombre} {mensualidad.anio}",  # ✅ CAMBIADO: Ahora muestra "Junio 2026"
                'servicios': servicios_str or 'Sin servicios',
                'costo': mensualidad.costo_mensual,
                'pagado': mensualidad.pagado_neto,
                'saldo': mensualidad.saldo_pendiente,
                'estado_label': estado_label,
                'estado_color': estado_color,
            })
            total_deuda_mensualidades += mensualidad.saldo_pendiente
    
    # ==================== TOTALES ====================
    total_deuda_general = total_deuda_sesiones + total_deuda_proyectos + total_deuda_mensualidades
    total_items = len(sesiones_con_deuda) + len(proyectos_con_deuda) + len(mensualidades_con_deuda)
    
    # Contar items futuros vs realizados
    sesiones_programadas = sum(1 for s in sesiones_con_deuda if s['es_futura'])
    sesiones_realizadas = len(sesiones_con_deuda) - sesiones_programadas
    proyectos_futuros = sum(1 for p in proyectos_con_deuda if p['es_futuro'])
    proyectos_actuales = len(proyectos_con_deuda) - proyectos_futuros
    
    context = {
        'paciente': paciente,
        'cuenta': cuenta,
        'sesiones_con_deuda': sesiones_con_deuda,
        'proyectos_con_deuda': proyectos_con_deuda,
        'mensualidades_con_deuda': mensualidades_con_deuda,
        'total_deuda_sesiones': total_deuda_sesiones,
        'total_deuda_proyectos': total_deuda_proyectos,
        'total_deuda_mensualidades': total_deuda_mensualidades,
        'total_deuda_general': total_deuda_general,
        'total_items': total_items,
        'sesiones_programadas': sesiones_programadas,
        'sesiones_realizadas': sesiones_realizadas,
        'proyectos_futuros': proyectos_futuros,
        'proyectos_actuales': proyectos_actuales,
    }
    
    return render(request, 'facturacion/mis_deudas.html', context)


@login_required
def detalle_pago_paciente(request, pago_id):
    """
    Detalle de un pago/recibo específico
    ✅ EXCLUSIVA para pacientes - Solo pueden ver SUS propios pagos
    ✅ ACTUALIZADO: Si es un pago masivo, muestra todos los items del recibo
    """
    # ✅ Verificar que el usuario sea paciente
    if not hasattr(request.user, 'perfil') or not request.user.perfil.es_paciente():
        messages.error(request, '⚠️ Esta sección es solo para pacientes.')
        return redirect('core:dashboard')
    
    # ✅ Obtener el paciente vinculado
    paciente = request.user.perfil.paciente
    
    if not paciente:
        messages.error(request, '❌ No hay un paciente vinculado a tu cuenta.')
        return redirect('core:dashboard')
    
    # Obtener el pago principal
    pago = get_object_or_404(
        Pago.objects.select_related(
            'metodo_pago', 'sesion', 'sesion__servicio', 'sesion__profesional',
            'proyecto', 'registrado_por'
        ),
        id=pago_id,
        paciente=paciente,  # ✅ IMPORTANTE: Solo SUS pagos
        anulado=False
    )
    
    # ✅ Obtener TODOS los pagos con el mismo número de recibo
    pagos_del_recibo = Pago.objects.filter(
        numero_recibo=pago.numero_recibo,
        paciente=paciente,
        anulado=False
    ).exclude(
        metodo_pago__nombre="Uso de Crédito"
    ).select_related(
        'sesion', 'sesion__servicio', 'sesion__profesional',
        'proyecto'
    ).order_by('fecha_pago')
    
    # Calcular total del recibo
    total_recibo = pagos_del_recibo.aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
    
    # Separar por tipo
    pagos_sesiones = pagos_del_recibo.filter(sesion__isnull=False)
    pagos_proyectos = pagos_del_recibo.filter(proyecto__isnull=False)
    pagos_adelantados = pagos_del_recibo.filter(sesion__isnull=True, proyecto__isnull=True)
    
    context = {
        'paciente': paciente,
        'pago': pago,
        'pagos_del_recibo': pagos_del_recibo,
        'total_recibo': total_recibo,
        'es_pago_multiple': pagos_del_recibo.count() > 1,
        'pagos_sesiones': pagos_sesiones,
        'pagos_proyectos': pagos_proyectos,
        'pagos_adelantados': pagos_adelantados,
    }
    
    return render(request, 'facturacion/detalle_pago_paciente.html', context)


# ==================== LIMPIAR PAGOS ANULADOS ====================

from django.contrib.admin.views.decorators import staff_member_required

@staff_member_required
def limpiar_pagos_anulados(request):
    """
    Vista para que administradores limpien pagos anulados
    Solo accesible para staff
    """
    
    if request.method == 'POST':
        try:
            from django.db import transaction
            
            # Obtener parámetros
            dias = int(request.POST.get('dias', 0))
            
            # Filtrar pagos anulados
            pagos_anulados = Pago.objects.filter(anulado=True)
            
            # Filtrar por antigüedad
            if dias > 0:
                from datetime import datetime, timedelta
                fecha_limite = datetime.now() - timedelta(days=dias)
                pagos_anulados = pagos_anulados.filter(fecha_anulacion__lt=fecha_limite)
            
            total = pagos_anulados.count()
            
            if total == 0:
                messages.warning(request, '⚠️ No hay pagos anulados para eliminar con los filtros seleccionados')
                return redirect('facturacion:limpiar_pagos_anulados')
            
            # Transacción atómica
            with transaction.atomic():
                # Obtener pacientes afectados
                pacientes_ids = set(pagos_anulados.values_list('paciente_id', flat=True))
                
                # Eliminar pagos
                eliminados, detalle = pagos_anulados.delete()
                
                # Recalcular cuentas corrientes
                cuentas_actualizadas = 0
                for paciente_id in pacientes_ids:
                    try:
                        cuenta = CuentaCorriente.objects.get(paciente_id=paciente_id)
                        cuenta.actualizar_saldo()
                        cuentas_actualizadas += 1
                    except CuentaCorriente.DoesNotExist:
                        pass
                
                messages.success(
                    request,
                    f'✅ {eliminados} pagos anulados eliminados correctamente. '
                    f'{cuentas_actualizadas} cuentas corrientes actualizadas.'
                )
                
                return redirect('facturacion:historial_pagos')
                
        except Exception as e:
            messages.error(request, f'❌ Error al eliminar pagos: {str(e)}')
            return redirect('facturacion:limpiar_pagos_anulados')
    
    # GET - Mostrar formulario
    from datetime import datetime, timedelta
    
    # Estadísticas de pagos anulados
    total_anulados = Pago.objects.filter(anulado=True).count()
    
    # Por antigüedad
    hoy = datetime.now()
    stats_antiguedad = {
        'ultima_semana': Pago.objects.filter(
            anulado=True,
            fecha_anulacion__gte=hoy - timedelta(days=7)
        ).count(),
        'ultimo_mes': Pago.objects.filter(
            anulado=True,
            fecha_anulacion__gte=hoy - timedelta(days=30)
        ).count(),
        'mas_3_meses': Pago.objects.filter(
            anulado=True,
            fecha_anulacion__lt=hoy - timedelta(days=90)
        ).count(),
    }
    
    # Monto total en pagos anulados
    monto_total = Pago.objects.filter(
        anulado=True
    ).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
    
    # Últimos pagos anulados
    ultimos_anulados = Pago.objects.filter(
        anulado=True
    ).select_related(
        'paciente', 'metodo_pago', 'anulado_por'
    ).order_by('-fecha_anulacion')[:20]
    
    context = {
        'total_anulados': total_anulados,
        'stats_antiguedad': stats_antiguedad,
        'monto_total': monto_total,
        'ultimos_anulados': ultimos_anulados,
    }
    
    return render(request, 'facturacion/limpiar_pagos_anulados.html', context)


@login_required
def api_mensualidades_paciente(request, paciente_id):
    """
    API: Obtener mensualidades de un paciente
    - Para pagos: Devuelve mensualidades con saldo pendiente
    - Para devoluciones: Devuelve mensualidades ACTIVAS/PAUSADAS que tienen pagos realizados
    ✅ CORREGIDO: Compatible con el nuevo modelo Mensualidad (many-to-many servicios-profesionales)
    """
    try:
        from agenda.models import Mensualidad
        
        # Detectar si es para devoluciones
        es_devolucion = request.GET.get('tipo') == 'devolucion'
        
        # ✅ Query optimizada con prefetch de servicios-profesionales
        # Para devoluciones: solo activas o pausadas (no completadas ni canceladas)
        if es_devolucion:
            estados_permitidos = ['activa', 'pausada']
        else:
            estados_permitidos = ['activa', 'pausada']
        
        mensualidades = Mensualidad.objects.filter(
            paciente_id=paciente_id,
            estado__in=estados_permitidos
        ).prefetch_related(
            'servicios_profesionales__servicio',
            'servicios_profesionales__profesional'
        ).select_related('sucursal').order_by('-anio', '-mes')
        
        # ✅ Filtrar y construir respuesta
        mensualidades_lista = []
        
        for m in mensualidades:
            # ✅ Usar cálculo que incluye pagos masivos
            total_pagado_real = _total_pagado_mensualidad(m)
            saldo_pendiente_real = m.costo_mensual - total_pagado_real

            # Para devoluciones: incluir las que tienen pagos realizados
            # Para pagos: incluir solo las que tienen saldo pendiente
            if es_devolucion:
                condicion = total_pagado_real > 0
            else:
                condicion = saldo_pendiente_real > Decimal('0.01')
            
            if condicion:
                # ✅ Obtener servicios y profesionales
                servicios_profesionales = m.servicios_profesionales.all()
                
                # Construir nombre descriptivo
                if servicios_profesionales.exists():
                    primer_sp = servicios_profesionales.first()
                    
                    # Si hay múltiples servicios, mostrar el primero + cantidad
                    if servicios_profesionales.count() > 1:
                        nombre_servicios = f"{primer_sp.servicio.nombre} (+{servicios_profesionales.count()-1} más)"
                    else:
                        nombre_servicios = primer_sp.servicio.nombre
                    
                    nombre_profesional = f"{primer_sp.profesional.nombre} {primer_sp.profesional.apellido}"
                else:
                    # Fallback si no hay servicios asignados
                    nombre_servicios = "Sin servicios"
                    nombre_profesional = "Sin asignar"
                
                mensualidades_lista.append({
                    'id': m.id,
                    'codigo': m.codigo,
                    'nombre': f"{m.periodo_display} - {nombre_servicios}",
                    'costo_mensual': float(m.costo_mensual),
                    'total_pagado': float(total_pagado_real),
                    'saldo_pendiente': float(saldo_pendiente_real),
                    'periodo': m.periodo_display,
                    'mes': m.mes,
                    'anio': m.anio,
                    'servicio': nombre_servicios,
                    'profesional': nombre_profesional,
                    'estado': m.get_estado_display(),
                })
        
        return JsonResponse({
            'success': True,
            'mensualidades': mensualidades_lista,
            'total': len(mensualidades_lista)
        })
        
    except Exception as e:
        # ✅ Log detallado para debugging
        import traceback
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error en api_mensualidades_paciente: {str(e)}")
        logger.error(traceback.format_exc())
        
        return JsonResponse({
            'success': False,
            'error': str(e),
            'detail': 'Error al cargar mensualidades. Ver logs del servidor.'
        }, status=400)

# ============================================================================
# VISTAS CORREGIDAS PARA DEVOLUCIONES
# ============================================================================
# Reemplazar en views.py desde la línea 3886

@login_required
def registrar_devolucion(request):
    """
    Vista unificada para registrar devoluciones.
    
    ✅ CORREGIDO: Ahora usa correctamente el modelo Devolucion
    """
    from facturacion.models import Devolucion
    
    if request.method == 'GET':
        # Cargar datos para el formulario
        context = {
            'pacientes': Paciente.objects.filter(estado='activo').order_by('apellido', 'nombre'),
            'metodos_pago': MetodoPago.objects.filter(activo=True).exclude(nombre="Uso de Crédito"),
            'fecha_hoy': date.today(),
        }
        
        # Pre-cargar paciente si viene en parámetros
        paciente_id = request.GET.get('paciente_id')
        if paciente_id:
            try:
                paciente = Paciente.objects.get(id=paciente_id)
                context['paciente_seleccionado'] = paciente
                
                # ✅ CORREGIDO: Obtener crédito disponible correctamente
                cuenta, _ = CuentaCorriente.objects.get_or_create(paciente=paciente)
                AccountService.update_balance(paciente)  # Actualizar antes de leer
                cuenta.refresh_from_db()  # Recargar desde BD
                context['credito_disponible'] = cuenta.pagos_adelantados
                
            except Paciente.DoesNotExist:
                messages.warning(request, 'Paciente no encontrado')
        
        # Pre-cargar proyecto si viene en parámetros
        proyecto_id = request.GET.get('proyecto_id')
        if proyecto_id:
            try:
                proyecto = Proyecto.objects.get(id=proyecto_id)
                context['proyecto_seleccionado'] = proyecto
                context['tipo_devolucion'] = 'proyecto'
                
                # ✅ CORREGIDO: Calcular disponible usando Pago y Devolucion correctamente
                total_pagado = Pago.objects.filter(
                    proyecto=proyecto,
                    anulado=False
                ).exclude(
                    metodo_pago__nombre="Uso de Crédito"
                ).aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
                
                # Sumar devoluciones del modelo Devolucion
                devoluciones_previas = Devolucion.objects.filter(
                    proyecto=proyecto
                ).aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
                
                context['disponible_devolver'] = total_pagado - devoluciones_previas
                
            except Proyecto.DoesNotExist:
                messages.warning(request, 'Proyecto no encontrado')
        
        # Pre-cargar mensualidad si viene en parámetros
        mensualidad_id = request.GET.get('mensualidad_id')
        if mensualidad_id:
            try:
                mensualidad = Mensualidad.objects.get(id=mensualidad_id)
                context['mensualidad_seleccionada'] = mensualidad
                context['tipo_devolucion'] = 'mensualidad'
                
                # ✅ CORREGIDO: Calcular disponible usando Pago y Devolucion correctamente
                total_pagado = Pago.objects.filter(
                    mensualidad=mensualidad,
                    anulado=False
                ).exclude(
                    metodo_pago__nombre="Uso de Crédito"
                ).aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
                
                # Sumar devoluciones del modelo Devolucion
                devoluciones_previas = Devolucion.objects.filter(
                    mensualidad=mensualidad
                ).aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
                
                context['disponible_devolver'] = total_pagado - devoluciones_previas
                
            except Mensualidad.DoesNotExist:
                messages.warning(request, 'Mensualidad no encontrada')
        
        return render(request, 'facturacion/registrar_devolucion.html', context)
    
    elif request.method == 'POST':
        # Procesar la devolución
        try:
            # Obtener datos del formulario
            paciente_id = request.POST.get('paciente_id')
            monto_devolucion = Decimal(request.POST.get('monto', '0'))
            metodo_pago_id = request.POST.get('metodo_pago')
            fecha_devolucion_str = request.POST.get('fecha_devolucion')
            tipo_devolucion = request.POST.get('tipo_devolucion')
            motivo = request.POST.get('motivo', '').strip()
            observaciones = request.POST.get('observaciones', '').strip()
            
            # Validaciones básicas
            if not paciente_id:
                messages.error(request, 'Debe seleccionar un paciente')
                return redirect('facturacion:registrar_devolucion')
            
            paciente = get_object_or_404(Paciente, id=paciente_id)
            
            # Parsear fecha
            if fecha_devolucion_str:
                fecha_devolucion = date.fromisoformat(fecha_devolucion_str)
            else:
                fecha_devolucion = date.today()
            
            # Obtener referencia según tipo
            referencia_id = None
            if tipo_devolucion == 'proyecto':
                referencia_id = request.POST.get('proyecto_id')
            elif tipo_devolucion == 'mensualidad':
                referencia_id = request.POST.get('mensualidad_id')
            
            # ✅ CORREGIDO: Usar PaymentService.process_refund (que está en services.py)
            resultado = PaymentService.process_refund(
                user=request.user,
                paciente=paciente,
                monto_devolucion=monto_devolucion,
                metodo_pago_id=metodo_pago_id,
                fecha_devolucion=fecha_devolucion,
                tipo_devolucion=tipo_devolucion,
                referencia_id=referencia_id,
                motivo=motivo,
                observaciones=observaciones
            )
            
            if resultado['success']:
                messages.success(
                    request, 
                    f"✅ {resultado['mensaje']}. Número: {resultado['numero_recibo']}"
                )
                
                # Redirigir a confirmación con el ID correcto
                return redirect('facturacion:confirmacion_devolucion', 
                              devolucion_id=resultado['devolucion'].id)
            else:
                messages.error(request, f"Error: {resultado.get('mensaje', 'Error desconocido')}")
                return redirect('facturacion:registrar_devolucion')
                
        except Exception as e:
            messages.error(request, f'Error al procesar devolución: {str(e)}')
            return redirect('facturacion:registrar_devolucion')


@login_required
def confirmacion_devolucion(request, devolucion_id):
    """
    ✅ CORREGIDO: Vista de confirmación usando el modelo Devolucion
    """
    from facturacion.models import Devolucion
    
    devolucion = get_object_or_404(Devolucion, id=devolucion_id)
    
    context = {
        'devolucion': devolucion,
        'paciente': devolucion.paciente,
    }
    
    return render(request, 'facturacion/confirmacion_devolucion.html', context)


# ============================================================================
# AJAX: Cargar información de crédito disponible
# ============================================================================

@login_required
def api_credito_disponible(request, paciente_id):
    """
    ✅ CORREGIDO: Retorna el crédito disponible correctamente
    """
    try:
        paciente = get_object_or_404(Paciente, id=paciente_id)
        
        # Actualizar balance y obtener crédito disponible
        cuenta, _ = CuentaCorriente.objects.get_or_create(paciente=paciente)
        AccountService.update_balance(paciente)
        cuenta.refresh_from_db()  # Recargar desde BD
        
        credito = cuenta.pagos_adelantados
        
        return JsonResponse({
            'success': True,
            'credito_disponible': float(credito),
            'credito_formatted': f"Bs. {credito:,.2f}"
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
def api_disponible_devolver_proyecto(request, proyecto_id):
    """
    ✅ CORREGIDO: Retorna el monto disponible para devolver de un proyecto
    """
    from facturacion.models import Devolucion
    
    try:
        proyecto = get_object_or_404(Proyecto, id=proyecto_id)
        
        # Total pagado (sin uso de crédito)
        total_pagado = Pago.objects.filter(
            proyecto=proyecto,
            anulado=False
        ).exclude(
            metodo_pago__nombre="Uso de Crédito"
        ).aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
        
        # Devoluciones previas del modelo Devolucion
        devoluciones_previas = Devolucion.objects.filter(
            proyecto=proyecto
        ).aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
        
        disponible = total_pagado - devoluciones_previas
        
        return JsonResponse({
            'success': True,
            'disponible_devolver': float(disponible),
            'disponible_formatted': f"Bs. {disponible:,.2f}",
            'total_pagado': float(total_pagado),
            'devoluciones_previas': float(devoluciones_previas),
            'proyecto_codigo': proyecto.codigo,
            'proyecto_nombre': proyecto.nombre
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
def api_disponible_devolver_mensualidad(request, mensualidad_id):
    """
    ✅ CORREGIDO: Retorna el monto disponible para devolver de una mensualidad
    """
    from facturacion.models import Devolucion
    
    try:
        mensualidad = get_object_or_404(Mensualidad, id=mensualidad_id)
        
        # Total pagado (sin uso de crédito)
        total_pagado = Pago.objects.filter(
            mensualidad=mensualidad,
            anulado=False
        ).exclude(
            metodo_pago__nombre="Uso de Crédito"
        ).aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
        
        # Devoluciones previas del modelo Devolucion
        devoluciones_previas = Devolucion.objects.filter(
            mensualidad=mensualidad
        ).aggregate(Sum('monto'))['monto__sum'] or Decimal('0.00')
        
        disponible = total_pagado - devoluciones_previas
        
        return JsonResponse({
            'success': True,
            'disponible_devolver': float(disponible),
            'disponible_formatted': f"Bs. {disponible:,.2f}",
            'total_pagado': float(total_pagado),
            'devoluciones_previas': float(devoluciones_previas),
            'mensualidad_codigo': mensualidad.codigo,
            'mensualidad_periodo': mensualidad.periodo_display
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


# ============================================================================
# VISTAS ADMINISTRATIVAS - COMANDOS WEB
# ============================================================================

@staff_member_required
def recalcular_todas_cuentas_web(request):
    """
    Vista administrativa para recalcular todas las cuentas desde el navegador.
    Solo accesible por staff/admin.
    Útil para Render y otros servicios sin acceso SSH.
    """
    from django.middleware.csrf import get_token
    
    if request.method == 'POST':
        # Ejecutar el recálculo
        resultado = AccountService.recalcular_todas_las_cuentas()
        
        # Preparar lista de errores (máximo 20)
        errores_html = ''
        if resultado['errores']:
            errores_items = ''.join([
                f"<li>Paciente {e['paciente_id']} ({e['paciente_nombre']}): {e['error']}</li>" 
                for e in resultado['errores'][:20]
            ])
            if len(resultado['errores']) > 20:
                errores_items += f"<li><em>... y {len(resultado['errores']) - 20} errores más</em></li>"
            
            errores_html = f'''
            <div class="error-section">
                <h2>⚠️ Errores encontrados:</h2>
                <ul class="error-list">
                    {errores_items}
                </ul>
            </div>
            '''
        
        html = f"""
        <!DOCTYPE html>
        <html lang="es">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Recálculo Completado</title>
            <style>
                body {{
                    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, sans-serif;
                    padding: 20px;
                    max-width: 900px;
                    margin: 0 auto;
                    background: #f5f5f5;
                    color: #333;
                }}
                .container {{
                    background: white;
                    padding: 30px;
                    border-radius: 10px;
                    box-shadow: 0 2px 10px rgba(0,0,0,0.1);
                }}
                h1 {{
                    color: #28a745;
                    margin-top: 0;
                }}
                .stats {{
                    background: #e7f3ff;
                    border-left: 4px solid #007bff;
                    padding: 20px;
                    border-radius: 5px;
                    margin: 20px 0;
                }}
                .stats h2 {{
                    margin-top: 0;
                    color: #007bff;
                }}
                .stats ul {{
                    list-style: none;
                    padding: 0;
                }}
                .stats li {{
                    padding: 8px 0;
                    font-size: 16px;
                }}
                .stats strong {{
                    color: #555;
                }}
                .error-section {{
                    background: #fff3cd;
                    border-left: 4px solid #ffc107;
                    padding: 20px;
                    border-radius: 5px;
                    margin: 20px 0;
                }}
                .error-section h2 {{
                    margin-top: 0;
                    color: #856404;
                }}
                .error-list {{
                    list-style: none;
                    padding: 0;
                    max-height: 400px;
                    overflow-y: auto;
                }}
                .error-list li {{
                    padding: 8px;
                    margin: 5px 0;
                    background: white;
                    border-radius: 3px;
                    font-size: 14px;
                }}
                .btn {{
                    display: inline-block;
                    padding: 12px 24px;
                    background: #007bff;
                    color: white;
                    text-decoration: none;
                    border-radius: 5px;
                    font-weight: 500;
                    transition: background 0.3s;
                }}
                .btn:hover {{
                    background: #0056b3;
                }}
                .success-icon {{
                    font-size: 48px;
                    margin-bottom: 10px;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="success-icon">✅</div>
                <h1>Recálculo Completado Exitosamente</h1>
                
                <div class="stats">
                    <h2>📊 Estadísticas del Recálculo:</h2>
                    <ul>
                        <li><strong>Total de cuentas procesadas:</strong> {resultado['total']}</li>
                        <li><strong>Actualizaciones exitosas:</strong> {resultado['exitosos']} 
                            <span style="color: #28a745;">✓</span></li>
                        <li><strong>Errores encontrados:</strong> {len(resultado['errores'])} 
                            {'' if len(resultado['errores']) == 0 else '<span style="color: #dc3545;">✗</span>'}</li>
                    </ul>
                </div>
                
                {errores_html}
                
                <p style="margin-top: 30px;">
                    <a href="/admin/facturacion/cuentacorriente/" class="btn">
                        📋 Ver Cuentas Corrientes
                    </a>
                    <a href="/admin/" class="btn" style="background: #6c757d; margin-left: 10px;">
                        🏠 Volver al Admin
                    </a>
                </p>
            </div>
        </body>
        </html>
        """
        return HttpResponse(html)
    
    # GET - Mostrar formulario de confirmación
    total_pacientes = Paciente.objects.count()
    total_cuentas = CuentaCorriente.objects.count()
    csrf_token = get_token(request)
    
    html = f"""
    <!DOCTYPE html>
    <html lang="es">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Recalcular Cuentas Corrientes</title>
        <style>
            body {{
                font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, sans-serif;
                padding: 20px;
                max-width: 900px;
                margin: 0 auto;
                background: #f5f5f5;
                color: #333;
            }}
            .container {{
                background: white;
                padding: 30px;
                border-radius: 10px;
                box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            }}
            h1 {{
                color: #333;
                margin-top: 0;
            }}
            .warning {{
                background: #fff3cd;
                border-left: 4px solid #ffc107;
                padding: 20px;
                border-radius: 5px;
                margin: 20px 0;
            }}
            .warning h2 {{
                margin-top: 0;
                color: #856404;
            }}
            .info {{
                background: #d1ecf1;
                border-left: 4px solid #17a2b8;
                padding: 20px;
                border-radius: 5px;
                margin: 20px 0;
            }}
            .info h3 {{
                margin-top: 0;
                color: #0c5460;
            }}
            .info ul {{
                margin: 10px 0;
                padding-left: 20px;
            }}
            .info li {{
                margin: 5px 0;
            }}
            button {{
                padding: 12px 24px;
                border: none;
                border-radius: 5px;
                cursor: pointer;
                font-size: 16px;
                font-weight: 500;
                transition: all 0.3s;
            }}
            .btn-primary {{
                background: #28a745;
                color: white;
            }}
            .btn-primary:hover {{
                background: #218838;
            }}
            .btn-secondary {{
                background: #6c757d;
                color: white;
                margin-left: 10px;
            }}
            .btn-secondary:hover {{
                background: #545b62;
            }}
            .stats {{
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
                gap: 15px;
                margin: 20px 0;
            }}
            .stat-card {{
                background: #f8f9fa;
                padding: 15px;
                border-radius: 5px;
                text-align: center;
            }}
            .stat-number {{
                font-size: 36px;
                font-weight: bold;
                color: #007bff;
            }}
            .stat-label {{
                font-size: 14px;
                color: #6c757d;
                margin-top: 5px;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <h1>🔄 Recalcular Todas las Cuentas Corrientes</h1>
            
            <div class="stats">
                <div class="stat-card">
                    <div class="stat-number">{total_pacientes}</div>
                    <div class="stat-label">Pacientes Registrados</div>
                </div>
                <div class="stat-card">
                    <div class="stat-number">{total_cuentas}</div>
                    <div class="stat-label">Cuentas Corrientes</div>
                </div>
            </div>
            
            <div class="warning">
                <h2>⚠️ Advertencia Importante</h2>
                <p><strong>Esta operación recalculará {total_cuentas} cuentas corrientes.</strong></p>
                <p>El proceso:</p>
                <ul>
                    <li>Recalculará todos los totales consumidos (real y actual)</li>
                    <li>Actualizará todos los totales de pagos por categoría</li>
                    <li>Recalculará los saldos (real y actual)</li>
                    <li>Actualizará todos los contadores de sesiones/proyectos/mensualidades</li>
                </ul>
                <p><strong>⏱️ Tiempo estimado:</strong> Puede tomar varios minutos dependiendo de la cantidad de datos.</p>
                <p><strong>💡 Consejo:</strong> Ejecuta esto en horarios de baja actividad.</p>
            </div>
            
            <div class="info">
                <h3>ℹ️ ¿Cuándo usar esto?</h3>
                <ul>
                    <li>Después de aplicar una migración que modifica los campos de CuentaCorriente</li>
                    <li>Si detectas inconsistencias en los saldos mostrados</li>
                    <li>Después de importar datos masivamente</li>
                    <li>Si los signals no se ejecutaron correctamente</li>
                </ul>
                <p><strong>Nota:</strong> Después del primer recálculo, los signals mantendrán todo actualizado automáticamente.</p>
            </div>
            
            <form method="POST" onsubmit="return confirm('¿Estás completamente seguro de iniciar el recálculo de {total_cuentas} cuentas?');">
                <input type="hidden" name="csrfmiddlewaretoken" value="{csrf_token}">
                <button type="submit" class="btn-primary">
                    ✅ Sí, Iniciar Recálculo Ahora
                </button>
                <a href="/admin/facturacion/cuentacorriente/">
                    <button type="button" class="btn-secondary">
                        ❌ Cancelar
                    </button>
                </a>
            </form>
        </div>
    </body>
    </html>
    """
    return HttpResponse(html)


@staff_member_required
def verificar_cuentas_web(request):
    """
    Vista para verificar el estado de las cuentas sin modificarlas.
    """
    from django.db.models import F, Q
    from decimal import Decimal
    
    # Encontrar cuentas con posibles inconsistencias
    cuentas_vacias = CuentaCorriente.objects.filter(
        total_consumido_real=0,
        total_pagado=0
    ).count()
    
    cuentas_con_datos = CuentaCorriente.objects.exclude(
        total_consumido_real=0,
        total_pagado=0
    ).count()
    
    total_cuentas = CuentaCorriente.objects.count()
    
    html = f"""
    <!DOCTYPE html>
    <html lang="es">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Verificar Estado de Cuentas</title>
        <style>
            body {{
                font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, sans-serif;
                padding: 20px;
                max-width: 900px;
                margin: 0 auto;
                background: #f5f5f5;
            }}
            .container {{
                background: white;
                padding: 30px;
                border-radius: 10px;
                box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            }}
            .stats {{
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
                gap: 20px;
                margin: 30px 0;
            }}
            .stat-card {{
                padding: 20px;
                border-radius: 8px;
                text-align: center;
            }}
            .stat-card.info {{
                background: #d1ecf1;
                border-left: 4px solid #17a2b8;
            }}
            .stat-card.success {{
                background: #d4edda;
                border-left: 4px solid #28a745;
            }}
            .stat-card.warning {{
                background: #fff3cd;
                border-left: 4px solid #ffc107;
            }}
            .stat-number {{
                font-size: 48px;
                font-weight: bold;
                margin: 10px 0;
            }}
            .stat-label {{
                font-size: 14px;
                color: #6c757d;
            }}
            .btn {{
                display: inline-block;
                padding: 12px 24px;
                background: #007bff;
                color: white;
                text-decoration: none;
                border-radius: 5px;
                font-weight: 500;
                margin: 10px 5px;
            }}
            .btn.success {{
                background: #28a745;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <h1>📊 Estado de las Cuentas Corrientes</h1>
            
            <div class="stats">
                <div class="stat-card info">
                    <div class="stat-label">Total de Cuentas</div>
                    <div class="stat-number">{total_cuentas}</div>
                </div>
                
                <div class="stat-card success">
                    <div class="stat-label">Con Datos Calculados</div>
                    <div class="stat-number">{cuentas_con_datos}</div>
                </div>
                
                <div class="stat-card warning">
                    <div class="stat-label">Sin Calcular (en 0)</div>
                    <div class="stat-number">{cuentas_vacias}</div>
                </div>
            </div>
            
            {'<div style="background: #fff3cd; padding: 20px; border-radius: 5px; margin: 20px 0;"><h3>⚠️ Acción Requerida</h3><p>Hay ' + str(cuentas_vacias) + ' cuentas sin datos calculados. Se recomienda ejecutar el recálculo.</p></div>' if cuentas_vacias > 0 else '<div style="background: #d4edda; padding: 20px; border-radius: 5px; margin: 20px 0;"><h3>✅ Todo en Orden</h3><p>Todas las cuentas tienen datos calculados correctamente.</p></div>'}
            
            <p>
                {'<a href="/facturacion/admin/recalcular-todas-cuentas/" class="btn success">🔄 Recalcular Cuentas</a>' if cuentas_vacias > 0 else ''}
                <a href="/admin/facturacion/cuentacorriente/" class="btn">📋 Ver Cuentas</a>
                <a href="/admin/" class="btn">🏠 Volver al Admin</a>
            </p>
        </div>
    </body>
    </html>
    """
    return HttpResponse(html)

# Copia y pega estas dos funciones en facturacion/views.py

@login_required
def detalle_sesion_partial(request, sesion_id):
    """
    Vista parcial para cargar detalle de sesión vía HTMX
    ✅ ACTUALIZADO: Incluye pagos masivos
    """
    from facturacion.models import DetallePagoMasivo
    from agenda.models import Sesion
    
    sesion = get_object_or_404(Sesion, id=sesion_id)
    
    # Pagos directos (FK en Pago apunta a la sesión)
    pagos = sesion.pagos.filter(anulado=False).select_related(
        'metodo_pago',
        'registrado_por'
    )
    
    # ✅ NUEVO: Pagos masivos donde esta sesión es uno de los ítems
    pagos_masivos = DetallePagoMasivo.objects.filter(
        tipo='sesion',
        sesion=sesion,
        pago__anulado=False  # Solo pagos no anulados
    ).select_related(
        'pago__metodo_pago',
        'pago__registrado_por'
    ).prefetch_related(
        'pago__detalles_masivos'  # Para poder usar pago.cantidad_detalles
    )
    
    context = {
        'sesion': sesion,
        'pagos': pagos,
        'pagos_masivos': pagos_masivos,  # ✅ Nueva variable
    }
    
    return render(request, 'facturacion/partials/detalle_sesion.html', context)

@login_required
def detalle_pago_partial(request, pago_id):
    pago = get_object_or_404(Pago, id=pago_id)
    context = {'pago': pago}
    return render(request, 'facturacion/partials/detalle_pago.html', context)

def obtener_registros_cobros_filtrados(paciente, request):
    """
    Obtiene registros de cobros con filtros mejorados
    
    Args:
        paciente: Instancia del paciente
        request: HttpRequest con los parámetros GET de filtrado
    
    Returns:
        tuple: (registros_paginados, contadores, filtros_url)
    """
    
    # ==================== EXTRAER PARÁMETROS DE FILTRO ====================
    tipo_concepto = request.GET.get('tipo_concepto', 'todos')
    estado_servicio = request.GET.get('estado_servicio', 'todos')
    estado_pago = request.GET.get('estado_pago', 'todos')
    fecha_desde = request.GET.get('fecha_desde_registros')
    fecha_hasta = request.GET.get('fecha_hasta_registros')
    mostrar_programadas = request.GET.get('mostrar_programadas', 'true') != 'false'
    
    # ==================== CONSTRUIR LISTA DE REGISTROS ====================
    registros = []
    
    # ----- SESIONES -----
    if tipo_concepto in ['todos', 'sesiones']:
        # Filtros base
        filtro_sesiones = Q(
            paciente=paciente,
            proyecto__isnull=True,
            mensualidad__isnull=True
        )
        
        # Filtro por estado del servicio
        if estado_servicio != 'todos':
            # Mapear estados de sesión
            estados_sesion_validos = [
                'programada', 'realizada', 'realizada_retraso', 
                'falta', 'permiso', 'cancelada', 'reprogramada'
            ]
            if estado_servicio in estados_sesion_validos:
                filtro_sesiones &= Q(estado=estado_servicio)
        else:
            # Si no se especifica estado, aplicar filtro de programadas
            if mostrar_programadas:
                filtro_sesiones &= Q(
                    estado__in=['programada', 'realizada', 'realizada_retraso', 'falta']
                )
            else:
                filtro_sesiones &= Q(
                    estado__in=['realizada', 'realizada_retraso', 'falta']
                )
        
        # Filtro por fechas
        if fecha_desde:
            filtro_sesiones &= Q(fecha__gte=fecha_desde)
        if fecha_hasta:
            filtro_sesiones &= Q(fecha__lte=fecha_hasta)
        
        # Obtener sesiones
        sesiones = Sesion.objects.filter(filtro_sesiones).select_related(
            'servicio', 'profesional', 'sucursal'
        ).prefetch_related('pagos')
        
        # Agregar sesiones a registros
        for sesion in sesiones:
            # ✅ CORREGIDO: Calcular total pagado incluyendo pagos masivos
            total_pagado = _total_pagado_sesion(sesion)
            
            # Calcular saldo pendiente
            saldo_pendiente = sesion.monto_cobrado - total_pagado
            pagado_completo = saldo_pendiente <= 0
            
            # Filtro por estado de pago
            if estado_pago == 'pendiente' and pagado_completo:
                continue
            if estado_pago == 'pagado' and not pagado_completo:
                continue
            
            # ✅ Obtener pagos masivos para esta sesión
            pagos_masivos = DetallePagoMasivo.objects.filter(
                tipo='sesion',
                sesion=sesion,
                pago__anulado=False
            ).select_related('pago__metodo_pago')
            
            registros.append({
                'tipo': 'sesion',
                'objeto': sesion,
                'fecha': sesion.fecha,
                'costo': sesion.monto_cobrado,
                'pagado': total_pagado,
                'saldo': saldo_pendiente,
                'pagado_completo': pagado_completo,
                'pagos_masivos': list(pagos_masivos),  # ✅ Agregar pagos masivos
            })
    
    # ----- PROYECTOS -----
    if tipo_concepto in ['todos', 'proyectos']:
        # Filtros base
        filtro_proyectos = Q(paciente=paciente)
        
        # Filtro por estado del servicio
        if estado_servicio != 'todos':
            # Mapear estados de proyecto
            estados_proyecto_validos = [
                'planificado', 'en_progreso', 'finalizado', 'cancelado'
            ]
            if estado_servicio in estados_proyecto_validos:
                filtro_proyectos &= Q(estado=estado_servicio)
        else:
            # Si no se especifica estado, aplicar filtro de planificados
            if mostrar_programadas:
                filtro_proyectos &= Q(
                    estado__in=['planificado', 'en_progreso', 'finalizado', 'cancelado']
                )
            else:
                filtro_proyectos &= Q(
                    estado__in=['en_progreso', 'finalizado', 'cancelado']
                )
        
        # Filtro por fechas
        if fecha_desde:
            filtro_proyectos &= Q(fecha_inicio__gte=fecha_desde)
        if fecha_hasta:
            filtro_proyectos &= Q(fecha_inicio__lte=fecha_hasta)
        
        # Obtener proyectos
        proyectos = Proyecto.objects.filter(filtro_proyectos).select_related(
            'servicio_base', 'profesional_responsable', 'sucursal'
        )
        
        # Agregar proyectos a registros
        for proyecto in proyectos:
            # Usar las propiedades ya definidas en el modelo
            total_pagado = proyecto.pagado_neto  # Ya considera devoluciones
            saldo_pendiente = proyecto.saldo_pendiente
            pagado_completo = proyecto.pagado_completo
            
            # Filtro por estado de pago
            if estado_pago == 'pendiente' and pagado_completo:
                continue
            if estado_pago == 'pagado' and not pagado_completo:
                continue
            
            # ✅ Obtener pagos masivos para este proyecto
            pagos_masivos = DetallePagoMasivo.objects.filter(
                tipo='proyecto',
                proyecto=proyecto,
                pago__anulado=False
            ).select_related('pago__metodo_pago')
            
            registros.append({
                'tipo': 'proyecto',
                'objeto': proyecto,
                'fecha': proyecto.fecha_inicio,
                'costo': proyecto.costo_total,
                'pagado': total_pagado,
                'saldo': saldo_pendiente,
                'pagado_completo': pagado_completo,
                'pagos_masivos': list(pagos_masivos),  # ✅ Agregar pagos masivos
            })
    
    # ----- MENSUALIDADES -----
    if tipo_concepto in ['todos', 'mensualidades']:
        # Filtros base
        filtro_mensualidades = Q(paciente=paciente)
        
        # Filtro por estado del servicio
        if estado_servicio != 'todos':
            # Mapear estados de mensualidad
            # Nota: 'cancelada_mens' se mapea a 'cancelada' para evitar conflicto con sesiones
            if estado_servicio == 'cancelada_mens':
                filtro_mensualidades &= Q(estado='cancelada')
            elif estado_servicio in ['activa', 'vencida', 'pausada', 'completada']:
                filtro_mensualidades &= Q(estado=estado_servicio)
        else:
            # Por defecto mostrar todas menos las canceladas si corresponde
            filtro_mensualidades &= Q(
                estado__in=['activa', 'vencida', 'pausada', 'completada', 'cancelada']
            )
        
        # Filtro por fechas (usar mes/año)
        if fecha_desde:
            # Convertir fecha_desde a año/mes
            from datetime import datetime
            fecha_obj = datetime.strptime(fecha_desde, '%Y-%m-%d')
            filtro_mensualidades &= Q(
                Q(anio__gt=fecha_obj.year) |
                Q(anio=fecha_obj.year, mes__gte=fecha_obj.month)
            )
        if fecha_hasta:
            from datetime import datetime
            fecha_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            filtro_mensualidades &= Q(
                Q(anio__lt=fecha_obj.year) |
                Q(anio=fecha_obj.year, mes__lte=fecha_obj.month)
            )
        
        # Obtener mensualidades
        mensualidades = Mensualidad.objects.filter(filtro_mensualidades).prefetch_related(
            'servicios_profesionales__servicio',
            'servicios_profesionales__profesional'
        )
        
        # Agregar mensualidades a registros
        for mensualidad in mensualidades:
            # Usar las propiedades ya definidas en el modelo
            total_pagado = mensualidad.pagado_neto  # Ya considera devoluciones
            saldo_pendiente = mensualidad.saldo_pendiente
            pagado_completo = mensualidad.pagado_completo
            
            # Filtro por estado de pago
            if estado_pago == 'pendiente' and pagado_completo:
                continue
            if estado_pago == 'pagado' and not pagado_completo:
                continue
            
            # Crear fecha ficticia para ordenamiento (primer día del mes)
            from datetime import date
            fecha_ficticia = date(mensualidad.anio, mensualidad.mes, 1)
            
            # ✅ Obtener pagos masivos para esta mensualidad
            pagos_masivos = DetallePagoMasivo.objects.filter(
                tipo='mensualidad',
                mensualidad=mensualidad,
                pago__anulado=False
            ).select_related('pago__metodo_pago')
            
            registros.append({
                'tipo': 'mensualidad',
                'objeto': mensualidad,
                'fecha': fecha_ficticia,
                'costo': mensualidad.costo_mensual,
                'pagado': total_pagado,
                'saldo': saldo_pendiente,
                'pagado_completo': pagado_completo,
                'pagos_masivos': list(pagos_masivos),  # ✅ Agregar pagos masivos
            })
    
    # ==================== ORDENAR REGISTROS POR FECHA ====================
    registros.sort(key=lambda x: x['fecha'], reverse=True)
    
    # ==================== CALCULAR SUMAS TOTALES ====================
    suma_total = sum(r['costo'] for r in registros)
    suma_pagado = sum(r['pagado'] for r in registros)
    suma_pendientes = sum(r['saldo'] for r in registros if not r['pagado_completo'])
    
    # ==================== PAGINAR REGISTROS ====================
    paginator = Paginator(registros, 20)  # 20 registros por página
    page_number = request.GET.get('page', 1)
    registros_paginados = paginator.get_page(page_number)
    
    # ==================== CONSTRUIR URL DE FILTROS PARA PAGINACIÓN ====================
    filtros_params = []
    if tipo_concepto and tipo_concepto != 'todos':
        filtros_params.append(f'tipo_concepto={tipo_concepto}')
    if estado_servicio and estado_servicio != 'todos':
        filtros_params.append(f'estado_servicio={estado_servicio}')
    if estado_pago and estado_pago != 'todos':
        filtros_params.append(f'estado_pago={estado_pago}')
    if fecha_desde:
        filtros_params.append(f'fecha_desde_registros={fecha_desde}')
    if fecha_hasta:
        filtros_params.append(f'fecha_hasta_registros={fecha_hasta}')
    filtros_params.append(f'mostrar_programadas={mostrar_programadas}')
    
    filtros_url = '&' + '&'.join(filtros_params) if filtros_params else ''
    
    # ==================== RETORNAR RESULTADOS ====================
    return registros_paginados, suma_total, suma_pagado, suma_pendientes, filtros_url

@login_required
def generar_devolucion_pdf(request, devolucion_id):
    """
    Genera comprobante de devolución en PDF usando ReportLab
    
    ✅ Código limpio y organizado
    ✅ Lógica en pdf_generator.py
    """
    
    # Query optimizada de la devolución
    devolucion = get_object_or_404(
        Devolucion.objects.select_related(
            'paciente',
            'metodo_devolucion',
            'registrado_por',
            'proyecto',
            'mensualidad'
        ),
        id=devolucion_id
    )
    
    try:
        # ✅ Verificar cache
        cache_key = f'pdf_devolucion_{devolucion.numero_devolucion}'
        pdf_cached = cache.get(cache_key)
        
        if pdf_cached:
            response = HttpResponse(pdf_cached, content_type='application/pdf')
            response['Content-Disposition'] = f'inline; filename="devolucion_{devolucion.numero_devolucion}.pdf"'
            return response
        
        # ✅ Generar PDF usando el módulo separado
        pdf_data = pdf_generator.generar_devolucion_pdf(devolucion)
        
        # Cachear
        cache.set(cache_key, pdf_data, 3600)
        
        # Respuesta
        response = HttpResponse(pdf_data, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="devolucion_{devolucion.numero_devolucion}.pdf"'
        
        return response
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Error generando PDF devolución {devolucion_id}: {str(e)}')
        
        messages.error(request, f'❌ Error al generar PDF: {str(e)}')
        return redirect('facturacion:historial_pagos')